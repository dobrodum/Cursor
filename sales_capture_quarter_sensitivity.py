#!/usr/bin/env python3
"""Compare % Sales Captured in DB and estimate total sold sensitivity.

Given quarter-level sales data, this script computes:

    % Sales Captured in DB = db_sales / total_sales * 100

for rolling windows of 1, 2, 3, and 4 quarters (or custom windows), and then
estimates "total sold" for a specific quarter (for example FY 4Q25):

    Estimated Total Sold = db_sales_in_estimate_quarter / (capture_pct / 100)
"""

from __future__ import annotations

import argparse
import csv
import math
import re
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Sequence, Tuple

try:
    from openpyxl import load_workbook
except ImportError:
    load_workbook = None


@dataclass(frozen=True, order=True)
class Quarter:
    year: int
    quarter: int

    def __str__(self) -> str:
        return f"{self.year}Q{self.quarter}"


@dataclass
class WindowResult:
    window_quarters: int
    start_quarter: Quarter
    end_quarter: Quarter
    db_sales_window: float
    total_sales_window: float
    pct_captured: float
    delta_vs_baseline_pp: float
    db_sales_estimate_quarter: float
    estimated_total_sold: float
    delta_estimate_vs_baseline: float
    delta_estimate_vs_baseline_pct: float


def safe_float(raw: Any) -> Optional[float]:
    if raw is None:
        return None
    if isinstance(raw, (int, float)):
        value = float(raw)
        if math.isnan(value) or math.isinf(value):
            return None
        return value
    text = str(raw).strip()
    if text == "":
        return None
    negative = text.startswith("(") and text.endswith(")")
    cleaned = text.replace(",", "").replace("$", "").replace("%", "")
    if negative:
        cleaned = cleaned[1:-1]
    try:
        value = float(cleaned)
    except ValueError:
        return None
    if negative:
        value = -value
    if math.isnan(value) or math.isinf(value):
        return None
    return value


def normalize_year(year_text: str) -> Optional[int]:
    if not year_text.isdigit():
        return None
    if len(year_text) == 4:
        return int(year_text)
    if len(year_text) == 2:
        # Interprets 25 as 2025, which matches common fiscal shorthand.
        return 2000 + int(year_text)
    return None


def parse_quarter(raw: Any) -> Optional[Quarter]:
    if raw is None:
        return None
    if isinstance(raw, datetime):
        qtr = ((raw.month - 1) // 3) + 1
        return Quarter(year=raw.year, quarter=qtr)

    text = str(raw).strip()
    if text == "":
        return None

    normalized = text.upper().replace(" ", "").replace("FY", "")
    quarter_patterns = (
        r"^(?P<year>\d{4})[-_/]?Q(?P<q>[1-4])$",  # 2025Q4, 2025-Q4
        r"^Q(?P<q>[1-4])[-_/]?(?P<year>\d{4})$",  # Q4-2025
        r"^(?P<q>[1-4])Q(?P<year>\d{2,4})$",      # 4Q25, 4Q2025
        r"^(?P<year>\d{2,4})Q(?P<q>[1-4])$",      # 25Q4, 2025Q4
    )
    for pattern in quarter_patterns:
        match = re.fullmatch(pattern, normalized)
        if not match:
            continue
        year = normalize_year(match.group("year"))
        if year is None:
            continue
        qtr = int(match.group("q"))
        return Quarter(year=year, quarter=qtr)

    # Supported date strings: 2024-03-31, 2024/03/31, etc.
    date_formats = ("%Y-%m-%d", "%Y/%m/%d", "%m/%d/%Y", "%d/%m/%Y")
    for fmt in date_formats:
        try:
            dt = datetime.strptime(str(raw), fmt)
            qtr = ((dt.month - 1) // 3) + 1
            return Quarter(year=dt.year, quarter=qtr)
        except ValueError:
            continue

    return None


def read_csv_rows(path: str) -> List[Dict[str, Any]]:
    with open(path, "r", encoding="utf-8", newline="") as handle:
        reader = csv.DictReader(handle)
        if not reader.fieldnames:
            raise ValueError("CSV has no headers.")
        return list(reader)


def read_excel_rows(path: str, sheet_name: str) -> List[Dict[str, Any]]:
    if load_workbook is None:
        raise ValueError(
            "Excel support requires openpyxl. Install it with: pip3 install openpyxl"
        )

    workbook = load_workbook(path, data_only=True, read_only=True)
    if sheet_name:
        if sheet_name not in workbook.sheetnames:
            raise ValueError(
                f"Sheet '{sheet_name}' not found. Available sheets: {', '.join(workbook.sheetnames)}"
            )
        sheet = workbook[sheet_name]
    else:
        sheet = workbook.active

    row_iter = sheet.iter_rows(values_only=True)
    header_row = next(row_iter, None)
    if header_row is None:
        raise ValueError("Excel sheet has no rows.")

    headers: List[str] = []
    indexes: List[int] = []
    for idx, value in enumerate(header_row):
        if value is None:
            continue
        name = str(value).strip()
        if not name:
            continue
        headers.append(name)
        indexes.append(idx)

    if not headers:
        raise ValueError("Excel sheet has no usable header row.")

    rows: List[Dict[str, Any]] = []
    for raw_row in row_iter:
        record: Dict[str, Any] = {}
        has_values = False
        for header, idx in zip(headers, indexes):
            value = raw_row[idx] if idx < len(raw_row) else None
            if value not in (None, ""):
                has_values = True
            record[header] = value
        if has_values:
            rows.append(record)
    return rows


def read_rows(path: str, sheet_name: str) -> List[Dict[str, Any]]:
    extension = Path(path).suffix.lower()
    if extension in {".csv", ".txt"}:
        return read_csv_rows(path)
    if extension in {".xlsx", ".xlsm"}:
        return read_excel_rows(path, sheet_name=sheet_name)
    if extension == ".xls":
        raise ValueError("Legacy .xls is not supported. Convert to .xlsx or .csv.")
    raise ValueError(f"Unsupported file extension '{extension}'. Use csv or xlsx.")


def aggregate_by_quarter(
    rows: Sequence[Dict[str, Any]],
    quarter_col: str,
    db_sales_col: str,
    total_sales_col: str,
) -> Dict[Quarter, Tuple[float, float]]:
    aggregates: Dict[Quarter, Tuple[float, float]] = {}

    for idx, row in enumerate(rows, start=2):
        quarter = parse_quarter(row.get(quarter_col, ""))
        db_sales = safe_float(row.get(db_sales_col, ""))
        total_sales = safe_float(row.get(total_sales_col, ""))

        if quarter is None:
            raise ValueError(
                f"Row {idx}: quarter value '{row.get(quarter_col, '')}' "
                f"could not be parsed. Use format like 2024Q1 or 2024-03-31."
            )
        if db_sales is None or total_sales is None:
            raise ValueError(
                f"Row {idx}: db_sales or total_sales is missing/non-numeric "
                f"('{row.get(db_sales_col, '')}', '{row.get(total_sales_col, '')}')."
            )

        existing_db, existing_total = aggregates.get(quarter, (0.0, 0.0))
        aggregates[quarter] = (existing_db + db_sales, existing_total + total_sales)

    return aggregates


def rolling_capture(
    ordered_quarters: Sequence[Quarter],
    aggregates: Dict[Quarter, Tuple[float, float]],
    window_size: int,
    end_quarter: Optional[Quarter] = None,
) -> Tuple[Quarter, Quarter, float, float, float]:
    if window_size < 1:
        raise ValueError("window_size must be at least 1.")

    if end_quarter is None:
        end_idx = len(ordered_quarters) - 1
    else:
        try:
            end_idx = ordered_quarters.index(end_quarter)
        except ValueError as err:
            raise ValueError(
                f"end_quarter '{end_quarter}' is not present in your data."
            ) from err

    if end_idx < 0:
        raise ValueError("No quarter data available.")

    start_idx = max(0, end_idx - window_size + 1)
    chosen = ordered_quarters[start_idx : end_idx + 1]
    if len(chosen) < window_size:
        raise ValueError(
            f"Not enough history for window={window_size}. "
            f"Only {len(chosen)} quarter(s) available up to {ordered_quarters[end_idx]}."
        )

    db_total = 0.0
    sales_total = 0.0
    for qtr in chosen:
        db_value, total_value = aggregates[qtr]
        db_total += db_value
        sales_total += total_value

    if sales_total == 0:
        raise ValueError(
            f"Total sales is 0 for window={window_size} ending {ordered_quarters[end_idx]}."
        )

    pct = (db_total / sales_total) * 100.0
    return chosen[0], chosen[-1], db_total, sales_total, pct


def render_table(results: Sequence[WindowResult]) -> str:
    headers = [
        "window_quarters",
        "start_quarter",
        "end_quarter",
        "db_sales_window",
        "total_sales_window",
        "%_sales_captured_in_db",
        "delta_vs_baseline_pp",
        "db_sales_in_estimate_qtr",
        "estimated_total_sold",
        "delta_estimate_vs_baseline",
        "delta_estimate_vs_baseline_pct",
    ]
    rows: List[List[str]] = []
    for item in results:
        rows.append(
            [
                str(item.window_quarters),
                str(item.start_quarter),
                str(item.end_quarter),
                f"{item.db_sales_window:.2f}",
                f"{item.total_sales_window:.2f}",
                f"{item.pct_captured:.4f}",
                f"{item.delta_vs_baseline_pp:+.4f}",
                f"{item.db_sales_estimate_quarter:.2f}",
                f"{item.estimated_total_sold:.2f}",
                f"{item.delta_estimate_vs_baseline:+.2f}",
                (
                    "NA"
                    if math.isnan(item.delta_estimate_vs_baseline_pct)
                    else f"{item.delta_estimate_vs_baseline_pct:+.4f}%"
                ),
            ]
        )

    widths = [len(h) for h in headers]
    for row in rows:
        for idx, cell in enumerate(row):
            widths[idx] = max(widths[idx], len(cell))

    def fmt(values: Iterable[str]) -> str:
        return " | ".join(v.ljust(widths[i]) for i, v in enumerate(values))

    output = [fmt(headers), "-+-".join("-" * w for w in widths)]
    output.extend(fmt(row) for row in rows)
    return "\n".join(output)


def parse_windows(raw: str) -> List[int]:
    values: List[int] = []
    for item in raw.split(","):
        text = item.strip()
        if not text:
            continue
        if not text.isdigit():
            raise ValueError(f"Window '{item}' is not an integer.")
        number = int(text)
        if number < 1:
            raise ValueError(f"Window '{item}' must be >= 1.")
        values.append(number)
    if not values:
        raise ValueError("No window values provided.")
    return sorted(set(values))


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description=(
            "Compare % Sales Captured in DB for quarter windows "
            "(e.g., 1,2,3,4 quarters)."
        )
    )
    parser.add_argument("--data", required=True, help="Path to CSV data.")
    parser.add_argument(
        "--sheet",
        default="",
        help="Excel sheet name (only used for .xlsx/.xlsm). Default: active sheet.",
    )
    parser.add_argument(
        "--quarter-col",
        "--date-col",
        default="quarter",
        dest="quarter_col",
        help="Column holding quarter or date values (default: quarter).",
    )
    parser.add_argument(
        "--db-sales-col",
        "--captured-sales-col",
        default="db_sales",
        dest="db_sales_col",
        help="Column holding sales captured in DB (default: db_sales).",
    )
    parser.add_argument(
        "--total-sales-col",
        default="total_sales",
        help="Column holding total sales (default: total_sales).",
    )
    parser.add_argument(
        "--windows",
        "--quarters",
        default="1,2,3,4",
        dest="windows",
        help="Comma-separated quarter windows to compare (default: 1,2,3,4).",
    )
    parser.add_argument(
        "--end-quarter",
        default="",
        help=(
            "Optional analysis endpoint quarter (example: 2025Q4). "
            "If omitted, latest available quarter is used."
        ),
    )
    parser.add_argument(
        "--estimate-quarter",
        default="",
        help=(
            "Quarter for Estimated Total Sold output (example: FY4Q25 or 4Q25). "
            "Defaults to --end-quarter (or latest quarter if --end-quarter omitted)."
        ),
    )
    parser.add_argument(
        "--baseline-quarter",
        type=int,
        default=4,
        help="Baseline window used for deltas (default: 4).",
    )
    parser.add_argument(
        "--output",
        default="",
        help="Optional path to write results as CSV.",
    )
    return parser


def write_results_csv(path: str, results: Sequence[WindowResult]) -> None:
    headers = [
        "window_quarters",
        "start_quarter",
        "end_quarter",
        "db_sales_window",
        "total_sales_window",
        "pct_sales_captured_in_db",
        "delta_vs_baseline_pp",
        "db_sales_in_estimate_qtr",
        "estimated_total_sold",
        "delta_estimate_vs_baseline",
        "delta_estimate_vs_baseline_pct",
    ]
    with open(path, "w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=headers)
        writer.writeheader()
        for row in results:
            writer.writerow(
                {
                    "window_quarters": row.window_quarters,
                    "start_quarter": str(row.start_quarter),
                    "end_quarter": str(row.end_quarter),
                    "db_sales_window": f"{row.db_sales_window:.6f}",
                    "total_sales_window": f"{row.total_sales_window:.6f}",
                    "pct_sales_captured_in_db": f"{row.pct_captured:.6f}",
                    "delta_vs_baseline_pp": f"{row.delta_vs_baseline_pp:.6f}",
                    "db_sales_in_estimate_qtr": f"{row.db_sales_estimate_quarter:.6f}",
                    "estimated_total_sold": f"{row.estimated_total_sold:.6f}",
                    "delta_estimate_vs_baseline": f"{row.delta_estimate_vs_baseline:.6f}",
                    "delta_estimate_vs_baseline_pct": (
                        ""
                        if math.isnan(row.delta_estimate_vs_baseline_pct)
                        else f"{row.delta_estimate_vs_baseline_pct:.6f}"
                    ),
                }
            )


def main() -> None:
    args = build_parser().parse_args()
    if args.baseline_quarter < 1:
        raise ValueError("--baseline-quarter must be >= 1.")

    windows = parse_windows(args.windows)
    rows = read_rows(args.data, sheet_name=args.sheet)
    aggregates = aggregate_by_quarter(
        rows=rows,
        quarter_col=args.quarter_col,
        db_sales_col=args.db_sales_col,
        total_sales_col=args.total_sales_col,
    )
    ordered_quarters = sorted(aggregates.keys())
    if not ordered_quarters:
        raise ValueError("No quarter data found.")

    end_quarter = None
    if args.end_quarter.strip():
        end_quarter = parse_quarter(args.end_quarter.strip())
        if end_quarter is None:
            raise ValueError(
                f"Could not parse --end-quarter value '{args.end_quarter}'."
            )

    if args.baseline_quarter not in windows:
        windows = sorted(set(windows + [args.baseline_quarter]))

    # Default estimate quarter to analysis endpoint, or latest quarter if no endpoint given.
    estimate_quarter = end_quarter if end_quarter is not None else ordered_quarters[-1]
    if args.estimate_quarter.strip():
        estimate_quarter = parse_quarter(args.estimate_quarter.strip())
        if estimate_quarter is None:
            raise ValueError(
                f"Could not parse --estimate-quarter value '{args.estimate_quarter}'."
            )
    if estimate_quarter not in aggregates:
        raise ValueError(
            f"estimate quarter '{estimate_quarter}' is not present in your data."
        )

    interim: List[Tuple[int, Quarter, Quarter, float, float, float]] = []
    for window_size in windows:
        start, end, db_sales, total_sales, pct = rolling_capture(
            ordered_quarters=ordered_quarters,
            aggregates=aggregates,
            window_size=window_size,
            end_quarter=end_quarter,
        )
        interim.append((window_size, start, end, db_sales, total_sales, pct))

    baseline_pct = next(
        (pct for win, *_rest, pct in interim if win == args.baseline_quarter), None
    )
    if baseline_pct is None:
        raise ValueError("Unable to compute baseline percentage.")

    estimate_db_sales = aggregates[estimate_quarter][0]
    if baseline_pct <= 0:
        raise ValueError("Baseline capture % is <= 0, cannot estimate total sold.")
    baseline_estimate_total_sold = estimate_db_sales / (baseline_pct / 100.0)

    results: List[WindowResult] = []
    for window_size, start, end, db_sales, total_sales, pct in interim:
        if pct <= 0:
            raise ValueError(
                f"Window {window_size}: capture % is <= 0, cannot estimate total sold."
            )
        estimate_total_sold = estimate_db_sales / (pct / 100.0)
        delta_abs = estimate_total_sold - baseline_estimate_total_sold
        if baseline_estimate_total_sold == 0:
            delta_pct = math.nan
        else:
            delta_pct = (delta_abs / baseline_estimate_total_sold) * 100.0

        results.append(
            WindowResult(
                window_quarters=window_size,
                start_quarter=start,
                end_quarter=end,
                db_sales_window=db_sales,
                total_sales_window=total_sales,
                pct_captured=pct,
                delta_vs_baseline_pp=pct - baseline_pct,
                db_sales_estimate_quarter=estimate_db_sales,
                estimated_total_sold=estimate_total_sold,
                delta_estimate_vs_baseline=delta_abs,
                delta_estimate_vs_baseline_pct=delta_pct,
            )
        )

    results = sorted(results, key=lambda item: item.window_quarters)
    print(render_table(results))
    print()
    print(
        f"Estimate quarter: {estimate_quarter} | "
        f"DB sales in quarter: {estimate_db_sales:.2f}"
    )
    print(
        f"Baseline window: {args.baseline_quarter} quarter(s) | "
        f"Baseline estimated total sold: {baseline_estimate_total_sold:.2f}"
    )

    min_result = min(results, key=lambda item: item.estimated_total_sold)
    max_result = max(results, key=lambda item: item.estimated_total_sold)
    print(
        f"Min estimated total sold: {min_result.estimated_total_sold:.2f} "
        f"(window={min_result.window_quarters})"
    )
    print(
        f"Max estimated total sold: {max_result.estimated_total_sold:.2f} "
        f"(window={max_result.window_quarters})"
    )

    if args.output.strip():
        write_results_csv(args.output.strip(), results)
        print(f"Saved result table to {args.output.strip()}")


if __name__ == "__main__":
    main()
