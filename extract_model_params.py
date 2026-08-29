#!/usr/bin/env python3
"""Extract empirical/regression model candidates from a folder of .xlsx files.

Runtime-focused behavior:
- Uses one hidden Excel app for the full run.
- Opens each source workbook once, processes both target sheets, then closes it.
- Uses R1C1 `.formula2` writes only when needed for calculations.
"""

from __future__ import annotations

import datetime as dt
import re
from pathlib import Path
from typing import Any, Dict, List, Optional, Sequence, Tuple

import xlwings as xw
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

# -----------------------------
# User-configurable paths
# -----------------------------
input_dir = Path("./input")
output_dir = Path("./output")


EMPIRICAL_COLUMNS = [
    "model",
    "ticker",
    "model_period",
    "model_date",
    "method",
    "parameter_name",
    "parameter_value",
    "num_quarters_used",
    "last_quarter_used",
    "forecast_value",
    "actual_value",
    "forecast_max",
    "forecast_min",
    "range_width",
    "avg_penetration_pct",
    "quarterly_sales",
    "reported_sales",
    "growth_rate_pct",
    "sales_captured_in_db_pct",
    "source_file",
]

REGRESSION_COLUMNS = [
    "model",
    "ticker",
    "model_period",
    "model_date",
    "method",
    "parameter_name",
    "parameter_value",
    "num_quarters_used",
    "forecast_value",
    "actual_value",
    "forecast_max",
    "forecast_min",
    "range_width",
    "intercept",
    "slope",
    "source_file",
]

PERIOD_DAY = {"early": 5, "mid": 15, "late": 25}


def is_blank(value: Any) -> bool:
    return value is None or (isinstance(value, str) and value.strip() == "")


def normalize_text(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip().lower()
    text = re.sub(r"[%/\\-]+", " ", text)
    text = re.sub(r"[^a-z0-9 ]+", "", text)
    text = re.sub(r"\s+", " ", text).strip()
    return text


def to_float(value: Any) -> Optional[float]:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        text = value.strip().replace(",", "")
        if text.endswith("%"):
            text = text[:-1]
        if text == "":
            return None
        try:
            return float(text)
        except ValueError:
            return None
    return None


def to_int(value: Any, default: Optional[int] = None) -> Optional[int]:
    num = to_float(value)
    if num is None:
        return default
    return int(round(num))


def clamp_col(col: int) -> int:
    return max(1, col)


def set_formula2(cell: xw.Range, formula: str) -> None:
    try:
        cell.formula2 = formula
    except Exception:
        # Conservative fallback for Excel variants that expose only `.formula`.
        cell.formula = formula


def get_output_path(src_input_dir: Path, dst_output_dir: Path) -> Path:
    dst_output_dir.mkdir(parents=True, exist_ok=True)
    base = f"{src_input_dir.name}_PARAM"
    first = dst_output_dir / f"{base}.xlsx"
    if not first.exists():
        return first

    index = 1
    while True:
        candidate = dst_output_dir / f"{base}.{index}.xlsx"
        if not candidate.exists():
            return candidate
        index += 1


def parse_file_labels(file_name: str) -> Dict[str, str]:
    stem = Path(file_name).stem
    # Typical: MedMiner_Model - AORT - MidJan2026_Send
    parts = [part.strip() for part in stem.split(" - ")]
    ticker = ""
    period_token = ""

    if len(parts) >= 3:
        ticker = parts[-2]
        period_token = re.sub(r"_send$", "", parts[-1], flags=re.IGNORECASE)
    else:
        ticker = stem
        period_token = stem

    match = re.match(r"(?i)^(early|mid|late)([a-z]+)(\d{4})$", period_token)
    if not match:
        model_period = period_token
        model_date = ""
        model = f"{ticker}_{model_period}" if ticker else model_period
        return {
            "ticker": ticker,
            "model_period": model_period,
            "model_date": model_date,
            "model": model,
        }

    period_part = match.group(1).lower()
    month_part = match.group(2)
    year = int(match.group(3))

    month_abbrev = month_part[:3].title()
    try:
        month_num = dt.datetime.strptime(month_abbrev, "%b").month
        day = PERIOD_DAY[period_part]
        model_date = dt.date(year, month_num, day).isoformat()
    except ValueError:
        model_date = ""

    model_period = f"{period_part.title()}{month_abbrev}_{year}"
    model = f"{ticker}_{model_period}" if ticker else model_period

    return {
        "ticker": ticker,
        "model_period": model_period,
        "model_date": model_date,
        "model": model,
    }


def get_sheet_case_insensitive(wb: xw.Book, sheet_name: str) -> Optional[xw.Sheet]:
    target = sheet_name.strip().lower()
    for sh in wb.sheets:
        if sh.name.strip().lower() == target:
            return sh
    return None


def safe_close_workbook(wb: xw.Book) -> None:
    try:
        wb.close(save=False)
        return
    except TypeError:
        pass
    except Exception:
        pass

    try:
        wb.api.Close(SaveChanges=False)
        return
    except Exception:
        pass

    try:
        wb.close()
    except Exception:
        pass


def used_range_matrix(sheet: xw.Sheet) -> Tuple[int, int, List[List[Any]]]:
    used = sheet.used_range
    start_row = used.row
    start_col = used.column
    values = used.value

    if values is None:
        return start_row, start_col, []
    if not isinstance(values, list):
        return start_row, start_col, [[values]]
    if values and not isinstance(values[0], list):
        return start_row, start_col, [values]
    return start_row, start_col, values


def find_anchor_max(sheet: xw.Sheet) -> Optional[Tuple[int, int]]:
    start_row, start_col, grid = used_range_matrix(sheet)
    if not grid:
        return None

    norm_max = "max"
    candidates: List[Tuple[int, int]] = []
    for r_idx, row_vals in enumerate(grid):
        for c_idx, value in enumerate(row_vals):
            if normalize_text(value) == norm_max:
                candidates.append((start_row + r_idx, start_col + c_idx))

    if not candidates:
        return None

    # Prefer a max label that appears next to a min label.
    for row, col in candidates:
        right = normalize_text(sheet.cells(row, col + 1).value)
        left = normalize_text(sheet.cells(row, col - 1).value)
        if right == "min" or left == "min":
            return row, col
    return candidates[0]


def get_header_map(sheet: xw.Sheet, header_row: int, min_col: int, max_col: int) -> Dict[str, int]:
    values = sheet.range((header_row, min_col), (header_row, max_col)).value
    if not isinstance(values, list):
        values = [values]
    mapping: Dict[str, int] = {}
    for idx, val in enumerate(values):
        norm = normalize_text(val)
        if norm and norm not in mapping:
            mapping[norm] = min_col + idx
    return mapping


def resolve_column(
    header_map: Dict[str, int],
    aliases: Sequence[str],
    default: Optional[int] = None,
) -> Optional[int]:
    alias_norms = [normalize_text(alias) for alias in aliases]

    for alias in alias_norms:
        if alias in header_map:
            return header_map[alias]

    for key, col in header_map.items():
        for alias in alias_norms:
            alias_tokens = alias.split()
            if alias_tokens and all(token in key for token in alias_tokens):
                return col

    return default


def build_empirical_rows(
    sheet: xw.Sheet,
    labels: Dict[str, str],
    source_file: str,
) -> List[Dict[str, Any]]:
    anchor = find_anchor_max(sheet)
    if anchor is None:
        print(f"  skipped empirical in {source_file}: max anchor not found")
        return []

    anchor_row, anchor_col = anchor
    header_map = get_header_map(sheet, anchor_row, max(1, anchor_col - 24), anchor_col + 8)

    num_col = resolve_column(
        header_map,
        ["num quarters used", "num quarters", "quarters used"],
        clamp_col(anchor_col - 8),
    )
    last_quarter_col = resolve_column(
        header_map,
        ["last quarter used", "last qtr used", "last quarter"],
        clamp_col(anchor_col - 7),
    )
    forecast_col = resolve_column(
        header_map,
        ["estimated total sold", "forecast value", "forecast", "total sold"],
        clamp_col(anchor_col - 6),
    )
    actual_col = resolve_column(
        header_map,
        ["reported sales", "actual sales", "actual value"],
        clamp_col(anchor_col - 5),
    )
    quarterly_sales_col = resolve_column(
        header_map,
        ["quarterly sales"],
        clamp_col(anchor_col - 4),
    )
    growth_rate_col = resolve_column(
        header_map,
        ["growth rate pct", "growth rate"],
        clamp_col(anchor_col - 3),
    )
    captured_col = resolve_column(
        header_map,
        ["sales captured in db pct", "sales captured in db", "captured in db pct"],
        clamp_col(anchor_col - 2),
    )
    avg_pen_col = resolve_column(
        header_map,
        ["avg penetration pct", "avg penetration", "average penetration"],
        None,
    )
    max_col = anchor_col
    min_col = resolve_column(header_map, ["min"], clamp_col(anchor_col + 1))

    n_quarters = 10
    start_row = anchor_row + 1
    data_rows = [start_row + i for i in range(n_quarters)]

    # If avg penetration is not directly available, create temporary R1C1 formulas.
    temp_avg_col = anchor_col + 4
    pending_formula_rows: List[int] = []
    for idx, row in enumerate(data_rows, start=1):
        if avg_pen_col is not None or captured_col is None:
            continue
        quarters_used = to_int(sheet.cells(row, num_col).value, default=idx)
        if quarters_used is None or quarters_used <= 0:
            continue
        start_col = max(1, captured_col - quarters_used + 1)
        start_offset = start_col - temp_avg_col
        end_offset = captured_col - temp_avg_col
        formula = f'=IFERROR(AVERAGE(RC[{start_offset}]:RC[{end_offset}]),"")'
        set_formula2(sheet.cells(row, temp_avg_col), formula)
        pending_formula_rows.append(row)

    if pending_formula_rows:
        sheet.book.app.calculate()

    temp_avg_values: Dict[int, Any] = {}
    for row in pending_formula_rows:
        temp_avg_values[row] = sheet.cells(row, temp_avg_col).value
        sheet.cells(row, temp_avg_col).clear_contents()

    rows: List[Dict[str, Any]] = []
    for idx, row in enumerate(data_rows, start=1):
        forecast_max = to_float(sheet.cells(row, max_col).value)
        forecast_min = to_float(sheet.cells(row, min_col).value)
        forecast_value = to_float(sheet.cells(row, forecast_col).value) if forecast_col else None
        actual_value = to_float(sheet.cells(row, actual_col).value) if actual_col else None
        num_quarters_used = to_int(sheet.cells(row, num_col).value, default=idx) if num_col else idx
        last_quarter_used = sheet.cells(row, last_quarter_col).value if last_quarter_col else None
        quarterly_sales = to_float(sheet.cells(row, quarterly_sales_col).value) if quarterly_sales_col else None
        reported_sales = to_float(sheet.cells(row, actual_col).value) if actual_col else None
        growth_rate = to_float(sheet.cells(row, growth_rate_col).value) if growth_rate_col else None
        captured_pct = to_float(sheet.cells(row, captured_col).value) if captured_col else None
        avg_penetration = (
            to_float(sheet.cells(row, avg_pen_col).value)
            if avg_pen_col
            else to_float(temp_avg_values.get(row))
        )

        key_values = [forecast_value, actual_value, forecast_max, forecast_min, avg_penetration]
        if all(value is None for value in key_values):
            continue

        range_width = (
            (forecast_max - forecast_min)
            if forecast_max is not None and forecast_min is not None
            else None
        )

        rows.append(
            {
                "model": labels["model"],
                "ticker": labels["ticker"],
                "model_period": labels["model_period"],
                "model_date": labels["model_date"],
                "method": "empirical",
                "parameter_name": "avg_penetration_pct",
                "parameter_value": avg_penetration,
                "num_quarters_used": num_quarters_used,
                "last_quarter_used": last_quarter_used,
                "forecast_value": forecast_value,
                "actual_value": actual_value,
                "forecast_max": forecast_max,
                "forecast_min": forecast_min,
                "range_width": range_width,
                "avg_penetration_pct": avg_penetration,
                "quarterly_sales": quarterly_sales,
                "reported_sales": reported_sales,
                "growth_rate_pct": growth_rate,
                "sales_captured_in_db_pct": captured_pct,
                "source_file": source_file,
            }
        )

    return rows


def find_last_numeric_row(sheet: xw.Sheet, col: int, from_row: int) -> Optional[int]:
    for row in range(from_row, 0, -1):
        if to_float(sheet.cells(row, col).value) is not None:
            return row
    return None


def build_regression_rows(
    sheet: xw.Sheet,
    labels: Dict[str, str],
    source_file: str,
) -> List[Dict[str, Any]]:
    anchor = find_anchor_max(sheet)
    if anchor is None:
        print(f"  skipped regression in {source_file}: max anchor not found")
        return []

    anchor_row, anchor_col = anchor
    y_col = anchor_col - 7
    x_col = anchor_col - 11
    y_col = clamp_col(y_col)
    x_col = clamp_col(x_col)

    header_map = get_header_map(sheet, anchor_row, max(1, anchor_col - 24), anchor_col + 8)
    num_col = resolve_column(
        header_map,
        ["num quarters used", "num quarters", "quarters used"],
        clamp_col(anchor_col - 8),
    )
    forecast_col = resolve_column(
        header_map,
        ["tot fcst w o sa", "tot fcst wo sa", "tot fcst without sa", "forecast value"],
        clamp_col(anchor_col - 2),
    )
    actual_col = resolve_column(header_map, ["actual value", "actual sales"], None)
    max_col = anchor_col
    min_col = resolve_column(header_map, ["min"], clamp_col(anchor_col + 1))

    last_data_row = find_last_numeric_row(sheet, y_col, anchor_row - 1)
    if last_data_row is None:
        print(f"  skipped regression in {source_file}: no numeric Y data found")
        return []

    n_quarters = 10
    start_row = anchor_row + 1
    data_rows = [start_row + i for i in range(n_quarters)]

    temp_intercept_col = anchor_col + 4
    temp_slope_col = anchor_col + 5
    rows_with_formulas: List[Tuple[int, int]] = []

    for idx, row in enumerate(data_rows, start=1):
        q_used = to_int(sheet.cells(row, num_col).value, default=idx) if num_col else idx
        if q_used is None or q_used <= 1:
            continue
        first_row = last_data_row - q_used + 1
        if first_row < 1:
            continue

        intercept_formula = (
            f'=IFERROR(INTERCEPT(R{first_row}C{y_col}:R{last_data_row}C{y_col},'
            f'R{first_row}C{x_col}:R{last_data_row}C{x_col}),"")'
        )
        slope_formula = (
            f'=IFERROR(SLOPE(R{first_row}C{y_col}:R{last_data_row}C{y_col},'
            f'R{first_row}C{x_col}:R{last_data_row}C{x_col}),"")'
        )
        set_formula2(sheet.cells(row, temp_intercept_col), intercept_formula)
        set_formula2(sheet.cells(row, temp_slope_col), slope_formula)
        rows_with_formulas.append((row, q_used))

    if rows_with_formulas:
        sheet.book.app.calculate()

    intercept_values: Dict[int, Any] = {}
    slope_values: Dict[int, Any] = {}
    for row, _ in rows_with_formulas:
        intercept_values[row] = sheet.cells(row, temp_intercept_col).value
        slope_values[row] = sheet.cells(row, temp_slope_col).value
        sheet.cells(row, temp_intercept_col).clear_contents()
        sheet.cells(row, temp_slope_col).clear_contents()

    rows: List[Dict[str, Any]] = []
    prev_signature: Optional[Tuple[Any, ...]] = None

    for idx, row in enumerate(data_rows, start=1):
        num_quarters_used = to_int(sheet.cells(row, num_col).value, default=idx) if num_col else idx
        forecast_value = to_float(sheet.cells(row, forecast_col).value) if forecast_col else None
        actual_value = to_float(sheet.cells(row, actual_col).value) if actual_col else None
        forecast_max = to_float(sheet.cells(row, max_col).value)
        forecast_min = to_float(sheet.cells(row, min_col).value) if min_col else None
        intercept = to_float(intercept_values.get(row))
        slope = to_float(slope_values.get(row))

        key_values = [forecast_value, forecast_max, forecast_min, intercept, slope]
        if all(value is None for value in key_values):
            continue

        signature = (
            num_quarters_used,
            forecast_value,
            forecast_max,
            forecast_min,
            intercept,
            slope,
        )
        if prev_signature is not None and signature == prev_signature:
            continue
        prev_signature = signature

        range_width = (
            (forecast_max - forecast_min)
            if forecast_max is not None and forecast_min is not None
            else None
        )

        rows.append(
            {
                "model": labels["model"],
                "ticker": labels["ticker"],
                "model_period": labels["model_period"],
                "model_date": labels["model_date"],
                "method": "regression",
                "parameter_name": "num_quarters_used",
                "parameter_value": num_quarters_used,
                "num_quarters_used": num_quarters_used,
                "forecast_value": forecast_value,
                "actual_value": actual_value if actual_value is not None else "",
                "forecast_max": forecast_max,
                "forecast_min": forecast_min,
                "range_width": range_width,
                "intercept": intercept,
                "slope": slope,
                "source_file": source_file,
            }
        )

    return rows


def write_rows_to_sheet(ws, columns: Sequence[str], rows: Sequence[Dict[str, Any]]) -> None:
    ws.append(list(columns))
    for row in rows:
        ws.append([row.get(col, "") for col in columns])

    for cell in ws[1]:
        cell.font = Font(bold=True)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    for idx, col_name in enumerate(columns, start=1):
        max_len = len(col_name)
        for r in range(2, ws.max_row + 1):
            cell_value = ws.cell(row=r, column=idx).value
            if cell_value is None:
                continue
            max_len = max(max_len, len(str(cell_value)))
        ws.column_dimensions[get_column_letter(idx)].width = min(max_len + 2, 48)


def save_output_workbook(
    output_path: Path,
    empirical_rows: Sequence[Dict[str, Any]],
    regression_rows: Sequence[Dict[str, Any]],
) -> None:
    wb = Workbook()
    ws_emp = wb.active
    ws_emp.title = "empirical_candidates"
    ws_reg = wb.create_sheet("regression_candidates")

    write_rows_to_sheet(ws_emp, EMPIRICAL_COLUMNS, empirical_rows)
    write_rows_to_sheet(ws_reg, REGRESSION_COLUMNS, regression_rows)
    wb.save(output_path)


def main() -> None:
    src_input_dir = input_dir.expanduser().resolve()
    dst_output_dir = output_dir.expanduser().resolve()

    if not src_input_dir.exists() or not src_input_dir.is_dir():
        raise FileNotFoundError(f"input_dir does not exist or is not a folder: {src_input_dir}")

    output_path = get_output_path(src_input_dir, dst_output_dir)

    empirical_rows: List[Dict[str, Any]] = []
    regression_rows: List[Dict[str, Any]] = []
    files_processed = 0

    app: Optional[xw.App] = None
    try:
        app = xw.App(visible=False, add_book=False)
        app.display_alerts = False
        app.screen_updating = False
        try:
            app.calculation = "manual"
        except Exception:
            # Some Excel environments may not expose this property through xlwings.
            pass

        for file_path in sorted(src_input_dir.iterdir(), key=lambda p: p.name.lower()):
            if not file_path.is_file():
                print(f"Skipped {file_path.name}: not a file")
                continue
            if file_path.name.startswith("~"):
                print(f"Skipped {file_path.name}: temp file")
                continue
            if file_path.suffix.lower() != ".xlsx":
                print(f"Skipped {file_path.name}: not .xlsx")
                continue

            print(f"Processing {file_path.name}")
            labels = parse_file_labels(file_path.name)
            wb: Optional[xw.Book] = None

            try:
                wb = app.books.open(str(file_path), update_links=False)
                emp_sheet = get_sheet_case_insensitive(wb, "Empirical Model")
                reg_sheet = get_sheet_case_insensitive(wb, "Regression Model")

                if emp_sheet is None:
                    print(f"  skipped empirical in {file_path.name}: sheet not found")
                else:
                    empirical_rows.extend(
                        build_empirical_rows(emp_sheet, labels, file_path.name)
                    )

                if reg_sheet is None:
                    print(f"  skipped regression in {file_path.name}: sheet not found")
                else:
                    regression_rows.extend(
                        build_regression_rows(reg_sheet, labels, file_path.name)
                    )

                files_processed += 1

            except Exception as exc:
                print(f"Skipped {file_path.name}: {exc}")
            finally:
                if wb is not None:
                    safe_close_workbook(wb)
    finally:
        if app is not None:
            try:
                app.quit()
            except Exception:
                pass

    save_output_workbook(output_path, empirical_rows, regression_rows)

    print(f"Output path: {output_path}")
    print(f"Files processed: {files_processed}")
    print(f"Empirical rows: {len(empirical_rows)}")
    print(f"Regression rows: {len(regression_rows)}")


if __name__ == "__main__":
    main()
