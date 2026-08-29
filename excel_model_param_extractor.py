from __future__ import annotations

import re
from pathlib import Path
from typing import Any

import xlwings as xw
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

# -------------------------
# User-configurable folders
# -------------------------
input_dir = r"./input"
output_dir = r"./output"


EMPIRICAL_COLUMNS = [
    "model",
    "ticker",
    "model_period",
    "model_date",
    "method",
    "parameter_name",
    "parameter_value",
    "num_quarters_used",
    "last_quarter_used",
    "forecast_value",
    "actual_value",
    "forecast_max",
    "forecast_min",
    "range_width",
    "avg_penetration_pct",
    "quarterly_sales",
    "reported_sales",
    "growth_rate_pct",
    "sales_captured_in_db_pct",
    "source_file",
]

REGRESSION_COLUMNS = [
    "model",
    "ticker",
    "model_period",
    "model_date",
    "method",
    "parameter_name",
    "parameter_value",
    "num_quarters_used",
    "forecast_value",
    "actual_value",
    "forecast_max",
    "forecast_min",
    "range_width",
    "intercept",
    "slope",
    "source_file",
]

MONTH_MAP = {
    "jan": 1,
    "feb": 2,
    "mar": 3,
    "apr": 4,
    "may": 5,
    "jun": 6,
    "jul": 7,
    "aug": 8,
    "sep": 9,
    "oct": 10,
    "nov": 11,
    "dec": 12,
}

PHASE_DAY_MAP = {"early": 5, "mid": 15, "late": 25}

EMPIRICAL_N_QUARTERS = 10
REGRESSION_N_QUARTERS = 10


def normalize_label(value: Any) -> str:
    if value is None:
        return ""
    label = str(value).strip().lower()
    label = label.replace("%", " pct ")
    label = re.sub(r"[^a-z0-9]+", " ", label)
    return label.strip()


def to_2d(values: Any) -> list[list[Any]]:
    if values is None:
        return []
    if not isinstance(values, list):
        return [[values]]
    if values and not isinstance(values[0], list):
        return [values]
    return values


def to_float(value: Any) -> float | None:
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace(",", "")
    if not text:
        return None
    if text.endswith("%"):
        number_part = text[:-1].strip()
        try:
            return float(number_part) / 100.0
        except ValueError:
            return None
    try:
        return float(text)
    except ValueError:
        return None


def to_int(value: Any) -> int | None:
    parsed = to_float(value)
    if parsed is None:
        return None
    return int(round(parsed))


def value_or_blank(value: Any) -> Any:
    if value is None:
        return ""
    return value


def build_unique_output_path(in_dir: Path, out_dir: Path) -> Path:
    base_name = f"{in_dir.name}_PARAM"
    output_path = out_dir / f"{base_name}.xlsx"
    suffix_index = 1
    while output_path.exists():
        output_path = out_dir / f"{base_name}.{suffix_index}.xlsx"
        suffix_index += 1
    return output_path


def parse_file_label(file_name: str) -> dict[str, str]:
    stem = Path(file_name).stem
    parts = [part.strip() for part in stem.split(" - ")]

    ticker = ""
    if len(parts) >= 2:
        ticker = re.sub(r"[^A-Za-z0-9]", "", parts[1]).upper()
    if not ticker:
        ticker_match = re.search(r"\b([A-Z]{1,8})\b", stem)
        if ticker_match:
            ticker = ticker_match.group(1)

    period_match = re.search(
        r"(Early|Mid|Late)(Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)(20\d{2})",
        stem,
        flags=re.IGNORECASE,
    )

    model_period = ""
    model_date = ""
    if period_match:
        phase = period_match.group(1)
        month_token = period_match.group(2)
        year_token = period_match.group(3)

        month_num = MONTH_MAP[month_token.lower()]
        day_num = PHASE_DAY_MAP[phase.lower()]

        model_period = f"{phase[:1].upper()}{phase[1:].lower()}{month_token[:1].upper()}{month_token[1:].lower()}_{year_token}"
        model_date = f"{year_token}-{month_num:02d}-{day_num:02d}"

    model = f"{ticker}_{model_period}" if ticker and model_period else ticker or stem
    return {
        "ticker": ticker,
        "model_period": model_period,
        "model_date": model_date,
        "model": model,
    }


def find_anchor_max(sheet: xw.Sheet) -> tuple[int, int, int, int, list[list[Any]]] | None:
    used = sheet.used_range
    used_values = to_2d(used.value)
    if not used_values:
        return None

    start_row = used.row
    start_col = used.column

    for r_idx, row in enumerate(used_values):
        for c_idx, cell_value in enumerate(row):
            if isinstance(cell_value, str) and cell_value.strip().lower() == "max":
                return (
                    start_row + r_idx,
                    start_col + c_idx,
                    start_row,
                    start_col,
                    used_values,
                )
    return None


def build_header_map(
    used_values: list[list[Any]],
    used_start_row: int,
    used_start_col: int,
    anchor_row: int,
) -> dict[str, int]:
    header_map: dict[str, int] = {}
    header_row_idx = anchor_row - used_start_row
    if header_row_idx < 0 or header_row_idx >= len(used_values):
        return header_map

    for col_offset, value in enumerate(used_values[header_row_idx]):
        key = normalize_label(value)
        if key and key not in header_map:
            header_map[key] = used_start_col + col_offset
    return header_map


def resolve_col(
    header_map: dict[str, int],
    synonyms: list[str],
    anchor_col: int,
    default_offset: int,
) -> int:
    normalized_synonyms = [normalize_label(s) for s in synonyms]

    for syn in normalized_synonyms:
        if syn in header_map:
            return header_map[syn]

    for header_key, col in header_map.items():
        for syn in normalized_synonyms:
            if syn and (syn in header_key or header_key in syn):
                return col

    return anchor_col + default_offset


def read_row_values(sheet: xw.Sheet, row_idx: int, col_map: dict[str, int]) -> dict[str, Any]:
    valid_cols = {k: v for k, v in col_map.items() if v > 0}
    if not valid_cols:
        return {k: None for k in col_map}

    left_col = min(valid_cols.values())
    right_col = max(valid_cols.values())
    range_values = sheet.range((row_idx, left_col), (row_idx, right_col)).value
    if not isinstance(range_values, list):
        range_values = [range_values]

    out: dict[str, Any] = {}
    for key, col in col_map.items():
        if col <= 0:
            out[key] = None
            continue
        out[key] = range_values[col - left_col] if left_col <= col <= right_col else None
    return out


def set_r1c1_formula2(cell: xw.Range, r1c1_formula: str) -> None:
    # Preferred path: COM Formula2R1C1 (fast, no A1 conversion).
    try:
        cell.api.Formula2R1C1 = r1c1_formula
        return
    except Exception:
        pass

    # Fallback still uses formula2 with R1C1 text where supported.
    try:
        cell.formula2 = r1c1_formula
        return
    except Exception:
        pass

    # Last resort.
    cell.formula = r1c1_formula


def close_workbook_no_save(wb: xw.Book) -> None:
    try:
        wb.close(save=False)
        return
    except TypeError:
        pass
    except Exception:
        pass

    try:
        wb.close(False)
        return
    except Exception:
        pass

    try:
        wb.api.Close(SaveChanges=False)
        return
    except Exception:
        pass

    try:
        wb.close()
    except Exception:
        pass


def safe_num(value: Any) -> float | None:
    parsed = to_float(value)
    if parsed is None:
        return None
    return float(parsed)


def same_or_close(a: Any, b: Any, tolerance: float = 1e-9) -> bool:
    a_num = to_float(a)
    b_num = to_float(b)
    if a_num is not None and b_num is not None:
        return abs(a_num - b_num) <= tolerance
    return a == b


def extract_empirical_rows(
    wb: xw.Book, meta: dict[str, str], source_file: str
) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []

    try:
        sheet = wb.sheets["Empirical Model"]
    except Exception:
        print(f"Skipped empirical for {source_file}: 'Empirical Model' sheet not found")
        return rows

    anchor_data = find_anchor_max(sheet)
    if anchor_data is None:
        print(f"Skipped empirical for {source_file}: 'max' anchor not found")
        return rows

    anchor_row, anchor_col, used_start_row, used_start_col, used_values = anchor_data
    header_map = build_header_map(used_values, used_start_row, used_start_col, anchor_row)

    col_map = {
        "num_quarters_used": resolve_col(
            header_map,
            ["num quarters used", "n quarters", "# quarters", "quarters used"],
            anchor_col,
            -10,
        ),
        "last_quarter_used": resolve_col(
            header_map,
            ["last quarter used", "last qtr used", "last quarter"],
            anchor_col,
            -9,
        ),
        "forecast_value": resolve_col(
            header_map,
            [
                "estimated total sold",
                "estimate total sold",
                "forecast value",
                "total forecast",
                "tot fcst w o sa",
            ],
            anchor_col,
            -1,
        ),
        "actual_value": resolve_col(
            header_map,
            ["reported sales", "actual value", "actual sales", "actual"],
            anchor_col,
            -2,
        ),
        "forecast_max": resolve_col(header_map, ["max", "forecast max"], anchor_col, 0),
        "forecast_min": resolve_col(header_map, ["min", "forecast min"], anchor_col, 1),
        "avg_penetration_pct": resolve_col(
            header_map,
            ["avg penetration pct", "average penetration pct", "avg penetration"],
            anchor_col,
            -4,
        ),
        "quarterly_sales": resolve_col(
            header_map,
            ["quarterly sales", "quarter sales", "qtr sales"],
            anchor_col,
            -7,
        ),
        "reported_sales": resolve_col(
            header_map,
            ["reported sales", "reported", "db reported sales"],
            anchor_col,
            -11,
        ),
        "growth_rate_pct": resolve_col(
            header_map,
            ["growth rate pct", "growth pct", "growth rate"],
            anchor_col,
            -6,
        ),
        "sales_captured_in_db_pct": resolve_col(
            header_map,
            ["sales captured in db pct", "captured in db pct", "db capture pct"],
            anchor_col,
            -5,
        ),
    }

    x_col = col_map["reported_sales"]
    y_col = col_map["quarterly_sales"]

    history_end_row = anchor_row - 1
    scratch_avg_cell = sheet.range((anchor_row + EMPIRICAL_N_QUARTERS + 5, anchor_col + 8))

    for n_quarters in range(1, EMPIRICAL_N_QUARTERS + 1):
        current_row = anchor_row + n_quarters
        raw_row = read_row_values(sheet, current_row, col_map)

        history_start_row = max(used_start_row, history_end_row - n_quarters + 1)
        avg_penetration_pct = safe_num(raw_row.get("avg_penetration_pct"))

        if x_col > 0 and y_col > 0 and history_end_row >= history_start_row:
            avg_formula = (
                f'=IFERROR(AVERAGE('
                f'R{history_start_row}C{y_col}:R{history_end_row}C{y_col}/'
                f'R{history_start_row}C{x_col}:R{history_end_row}C{x_col}'
                f'),"")'
            )
            set_r1c1_formula2(scratch_avg_cell, avg_formula)
            wb.app.calculate()
            calculated_avg = safe_num(scratch_avg_cell.value)
            if calculated_avg is not None:
                avg_penetration_pct = calculated_avg

        num_quarters_used = to_int(raw_row.get("num_quarters_used")) or n_quarters
        last_quarter_used = value_or_blank(raw_row.get("last_quarter_used"))
        forecast_value = safe_num(raw_row.get("forecast_value"))
        actual_value = safe_num(raw_row.get("actual_value"))
        forecast_max = safe_num(raw_row.get("forecast_max"))
        forecast_min = safe_num(raw_row.get("forecast_min"))
        quarterly_sales = safe_num(raw_row.get("quarterly_sales"))
        reported_sales = safe_num(raw_row.get("reported_sales"))
        growth_rate_pct = safe_num(raw_row.get("growth_rate_pct"))
        sales_captured_in_db_pct = safe_num(raw_row.get("sales_captured_in_db_pct"))

        if forecast_value is None and actual_value is not None and avg_penetration_pct:
            # Fallback estimate when workbook does not expose a direct forecast column.
            forecast_value = actual_value / avg_penetration_pct if avg_penetration_pct else None

        if (
            forecast_value is None
            and actual_value is None
            and forecast_max is None
            and forecast_min is None
            and avg_penetration_pct is None
            and quarterly_sales is None
            and reported_sales is None
            and growth_rate_pct is None
            and sales_captured_in_db_pct is None
        ):
            continue

        range_width = (
            (forecast_max - forecast_min)
            if forecast_max is not None and forecast_min is not None
            else None
        )

        rows.append(
            {
                "model": meta["model"],
                "ticker": meta["ticker"],
                "model_period": meta["model_period"],
                "model_date": meta["model_date"],
                "method": "empirical",
                "parameter_name": "avg_penetration_pct",
                "parameter_value": avg_penetration_pct,
                "num_quarters_used": num_quarters_used,
                "last_quarter_used": last_quarter_used,
                "forecast_value": forecast_value,
                "actual_value": actual_value,
                "forecast_max": forecast_max,
                "forecast_min": forecast_min,
                "range_width": range_width,
                "avg_penetration_pct": avg_penetration_pct,
                "quarterly_sales": quarterly_sales,
                "reported_sales": reported_sales,
                "growth_rate_pct": growth_rate_pct,
                "sales_captured_in_db_pct": sales_captured_in_db_pct,
                "source_file": source_file,
            }
        )

    try:
        scratch_avg_cell.clear_contents()
    except Exception:
        pass

    return rows


def extract_regression_rows(
    wb: xw.Book, meta: dict[str, str], source_file: str
) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []

    try:
        sheet = wb.sheets["Regression Model"]
    except Exception:
        print(f"Skipped regression for {source_file}: 'Regression Model' sheet not found")
        return rows

    anchor_data = find_anchor_max(sheet)
    if anchor_data is None:
        print(f"Skipped regression for {source_file}: 'max' anchor not found")
        return rows

    anchor_row, anchor_col, used_start_row, used_start_col, used_values = anchor_data
    header_map = build_header_map(used_values, used_start_row, used_start_col, anchor_row)

    # Required by spec.
    y_col = anchor_col - 7
    x_col = anchor_col - 11

    col_map = {
        "num_quarters_used": resolve_col(
            header_map,
            ["num quarters used", "n quarters", "# quarters", "quarters used"],
            anchor_col,
            -10,
        ),
        "forecast_value": resolve_col(
            header_map,
            ["tot fcst w o sa", "tot fcst wo sa", "total forecast", "forecast value"],
            anchor_col,
            -1,
        ),
        "actual_value": resolve_col(
            header_map,
            ["actual value", "actual sales", "reported sales", "actual"],
            anchor_col,
            -2,
        ),
        "forecast_max": resolve_col(header_map, ["max", "forecast max"], anchor_col, 0),
        "forecast_min": resolve_col(header_map, ["min", "forecast min"], anchor_col, 1),
    }

    history_end_row = anchor_row - 1
    scratch_intercept = sheet.range((anchor_row + REGRESSION_N_QUARTERS + 5, anchor_col + 8))
    scratch_slope = sheet.range((anchor_row + REGRESSION_N_QUARTERS + 5, anchor_col + 9))

    for n_quarters in range(1, REGRESSION_N_QUARTERS + 1):
        current_row = anchor_row + n_quarters
        raw_row = read_row_values(sheet, current_row, col_map)

        history_start_row = max(used_start_row, history_end_row - n_quarters + 1)
        intercept = None
        slope = None

        if x_col > 0 and y_col > 0 and (history_end_row - history_start_row + 1) >= 2:
            intercept_formula = (
                f'=IFERROR(INTERCEPT('
                f'R{history_start_row}C{y_col}:R{history_end_row}C{y_col},'
                f'R{history_start_row}C{x_col}:R{history_end_row}C{x_col}'
                f'),"")'
            )
            slope_formula = (
                f'=IFERROR(SLOPE('
                f'R{history_start_row}C{y_col}:R{history_end_row}C{y_col},'
                f'R{history_start_row}C{x_col}:R{history_end_row}C{x_col}'
                f'),"")'
            )
            set_r1c1_formula2(scratch_intercept, intercept_formula)
            set_r1c1_formula2(scratch_slope, slope_formula)
            wb.app.calculate()
            intercept = safe_num(scratch_intercept.value)
            slope = safe_num(scratch_slope.value)

        num_quarters_used = to_int(raw_row.get("num_quarters_used")) or n_quarters
        forecast_value = safe_num(raw_row.get("forecast_value"))
        actual_value = safe_num(raw_row.get("actual_value"))
        forecast_max = safe_num(raw_row.get("forecast_max"))
        forecast_min = safe_num(raw_row.get("forecast_min"))

        if (
            forecast_value is None
            and forecast_max is None
            and forecast_min is None
            and intercept is None
            and slope is None
        ):
            continue

        range_width = (
            (forecast_max - forecast_min)
            if forecast_max is not None and forecast_min is not None
            else None
        )

        new_row = {
            "model": meta["model"],
            "ticker": meta["ticker"],
            "model_period": meta["model_period"],
            "model_date": meta["model_date"],
            "method": "regression",
            "parameter_name": "num_quarters_used",
            "parameter_value": num_quarters_used,
            "num_quarters_used": num_quarters_used,
            "forecast_value": forecast_value,
            "actual_value": value_or_blank(actual_value),
            "forecast_max": forecast_max,
            "forecast_min": forecast_min,
            "range_width": range_width,
            "intercept": intercept,
            "slope": slope,
            "source_file": source_file,
        }

        if rows:
            previous = rows[-1]
            duplicate = (
                same_or_close(new_row["forecast_value"], previous["forecast_value"])
                and same_or_close(new_row["forecast_max"], previous["forecast_max"])
                and same_or_close(new_row["forecast_min"], previous["forecast_min"])
                and same_or_close(new_row["intercept"], previous["intercept"])
                and same_or_close(new_row["slope"], previous["slope"])
            )
            if duplicate:
                continue

        rows.append(new_row)

    try:
        scratch_intercept.clear_contents()
        scratch_slope.clear_contents()
    except Exception:
        pass

    return rows


def write_output_sheet(
    ws,
    columns: list[str],
    rows: list[dict[str, Any]],
) -> None:
    ws.append(columns)

    for row in rows:
        ws.append([value_or_blank(row.get(col)) for col in columns])

    for cell in ws[1]:
        cell.font = Font(bold=True)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    for col_idx, col_name in enumerate(columns, start=1):
        max_len = len(col_name)
        for row_idx in range(2, ws.max_row + 1):
            value = ws.cell(row=row_idx, column=col_idx).value
            if value is None:
                continue
            max_len = max(max_len, len(str(value)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 40)


def write_output_workbook(
    output_path: Path,
    empirical_rows: list[dict[str, Any]],
    regression_rows: list[dict[str, Any]],
) -> None:
    output_wb = Workbook()
    default_sheet = output_wb.active
    output_wb.remove(default_sheet)

    empirical_ws = output_wb.create_sheet("empirical_candidates")
    regression_ws = output_wb.create_sheet("regression_candidates")

    write_output_sheet(empirical_ws, EMPIRICAL_COLUMNS, empirical_rows)
    write_output_sheet(regression_ws, REGRESSION_COLUMNS, regression_rows)

    output_wb.save(output_path)


def should_process_file(path: Path) -> tuple[bool, str]:
    if not path.is_file():
        return False, "not a file"
    if path.name.startswith("~"):
        return False, "temporary file"
    if path.suffix.lower() != ".xlsx":
        return False, "not an .xlsx file"
    return True, ""


def process_workbook(
    app: xw.App,
    file_path: Path,
    empirical_rows: list[dict[str, Any]],
    regression_rows: list[dict[str, Any]],
) -> None:
    wb = app.books.open(str(file_path), update_links=False)
    try:
        metadata = parse_file_label(file_path.name)
        empirical_rows.extend(extract_empirical_rows(wb, metadata, file_path.name))
        regression_rows.extend(extract_regression_rows(wb, metadata, file_path.name))
    finally:
        close_workbook_no_save(wb)


def main() -> None:
    in_dir = Path(input_dir).expanduser().resolve()
    out_dir = Path(output_dir).expanduser().resolve()
    out_dir.mkdir(parents=True, exist_ok=True)

    if not in_dir.exists():
        raise FileNotFoundError(f"Input directory does not exist: {in_dir}")
    if not in_dir.is_dir():
        raise NotADirectoryError(f"Input path is not a directory: {in_dir}")

    output_path = build_unique_output_path(in_dir, out_dir)

    empirical_rows: list[dict[str, Any]] = []
    regression_rows: list[dict[str, Any]] = []
    processed_count = 0

    app = xw.App(visible=False, add_book=False)
    app.display_alerts = False
    app.screen_updating = False

    try:
        for file_path in sorted(in_dir.iterdir()):
            should_process, reason = should_process_file(file_path)
            if not should_process:
                print(f"Skipped: {file_path.name} ({reason})")
                continue

            try:
                process_workbook(app, file_path, empirical_rows, regression_rows)
                processed_count += 1
                print(f"Processed: {file_path.name}")
            except Exception as exc:
                print(f"Skipped: {file_path.name} (processing error: {exc})")
    finally:
        try:
            app.quit()
        except Exception:
            pass

    write_output_workbook(output_path, empirical_rows, regression_rows)

    print(f"Output path: {output_path}")
    print(f"Files processed: {processed_count}")
    print(f"Empirical rows: {len(empirical_rows)}")
    print(f"Regression rows: {len(regression_rows)}")


if __name__ == "__main__":
    main()
