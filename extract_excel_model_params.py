from __future__ import annotations

import re
from dataclasses import dataclass
from datetime import date
from pathlib import Path
from typing import Any

import xlwings as xw
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter


# =========================
# User-configurable inputs
# =========================
input_dir = Path("/path/to/input")
output_dir = Path("/path/to/output")


MAX_QUARTERS = 10

EMPIRICAL_COLUMNS = [
    "model",
    "ticker",
    "model_period",
    "model_date",
    "method",
    "parameter_name",
    "parameter_value",
    "num_quarters_used",
    "last_quarter_used",
    "forecast_value",
    "actual_value",
    "forecast_max",
    "forecast_min",
    "range_width",
    "avg_penetration_pct",
    "quarterly_sales",
    "reported_sales",
    "growth_rate_pct",
    "sales_captured_in_db_pct",
    "source_file",
]

REGRESSION_COLUMNS = [
    "model",
    "ticker",
    "model_period",
    "model_date",
    "method",
    "parameter_name",
    "parameter_value",
    "num_quarters_used",
    "forecast_value",
    "actual_value",
    "forecast_max",
    "forecast_min",
    "range_width",
    "intercept",
    "slope",
    "source_file",
]

MONTH_TO_NUM = {
    "Jan": 1,
    "Feb": 2,
    "Mar": 3,
    "Apr": 4,
    "May": 5,
    "Jun": 6,
    "Jul": 7,
    "Aug": 8,
    "Sep": 9,
    "Oct": 10,
    "Nov": 11,
    "Dec": 12,
}

PERIOD_DAY = {
    "Early": 5,
    "Mid": 15,
    "Late": 25,
}

FILE_PATTERN = re.compile(
    r"^(?P<prefix>.+?)\s*-\s*(?P<ticker>[A-Za-z0-9]+)\s*-\s*"
    r"(?P<period>(Early|Mid|Late)(Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)\d{4})"
    r"(?:[_\-\s].*)?\.xlsx$",
    re.IGNORECASE,
)


@dataclass(frozen=True)
class FileMeta:
    model: str
    ticker: str
    model_period: str
    model_date: str
    source_file: str


def to_float(value: Any) -> float | None:
    if value is None or value == "":
        return None
    if isinstance(value, bool):
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def parse_file_meta(file_name: str) -> FileMeta | None:
    match = FILE_PATTERN.match(file_name)
    if not match:
        return None

    ticker = match.group("ticker").upper()
    period_token = match.group("period")
    period_match = re.match(
        r"(?P<segment>Early|Mid|Late)(?P<month>[A-Za-z]{3})(?P<year>\d{4})",
        period_token,
        flags=re.IGNORECASE,
    )
    if not period_match:
        return None

    segment = period_match.group("segment").title()
    month_abbrev = period_match.group("month").title()
    year = int(period_match.group("year"))

    month_num = MONTH_TO_NUM.get(month_abbrev)
    day = PERIOD_DAY.get(segment)
    if month_num is None or day is None:
        return None

    model_period = f"{segment}{month_abbrev}_{year}"
    model_date = date(year, month_num, day).isoformat()
    model = f"{ticker}_{model_period}"

    return FileMeta(
        model=model,
        ticker=ticker,
        model_period=model_period,
        model_date=model_date,
        source_file=file_name,
    )


def next_output_path(input_folder: Path, out_dir: Path) -> Path:
    base_name = f"{input_folder.name}_PARAM"
    candidate = out_dir / f"{base_name}.xlsx"
    if not candidate.exists():
        return candidate

    idx = 1
    while True:
        candidate = out_dir / f"{base_name}.{idx}.xlsx"
        if not candidate.exists():
            return candidate
        idx += 1


def close_workbook_no_save(workbook: xw.Book) -> None:
    try:
        workbook.close(save=False)
    except TypeError:
        try:
            workbook.api.Close(SaveChanges=False)
        except Exception:
            workbook.close()
    except Exception:
        try:
            workbook.api.Close(SaveChanges=False)
        except Exception:
            pass


def normalize_used_values(values: Any) -> list[list[Any]]:
    if values is None:
        return []
    if isinstance(values, list):
        if not values:
            return []
        if isinstance(values[0], list):
            return values
        return [values]
    return [[values]]


def find_anchor_max(sheet: xw.Sheet) -> tuple[int, int] | None:
    used = sheet.used_range
    matrix = normalize_used_values(used.value)
    if not matrix:
        return None

    start_row = used.row
    start_col = used.column
    for r_idx, row_vals in enumerate(matrix):
        for c_idx, cell_val in enumerate(row_vals):
            if isinstance(cell_val, str) and cell_val.strip().lower() == "max":
                return start_row + r_idx, start_col + c_idx
    return None


def extract_xy_series(sheet: xw.Sheet, anchor_row: int, x_col: int, y_col: int) -> list[tuple[int, float, float]]:
    series: list[tuple[int, float, float]] = []

    row = anchor_row - 1
    while row >= 1:
        x_val = to_float(sheet.cells(row, x_col).value)
        y_val = to_float(sheet.cells(row, y_col).value)
        if x_val is None or y_val is None:
            if series:
                break
            row -= 1
            continue
        series.append((row, x_val, y_val))
        row -= 1

    if not series:
        row = anchor_row + 1
        while row <= sheet.cells.last_cell.row:
            x_val = to_float(sheet.cells(row, x_col).value)
            y_val = to_float(sheet.cells(row, y_col).value)
            if x_val is None or y_val is None:
                if series:
                    break
                row += 1
                continue
            series.append((row, x_val, y_val))
            row += 1

    series.reverse()
    return series


def to_number_or_blank(value: float | None) -> float | None:
    return None if value is None else float(value)


def build_empirical_rows(wb: xw.Book, meta: FileMeta) -> list[dict[str, Any]]:
    try:
        sheet = wb.sheets["Empirical Model"]
    except Exception:
        print(f"SKIPPED empirical: {meta.source_file} (missing 'Empirical Model' sheet)")
        return []

    anchor = find_anchor_max(sheet)
    if anchor is None:
        print(f"SKIPPED empirical: {meta.source_file} (could not find 'max' anchor)")
        return []

    anchor_row, anchor_col = anchor
    x_col = anchor_col - 11
    y_col = anchor_col - 7
    series = extract_xy_series(sheet, anchor_row, x_col, y_col)
    if len(series) < 2:
        print(f"SKIPPED empirical: {meta.source_file} (insufficient x/y data)")
        return []

    n_limit = min(MAX_QUARTERS, len(series))
    avg_pen_cells: list[tuple[int, int, int]] = []  # (n_quarters, scratch_row, last_row)
    scratch_col = anchor_col + 3

    for n_quarters in range(1, n_limit + 1):
        subset = series[-n_quarters:]
        first_row = subset[0][0]
        last_row = subset[-1][0]
        scratch_row = anchor_row + n_quarters
        avg_cell = sheet.cells(scratch_row, scratch_col)
        avg_cell.formula2 = (
            f'=IFERROR(AVERAGE(IFERROR(R{first_row}C{x_col}:R{last_row}C{x_col}/'
            f'R{first_row}C{y_col}:R{last_row}C{y_col},"")), "")'
        )
        avg_pen_cells.append((n_quarters, scratch_row, last_row))

    wb.app.calculate()

    rows: list[dict[str, Any]] = []
    for n_quarters, scratch_row, last_row in avg_pen_cells:
        subset = series[-n_quarters:]
        ratios = [x / y for _, x, y in subset if y]
        if not ratios:
            continue

        _, quarterly_sales, reported_sales = subset[-1]
        avg_pen_ratio = to_float(sheet.cells(scratch_row, scratch_col).value)
        if avg_pen_ratio is None or avg_pen_ratio <= 0:
            avg_pen_ratio = sum(ratios) / len(ratios)

        min_pen_ratio = min(ratios)
        max_pen_ratio = max(ratios)

        forecast_value = quarterly_sales / avg_pen_ratio if avg_pen_ratio > 0 else None
        forecast_max = quarterly_sales / min_pen_ratio if min_pen_ratio > 0 else None
        forecast_min = quarterly_sales / max_pen_ratio if max_pen_ratio > 0 else None
        range_width = None
        if forecast_max is not None and forecast_min is not None:
            range_width = forecast_max - forecast_min

        growth_rate_pct = None
        if len(subset) >= 2 and subset[-2][1] != 0:
            growth_rate_pct = ((subset[-1][1] / subset[-2][1]) - 1) * 100

        sales_captured_in_db_pct = None
        if reported_sales != 0:
            sales_captured_in_db_pct = (quarterly_sales / reported_sales) * 100

        last_quarter_used = sheet.cells(last_row, x_col - 1).value
        if last_quarter_used is None or last_quarter_used == "":
            last_quarter_used = f"row_{last_row}"

        avg_penetration_pct = avg_pen_ratio * 100 if avg_pen_ratio is not None else None

        rows.append(
            {
                "model": meta.model,
                "ticker": meta.ticker,
                "model_period": meta.model_period,
                "model_date": meta.model_date,
                "method": "empirical",
                "parameter_name": "avg_penetration_pct",
                "parameter_value": to_number_or_blank(avg_penetration_pct),
                "num_quarters_used": n_quarters,
                "last_quarter_used": last_quarter_used,
                "forecast_value": to_number_or_blank(forecast_value),
                "actual_value": to_number_or_blank(reported_sales),
                "forecast_max": to_number_or_blank(forecast_max),
                "forecast_min": to_number_or_blank(forecast_min),
                "range_width": to_number_or_blank(range_width),
                "avg_penetration_pct": to_number_or_blank(avg_penetration_pct),
                "quarterly_sales": to_number_or_blank(quarterly_sales),
                "reported_sales": to_number_or_blank(reported_sales),
                "growth_rate_pct": to_number_or_blank(growth_rate_pct),
                "sales_captured_in_db_pct": to_number_or_blank(sales_captured_in_db_pct),
                "source_file": meta.source_file,
            }
        )

    return rows


def _signature(*values: Any) -> tuple[Any, ...]:
    sig: list[Any] = []
    for value in values:
        if isinstance(value, float):
            sig.append(round(value, 10))
        else:
            sig.append(value)
    return tuple(sig)


def build_regression_rows(wb: xw.Book, meta: FileMeta) -> list[dict[str, Any]]:
    try:
        sheet = wb.sheets["Regression Model"]
    except Exception:
        print(f"SKIPPED regression: {meta.source_file} (missing 'Regression Model' sheet)")
        return []

    anchor = find_anchor_max(sheet)
    if anchor is None:
        print(f"SKIPPED regression: {meta.source_file} (could not find 'max' anchor)")
        return []

    anchor_row, anchor_col = anchor
    y_col = anchor_col - 7
    x_col = anchor_col - 11
    series = extract_xy_series(sheet, anchor_row, x_col, y_col)
    if len(series) < 2:
        print(f"SKIPPED regression: {meta.source_file} (insufficient x/y data)")
        return []

    n_limit = min(MAX_QUARTERS, len(series))
    scratch_rows: list[tuple[int, int]] = []

    for n_quarters in range(1, n_limit + 1):
        subset = series[-n_quarters:]
        first_row = subset[0][0]
        last_row = subset[-1][0]
        scratch_row = anchor_row + n_quarters
        intercept_cell = sheet.cells(scratch_row, anchor_col + 1)
        slope_cell = sheet.cells(scratch_row, anchor_col + 2)
        forecast_cell = sheet.cells(scratch_row, anchor_col + 3)
        max_cell = sheet.cells(scratch_row, anchor_col + 4)
        min_cell = sheet.cells(scratch_row, anchor_col + 5)

        intercept_cell.formula2 = (
            f'=IFERROR(INTERCEPT(R{first_row}C{y_col}:R{last_row}C{y_col},'
            f'R{first_row}C{x_col}:R{last_row}C{x_col}), "")'
        )
        slope_cell.formula2 = (
            f'=IFERROR(SLOPE(R{first_row}C{y_col}:R{last_row}C{y_col},'
            f'R{first_row}C{x_col}:R{last_row}C{x_col}), "")'
        )
        forecast_cell.formula2 = f'=IFERROR(RC[-2] + (RC[-1] * R{last_row}C{x_col}), "")'
        max_cell.formula2 = f'=IFERROR(MAX(R{first_row}C{y_col}:R{last_row}C{y_col}), "")'
        min_cell.formula2 = f'=IFERROR(MIN(R{first_row}C{y_col}:R{last_row}C{y_col}), "")'
        scratch_rows.append((n_quarters, scratch_row))

    wb.app.calculate()

    rows: list[dict[str, Any]] = []
    prev_sig: tuple[Any, ...] | None = None

    for n_quarters, scratch_row in scratch_rows:
        intercept = to_float(sheet.cells(scratch_row, anchor_col + 1).value)
        slope = to_float(sheet.cells(scratch_row, anchor_col + 2).value)
        forecast_total_without_sa = to_float(sheet.cells(scratch_row, anchor_col + 3).value)
        forecast_max = to_float(sheet.cells(scratch_row, anchor_col + 4).value)
        forecast_min = to_float(sheet.cells(scratch_row, anchor_col + 5).value)

        row_sig = _signature(
            n_quarters,
            intercept,
            slope,
            forecast_total_without_sa,
            forecast_max,
            forecast_min,
        )
        if prev_sig is not None and row_sig == prev_sig:
            continue
        prev_sig = row_sig

        range_width = None
        if forecast_max is not None and forecast_min is not None:
            range_width = forecast_max - forecast_min

        rows.append(
            {
                "model": meta.model,
                "ticker": meta.ticker,
                "model_period": meta.model_period,
                "model_date": meta.model_date,
                "method": "regression",
                "parameter_name": "num_quarters_used",
                "parameter_value": n_quarters,
                "num_quarters_used": n_quarters,
                "forecast_value": to_number_or_blank(forecast_total_without_sa),
                "actual_value": None,
                "forecast_max": to_number_or_blank(forecast_max),
                "forecast_min": to_number_or_blank(forecast_min),
                "range_width": to_number_or_blank(range_width),
                "intercept": to_number_or_blank(intercept),
                "slope": to_number_or_blank(slope),
                "source_file": meta.source_file,
            }
        )

    return rows


def write_sheet(workbook: Workbook, name: str, columns: list[str], rows: list[dict[str, Any]]) -> None:
    ws = workbook.create_sheet(title=name)
    ws.append(columns)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for row in rows:
        ws.append([row.get(col) for col in columns])

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(columns))}{max(1, ws.max_row)}"

    for idx, col_name in enumerate(columns, start=1):
        max_len = len(col_name)
        for row_idx in range(2, ws.max_row + 1):
            value = ws.cell(row=row_idx, column=idx).value
            if value is None:
                continue
            if isinstance(value, float):
                text = f"{value:.10g}"
            else:
                text = str(value)
            if len(text) > max_len:
                max_len = len(text)
        ws.column_dimensions[get_column_letter(idx)].width = min(max_len + 2, 48)


def build_output_workbook(
    empirical_rows: list[dict[str, Any]],
    regression_rows: list[dict[str, Any]],
    output_path: Path,
) -> None:
    wb_out = Workbook()
    wb_out.remove(wb_out.active)
    write_sheet(wb_out, "empirical_candidates", EMPIRICAL_COLUMNS, empirical_rows)
    write_sheet(wb_out, "regression_candidates", REGRESSION_COLUMNS, regression_rows)
    wb_out.save(output_path)


def iter_excel_files(folder: Path) -> list[Path]:
    files: list[Path] = []
    for path in sorted(folder.iterdir(), key=lambda p: p.name.lower()):
        if not path.is_file():
            print(f"SKIPPED {path.name}: not a file")
            continue
        if path.name.startswith("~"):
            print(f"SKIPPED {path.name}: temp file")
            continue
        if path.suffix.lower() != ".xlsx":
            print(f"SKIPPED {path.name}: not an .xlsx file")
            continue
        files.append(path)
    return files


def run() -> None:
    if not input_dir.exists() or not input_dir.is_dir():
        raise FileNotFoundError(f"input_dir does not exist or is not a folder: {input_dir}")

    output_dir.mkdir(parents=True, exist_ok=True)
    files = iter_excel_files(input_dir)

    empirical_rows: list[dict[str, Any]] = []
    regression_rows: list[dict[str, Any]] = []
    processed_files = 0

    app = xw.App(visible=False, add_book=False)
    app.display_alerts = False
    app.screen_updating = False

    try:
        for file_path in files:
            meta = parse_file_meta(file_path.name)
            if meta is None:
                print(f"SKIPPED {file_path.name}: filename does not match expected pattern")
                continue

            print(f"PROCESSING {file_path.name}")
            wb: xw.Book | None = None
            try:
                wb = app.books.open(str(file_path), update_links=False)
                empirical_file_rows = build_empirical_rows(wb, meta)
                regression_file_rows = build_regression_rows(wb, meta)
                empirical_rows.extend(empirical_file_rows)
                regression_rows.extend(regression_file_rows)
                processed_files += 1
                print(
                    f"PROCESSED {file_path.name}: "
                    f"empirical_rows={len(empirical_file_rows)}, "
                    f"regression_rows={len(regression_file_rows)}"
                )
            except Exception as exc:
                print(f"SKIPPED {file_path.name}: processing error -> {exc}")
            finally:
                if wb is not None:
                    close_workbook_no_save(wb)
    finally:
        app.quit()

    output_path = next_output_path(input_dir, output_dir)
    build_output_workbook(empirical_rows, regression_rows, output_path)

    print(f"OUTPUT {output_path}")
    print(f"FILES_PROCESSED {processed_files}")
    print(f"EMPIRICAL_ROWS {len(empirical_rows)}")
    print(f"REGRESSION_ROWS {len(regression_rows)}")


if __name__ == "__main__":
    run()
