#!/usr/bin/env python3
"""Extract empirical/regression candidates from model Excel files."""

from __future__ import annotations

import re
from pathlib import Path
from typing import Any

import xlwings as xw
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

# ===== User-configurable paths =====
input_dir = Path("./input")
output_dir = Path("./output")

EMPIRICAL_COLUMNS = [
    "model",
    "ticker",
    "model_period",
    "model_date",
    "method",
    "parameter_name",
    "parameter_value",
    "num_quarters_used",
    "last_quarter_used",
    "forecast_value",
    "actual_value",
    "forecast_max",
    "forecast_min",
    "range_width",
    "avg_penetration_pct",
    "quarterly_sales",
    "reported_sales",
    "growth_rate_pct",
    "sales_captured_in_db_pct",
    "source_file",
]

REGRESSION_COLUMNS = [
    "model",
    "ticker",
    "model_period",
    "model_date",
    "method",
    "parameter_name",
    "parameter_value",
    "num_quarters_used",
    "forecast_value",
    "actual_value",
    "forecast_max",
    "forecast_min",
    "range_width",
    "intercept",
    "slope",
    "source_file",
]

PHASE_DAY = {"early": 5, "mid": 15, "late": 25}
MONTH_MAP = {
    "jan": 1,
    "feb": 2,
    "mar": 3,
    "apr": 4,
    "may": 5,
    "jun": 6,
    "jul": 7,
    "aug": 8,
    "sep": 9,
    "oct": 10,
    "nov": 11,
    "dec": 12,
}


def normalize_text(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip().lower()
    if not text:
        return ""
    text = re.sub(r"[^a-z0-9]+", "_", text)
    return text.strip("_")


def is_blank(value: Any) -> bool:
    return value is None or (isinstance(value, str) and value.strip() == "")


def as_number(value: Any) -> float | None:
    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        raw = value.strip().replace(",", "").replace("$", "")
        if raw.endswith("%"):
            raw = raw[:-1]
        if not raw:
            return None
        try:
            return float(raw)
        except ValueError:
            return None
    return None


def to_int_if_possible(value: Any, fallback: int) -> int:
    num = as_number(value)
    if num is None:
        return fallback
    try:
        return int(round(num))
    except Exception:
        return fallback


def safe_close_workbook(wb: xw.Book | None) -> None:
    if wb is None:
        return
    try:
        wb.close(save=False)
        return
    except TypeError:
        pass
    except Exception:
        pass
    try:
        wb.close(False)
        return
    except Exception:
        pass
    try:
        wb.api.Close(SaveChanges=False)
    except Exception:
        pass


def parse_file_label(file_name: str) -> dict[str, str]:
    stem = Path(file_name).stem
    parts = [part.strip() for part in stem.split(" - ") if part.strip()]

    ticker = "UNKNOWN"
    if len(parts) >= 2:
        ticker = parts[1].split("_")[0].strip().upper()
    else:
        m_ticker = re.search(r"\b([A-Z]{1,8})\b", stem)
        if m_ticker:
            ticker = m_ticker.group(1).upper()

    period_token = ""
    if len(parts) >= 3:
        period_token = parts[2].split("_")[0].strip()
    if not period_token:
        m_period = re.search(r"(Early|Mid|Late)[A-Za-z]{3,9}\d{4}", stem, re.IGNORECASE)
        period_token = m_period.group(0) if m_period else "MidJan2026"

    m = re.search(r"(Early|Mid|Late)\s*([A-Za-z]{3,9})\s*(\d{4})", period_token, re.IGNORECASE)
    if m:
        phase_raw, month_raw, year_raw = m.groups()
        phase_key = phase_raw.lower()
        phase = phase_raw.capitalize()
        month_abbr = month_raw[:3].title()
        year = int(year_raw)
        month_num = MONTH_MAP.get(month_abbr.lower(), 1)
        day = PHASE_DAY.get(phase_key, 15)
        model_period = f"{phase}{month_abbr}_{year}"
        model_date = f"{year:04d}-{month_num:02d}-{day:02d}"
    else:
        model_period = period_token.replace(" ", "_")
        model_date = ""

    model = f"{ticker}_{model_period}"
    return {
        "model": model,
        "ticker": ticker,
        "model_period": model_period,
        "model_date": model_date,
    }


def next_output_path(input_path: Path, out_dir: Path) -> Path:
    base_name = f"{input_path.name}_PARAM"
    candidate = out_dir / f"{base_name}.xlsx"
    if not candidate.exists():
        return candidate

    version = 1
    while True:
        candidate = out_dir / f"{base_name}.{version}.xlsx"
        if not candidate.exists():
            return candidate
        version += 1


def get_sheet_by_name(wb: xw.Book, name: str) -> xw.Sheet | None:
    target = name.strip().lower()
    for sheet in wb.sheets:
        if sheet.name.strip().lower() == target:
            return sheet
    return None


def get_used_grid(sheet: xw.Sheet) -> tuple[int, int, list[list[Any]]]:
    used = sheet.used_range
    values = used.options(ndim=2).value
    if values is None:
        values = [[None]]
    return used.row, used.column, values


def find_anchor(values: list[list[Any]], base_row: int, base_col: int, target: str = "max") -> tuple[int, int] | None:
    target_norm = target.strip().lower()
    for r_idx, row in enumerate(values):
        for c_idx, value in enumerate(row):
            if isinstance(value, str) and value.strip().lower() == target_norm:
                return base_row + r_idx, base_col + c_idx
    return None


def collect_header_offsets(
    values: list[list[Any]],
    base_row: int,
    base_col: int,
    anchor_row: int,
    anchor_col: int,
    window: int = 40,
) -> dict[str, int]:
    row_idx = anchor_row - base_row
    if row_idx < 0 or row_idx >= len(values):
        return {}

    offsets: dict[str, int] = {}
    header_row = values[row_idx]
    for c_idx, value in enumerate(header_row):
        col = base_col + c_idx
        if abs(col - anchor_col) > window:
            continue
        key = normalize_text(value)
        if key and key not in offsets:
            offsets[key] = col - anchor_col
    return offsets


def choose_offset(header_offsets: dict[str, int], aliases: list[str], default: int | None = None) -> int | None:
    alias_norm = [normalize_text(alias) for alias in aliases if alias]
    for alias in alias_norm:
        for key, offset in header_offsets.items():
            if alias and alias in key:
                return offset
    return default


def find_header_column(
    values: list[list[Any]],
    base_row: int,
    base_col: int,
    max_abs_row: int,
    include_tokens: list[str],
    exclude_tokens: list[str] | None = None,
) -> int | None:
    include_norm = [normalize_text(token) for token in include_tokens if token]
    exclude_norm = [normalize_text(token) for token in (exclude_tokens or []) if token]

    max_idx = min(len(values) - 1, max(0, max_abs_row - base_row))
    for r_idx in range(max_idx + 1):
        row = values[r_idx]
        for c_idx, value in enumerate(row):
            key = normalize_text(value)
            if not key:
                continue
            if include_norm and not all(token in key for token in include_norm):
                continue
            if any(token in key for token in exclude_norm):
                continue
            return base_col + c_idx
    return None


def get_numeric_series(sheet: xw.Sheet, col: int | None, start_row: int, end_row: int) -> list[tuple[int, float]]:
    if col is None or end_row < start_row:
        return []

    values = sheet.range((start_row, col), (end_row, col)).options(ndim=1).value
    if values is None:
        return []
    if not isinstance(values, list):
        values = [values]

    out: list[tuple[int, float]] = []
    for idx, value in enumerate(values, start=start_row):
        numeric = as_number(value)
        if numeric is not None:
            out.append((idx, numeric))
    return out


def get_nonblank_series(sheet: xw.Sheet, col: int | None, start_row: int, end_row: int) -> list[tuple[int, Any]]:
    if col is None or end_row < start_row:
        return []
    values = sheet.range((start_row, col), (end_row, col)).options(ndim=1).value
    if values is None:
        return []
    if not isinstance(values, list):
        values = [values]
    out: list[tuple[int, Any]] = []
    for idx, value in enumerate(values, start=start_row):
        if not is_blank(value):
            out.append((idx, value))
    return out


def read_by_offset(sheet: xw.Sheet, row: int, anchor_col: int, offset: int | None) -> Any:
    if offset is None:
        return None
    col = anchor_col + offset
    if col < 1:
        return None
    try:
        return sheet.range((row, col)).value
    except Exception:
        return None


def set_formula2(cell: xw.Range, formula_r1c1: str) -> None:
    try:
        cell.formula2 = formula_r1c1
    except Exception:
        cell.formula = formula_r1c1


def max_used_col(values: list[list[Any]], base_col: int) -> int:
    widest = max((len(row) for row in values), default=1)
    return base_col + max(widest, 1) - 1


def extract_empirical_rows(wb: xw.Book, metadata: dict[str, str], source_file: str) -> list[dict[str, Any]]:
    sheet = get_sheet_by_name(wb, "Empirical Model")
    if sheet is None:
        print(f"  skipped empirical: sheet 'Empirical Model' not found ({source_file})")
        return []

    base_row, base_col, values = get_used_grid(sheet)
    anchor = find_anchor(values, base_row, base_col, target="max")
    if anchor is None:
        print(f"  skipped empirical: 'max' anchor not found ({source_file})")
        return []
    anchor_row, anchor_col = anchor

    header_offsets = collect_header_offsets(values, base_row, base_col, anchor_row, anchor_col)
    min_offset = choose_offset(header_offsets, ["min"], default=-1)
    forecast_offset = choose_offset(header_offsets, ["estimated_total_sold", "forecast_value", "forecast_total"])
    actual_offset = choose_offset(header_offsets, ["reported_sales", "actual_value", "actual", "reported"])
    num_quarters_offset = choose_offset(header_offsets, ["num_quarters_used", "num_quarters", "quarters_used", "num_qtrs"])
    last_quarter_offset = choose_offset(header_offsets, ["last_quarter_used", "last_quarter"])
    avg_pen_offset = choose_offset(header_offsets, ["avg_penetration_pct", "avg_penetration", "average_penetration"])
    q_sales_offset = choose_offset(header_offsets, ["quarterly_sales", "qtr_sales"])
    growth_offset = choose_offset(header_offsets, ["growth_rate_pct", "growth_rate", "growth"])
    captured_offset = choose_offset(header_offsets, ["sales_captured_in_db_pct", "captured_in_db", "captured"])

    # Pull supporting quarter-level columns once and compute rolling averages via formula2.
    penetration_col = find_header_column(
        values,
        base_row,
        base_col,
        anchor_row,
        include_tokens=["penetration"],
        exclude_tokens=["avg", "average"],
    )
    quarterly_col = find_header_column(values, base_row, base_col, anchor_row, include_tokens=["quarterly", "sales"])
    reported_col = find_header_column(values, base_row, base_col, anchor_row, include_tokens=["reported", "sales"])
    if reported_col is None:
        reported_col = find_header_column(values, base_row, base_col, anchor_row, include_tokens=["actual", "sales"])
    growth_col = find_header_column(values, base_row, base_col, anchor_row, include_tokens=["growth"])
    captured_col = find_header_column(values, base_row, base_col, anchor_row, include_tokens=["captured"])
    quarter_label_col = find_header_column(
        values,
        base_row,
        base_col,
        anchor_row,
        include_tokens=["quarter"],
        exclude_tokens=["num"],
    )

    pen_series = get_numeric_series(sheet, penetration_col, 1, anchor_row - 1)
    quarterly_series = get_numeric_series(sheet, quarterly_col, 1, anchor_row - 1)
    reported_series = get_numeric_series(sheet, reported_col, 1, anchor_row - 1)
    growth_series = get_numeric_series(sheet, growth_col, 1, anchor_row - 1)
    captured_series = get_numeric_series(sheet, captured_col, 1, anchor_row - 1)
    last_quarter_series = get_nonblank_series(sheet, quarter_label_col, 1, anchor_row - 1)

    n_quarters = 10
    avg_pen_formula_values: list[float | None] = [None] * n_quarters
    if pen_series:
        scratch_col = max(anchor_col + 25, max_used_col(values, base_col) + 2)
        scratch_start_row = anchor_row + 1
        for i in range(n_quarters):
            n_use = min(i + 1, len(pen_series))
            start_row = pen_series[-n_use][0]
            end_row = pen_series[-1][0]
            formula = f"=AVERAGE(R{start_row}C{penetration_col}:R{end_row}C{penetration_col})"
            set_formula2(sheet.range((scratch_start_row + i, scratch_col)), formula)
        wb.app.calculate()
        avg_vals = sheet.range(
            (scratch_start_row, scratch_col),
            (scratch_start_row + n_quarters - 1, scratch_col),
        ).options(ndim=1).value
        if not isinstance(avg_vals, list):
            avg_vals = [avg_vals]
        for i, value in enumerate(avg_vals[:n_quarters]):
            avg_pen_formula_values[i] = as_number(value)

    latest_quarterly = quarterly_series[-1][1] if quarterly_series else None
    latest_reported = reported_series[-1][1] if reported_series else None
    latest_growth = growth_series[-1][1] if growth_series else None
    latest_captured = captured_series[-1][1] if captured_series else None
    latest_last_quarter = last_quarter_series[-1][1] if last_quarter_series else None

    rows: list[dict[str, Any]] = []
    for i in range(1, n_quarters + 1):
        row_idx = anchor_row + i
        num_quarters_raw = read_by_offset(sheet, row_idx, anchor_col, num_quarters_offset)
        last_quarter_raw = read_by_offset(sheet, row_idx, anchor_col, last_quarter_offset)
        forecast_raw = read_by_offset(sheet, row_idx, anchor_col, forecast_offset)
        actual_raw = read_by_offset(sheet, row_idx, anchor_col, actual_offset)
        max_raw = read_by_offset(sheet, row_idx, anchor_col, 0)
        min_raw = read_by_offset(sheet, row_idx, anchor_col, min_offset)
        avg_pen_raw = read_by_offset(sheet, row_idx, anchor_col, avg_pen_offset)
        q_sales_raw = read_by_offset(sheet, row_idx, anchor_col, q_sales_offset)
        growth_raw = read_by_offset(sheet, row_idx, anchor_col, growth_offset)
        captured_raw = read_by_offset(sheet, row_idx, anchor_col, captured_offset)

        probe = [num_quarters_raw, last_quarter_raw, forecast_raw, actual_raw, max_raw, min_raw, avg_pen_raw, q_sales_raw]
        if all(is_blank(value) for value in probe) and avg_pen_formula_values[i - 1] is None:
            continue

        num_quarters_used = to_int_if_possible(num_quarters_raw, fallback=i)
        last_quarter_used = last_quarter_raw if not is_blank(last_quarter_raw) else latest_last_quarter

        avg_penetration_pct = as_number(avg_pen_raw)
        if avg_penetration_pct is None:
            avg_penetration_pct = avg_pen_formula_values[i - 1]

        quarterly_sales = as_number(q_sales_raw)
        if quarterly_sales is None:
            quarterly_sales = latest_quarterly

        reported_sales = as_number(actual_raw)
        if reported_sales is None:
            reported_sales = latest_reported

        forecast_value = as_number(forecast_raw)
        if forecast_value is None and quarterly_sales is not None and avg_penetration_pct not in (None, 0):
            ratio = avg_penetration_pct / 100.0 if avg_penetration_pct > 1 else avg_penetration_pct
            if ratio != 0:
                forecast_value = quarterly_sales / ratio

        forecast_max = as_number(max_raw)
        forecast_min = as_number(min_raw)
        if forecast_max is None and forecast_value is not None:
            forecast_max = forecast_value
        if forecast_min is None and forecast_value is not None:
            forecast_min = forecast_value

        range_width = None
        if forecast_max is not None and forecast_min is not None:
            range_width = forecast_max - forecast_min

        growth_rate_pct = as_number(growth_raw)
        if growth_rate_pct is None:
            growth_rate_pct = latest_growth

        sales_captured_in_db_pct = as_number(captured_raw)
        if sales_captured_in_db_pct is None:
            sales_captured_in_db_pct = latest_captured

        row = {
            "model": metadata["model"],
            "ticker": metadata["ticker"],
            "model_period": metadata["model_period"],
            "model_date": metadata["model_date"],
            "method": "empirical",
            "parameter_name": "avg_penetration_pct",
            "parameter_value": avg_penetration_pct,
            "num_quarters_used": num_quarters_used,
            "last_quarter_used": last_quarter_used,
            "forecast_value": forecast_value,
            "actual_value": reported_sales,
            "forecast_max": forecast_max,
            "forecast_min": forecast_min,
            "range_width": range_width,
            "avg_penetration_pct": avg_penetration_pct,
            "quarterly_sales": quarterly_sales,
            "reported_sales": reported_sales,
            "growth_rate_pct": growth_rate_pct,
            "sales_captured_in_db_pct": sales_captured_in_db_pct,
            "source_file": source_file,
        }
        rows.append(row)

    return rows


def values_close(a: Any, b: Any, tol: float = 1e-9) -> bool:
    num_a = as_number(a)
    num_b = as_number(b)
    if num_a is not None and num_b is not None:
        scale = max(1.0, abs(num_a), abs(num_b))
        return abs(num_a - num_b) <= tol * scale
    if is_blank(a) and is_blank(b):
        return True
    return a == b


def is_duplicate_regression_row(prev: dict[str, Any], curr: dict[str, Any]) -> bool:
    keys = ["num_quarters_used", "intercept", "slope", "forecast_value", "forecast_max", "forecast_min"]
    return all(values_close(prev.get(key), curr.get(key)) for key in keys)


def extract_regression_rows(wb: xw.Book, metadata: dict[str, str], source_file: str) -> list[dict[str, Any]]:
    sheet = get_sheet_by_name(wb, "Regression Model")
    if sheet is None:
        print(f"  skipped regression: sheet 'Regression Model' not found ({source_file})")
        return []

    base_row, base_col, values = get_used_grid(sheet)
    anchor = find_anchor(values, base_row, base_col, target="max")
    if anchor is None:
        print(f"  skipped regression: 'max' anchor not found ({source_file})")
        return []
    anchor_row, anchor_col = anchor

    # Fixed anchor-based offsets from requirement.
    y_col = anchor_col - 7
    x_col = anchor_col - 11
    if x_col < 1 or y_col < 1:
        print(f"  skipped regression: invalid x/y columns from anchor ({source_file})")
        return []

    x_series = get_numeric_series(sheet, x_col, 1, anchor_row - 1)
    y_series = get_numeric_series(sheet, y_col, 1, anchor_row - 1)
    x_map = {row: value for row, value in x_series}
    y_map = {row: value for row, value in y_series}
    common_rows = sorted(set(x_map) & set(y_map))
    xy_rows = [(row, x_map[row], y_map[row]) for row in common_rows]
    if len(xy_rows) < 2:
        print(f"  skipped regression: not enough x/y points ({source_file})")
        return []

    header_offsets = collect_header_offsets(values, base_row, base_col, anchor_row, anchor_col)
    min_offset = choose_offset(header_offsets, ["min"], default=-1)
    num_quarters_offset = choose_offset(header_offsets, ["num_quarters_used", "num_quarters", "quarters_used", "num_qtrs"])
    forecast_offset = choose_offset(
        header_offsets,
        ["tot_fcst_w_o_sa", "tot_fcst_wo_sa", "forecast_total_without_sa", "without_sa", "forecast_value"],
    )
    actual_offset = choose_offset(header_offsets, ["actual_value", "actual", "reported_sales", "reported"])

    n_rows = 10
    scratch_col = max(anchor_col + 25, max_used_col(values, base_col) + 2)
    scratch_start_row = anchor_row + 1

    n_use_by_row: list[int] = []
    for i in range(n_rows):
        table_row = anchor_row + 1 + i
        num_q_raw = read_by_offset(sheet, table_row, anchor_col, num_quarters_offset)
        requested = to_int_if_possible(num_q_raw, fallback=i + 1)
        requested = max(2, min(len(xy_rows), requested))
        n_use_by_row.append(requested)

        start_row = xy_rows[-requested][0]
        end_row = xy_rows[-1][0]
        output_row = scratch_start_row + i

        intercept_formula = (
            f"=INTERCEPT(R{start_row}C{y_col}:R{end_row}C{y_col},"
            f"R{start_row}C{x_col}:R{end_row}C{x_col})"
        )
        slope_formula = (
            f"=SLOPE(R{start_row}C{y_col}:R{end_row}C{y_col},"
            f"R{start_row}C{x_col}:R{end_row}C{x_col})"
        )
        next_x = xy_rows[-1][1] + 1.0
        forecast_formula = f"=R{output_row}C{scratch_col}+R{output_row}C{scratch_col + 1}*{next_x}"

        set_formula2(sheet.range((output_row, scratch_col)), intercept_formula)
        set_formula2(sheet.range((output_row, scratch_col + 1)), slope_formula)
        set_formula2(sheet.range((output_row, scratch_col + 2)), forecast_formula)

    wb.app.calculate()

    intercept_vals = sheet.range(
        (scratch_start_row, scratch_col),
        (scratch_start_row + n_rows - 1, scratch_col),
    ).options(ndim=1).value
    slope_vals = sheet.range(
        (scratch_start_row, scratch_col + 1),
        (scratch_start_row + n_rows - 1, scratch_col + 1),
    ).options(ndim=1).value
    forecast_calc_vals = sheet.range(
        (scratch_start_row, scratch_col + 2),
        (scratch_start_row + n_rows - 1, scratch_col + 2),
    ).options(ndim=1).value

    if not isinstance(intercept_vals, list):
        intercept_vals = [intercept_vals]
    if not isinstance(slope_vals, list):
        slope_vals = [slope_vals]
    if not isinstance(forecast_calc_vals, list):
        forecast_calc_vals = [forecast_calc_vals]

    rows: list[dict[str, Any]] = []
    for i in range(n_rows):
        table_row = anchor_row + 1 + i
        num_q_used = n_use_by_row[i]
        forecast_table = as_number(read_by_offset(sheet, table_row, anchor_col, forecast_offset))
        actual_value = read_by_offset(sheet, table_row, anchor_col, actual_offset)
        if is_blank(actual_value):
            actual_value = None

        forecast_max = as_number(read_by_offset(sheet, table_row, anchor_col, 0))
        forecast_min = as_number(read_by_offset(sheet, table_row, anchor_col, min_offset))
        intercept = as_number(intercept_vals[i]) if i < len(intercept_vals) else None
        slope = as_number(slope_vals[i]) if i < len(slope_vals) else None
        forecast_calc = as_number(forecast_calc_vals[i]) if i < len(forecast_calc_vals) else None

        forecast_value = forecast_table if forecast_table is not None else forecast_calc
        range_width = None
        if forecast_max is not None and forecast_min is not None:
            range_width = forecast_max - forecast_min

        probe = [forecast_value, forecast_max, forecast_min, intercept, slope]
        if all(is_blank(value) for value in probe):
            continue

        row = {
            "model": metadata["model"],
            "ticker": metadata["ticker"],
            "model_period": metadata["model_period"],
            "model_date": metadata["model_date"],
            "method": "regression",
            "parameter_name": "num_quarters_used",
            "parameter_value": num_q_used,
            "num_quarters_used": num_q_used,
            "forecast_value": forecast_value,
            "actual_value": actual_value,
            "forecast_max": forecast_max,
            "forecast_min": forecast_min,
            "range_width": range_width,
            "intercept": intercept,
            "slope": slope,
            "source_file": source_file,
        }

        if rows and is_duplicate_regression_row(rows[-1], row):
            continue
        rows.append(row)

    return rows


def write_sheet(ws, columns: list[str], rows: list[dict[str, Any]]) -> None:
    ws.append(columns)
    for row in rows:
        ws.append([row.get(col) for col in columns])

    for cell in ws[1]:
        cell.font = Font(bold=True)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}{max(ws.max_row, 1)}"

    for col_idx in range(1, ws.max_column + 1):
        col_letter = get_column_letter(col_idx)
        max_len = 0
        for row_idx in range(1, ws.max_row + 1):
            value = ws.cell(row=row_idx, column=col_idx).value
            if value is None:
                continue
            max_len = max(max_len, len(str(value)))
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 12), 48)


def save_output_workbook(
    empirical_rows: list[dict[str, Any]],
    regression_rows: list[dict[str, Any]],
    output_path: Path,
) -> None:
    wb_out = Workbook()
    ws_emp = wb_out.active
    ws_emp.title = "empirical_candidates"
    ws_reg = wb_out.create_sheet("regression_candidates")

    write_sheet(ws_emp, EMPIRICAL_COLUMNS, empirical_rows)
    write_sheet(ws_reg, REGRESSION_COLUMNS, regression_rows)

    wb_out.save(output_path)


def collect_source_files(input_path: Path) -> list[Path]:
    files: list[Path] = []
    for file_path in sorted(input_path.iterdir(), key=lambda p: p.name.lower()):
        if not file_path.is_file():
            continue
        if file_path.name.startswith("~"):
            print(f"skipped file: {file_path.name} (temporary file)")
            continue
        if file_path.suffix.lower() != ".xlsx":
            print(f"skipped file: {file_path.name} (not .xlsx)")
            continue
        files.append(file_path)
    return files


def main() -> None:
    input_path = Path(input_dir).expanduser().resolve()
    out_path = Path(output_dir).expanduser().resolve()
    out_path.mkdir(parents=True, exist_ok=True)

    if not input_path.exists() or not input_path.is_dir():
        raise FileNotFoundError(f"input_dir not found or not a folder: {input_path}")

    files = collect_source_files(input_path)

    empirical_rows: list[dict[str, Any]] = []
    regression_rows: list[dict[str, Any]] = []
    processed_files = 0

    app: xw.App | None = None
    try:
        app = xw.App(visible=False, add_book=False)
        app.display_alerts = False
        app.screen_updating = False
        try:
            app.api.EnableEvents = False
        except Exception:
            pass

        for file_path in files:
            print(f"processing file: {file_path.name}")
            wb: xw.Book | None = None
            try:
                # Open each source workbook once and process both model sheets while open.
                wb = app.books.open(str(file_path), update_links=False)
                metadata = parse_file_label(file_path.name)
                empirical_rows.extend(extract_empirical_rows(wb, metadata, file_path.name))
                regression_rows.extend(extract_regression_rows(wb, metadata, file_path.name))
                processed_files += 1
            except Exception as exc:
                print(f"skipped file: {file_path.name} (error: {exc})")
            finally:
                safe_close_workbook(wb)
    finally:
        if app is not None:
            try:
                app.quit()
            except Exception:
                pass

    output_file = next_output_path(input_path, out_path)
    save_output_workbook(empirical_rows, regression_rows, output_file)

    print(f"output path: {output_file}")
    print(f"files processed: {processed_files}")
    print(f"empirical rows: {len(empirical_rows)}")
    print(f"regression rows: {len(regression_rows)}")


if __name__ == "__main__":
    main()
