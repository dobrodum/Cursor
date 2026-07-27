#!/usr/bin/env python3
"""Extract empirical and regression parameter candidates from Excel models."""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
import re
from typing import Any, Dict, Iterable, List, Optional, Sequence, Tuple

import xlwings as xw
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.worksheet.worksheet import Worksheet


# ==========================
# User-configurable settings
# ==========================
input_dir = Path("/path/to/input")
output_dir = Path("/path/to/output")


EMPIRICAL_SHEET_NAME = "Empirical Model"
REGRESSION_SHEET_NAME = "Regression Model"
OUTPUT_EMPIRICAL_SHEET_NAME = "empirical_candidates"
OUTPUT_REGRESSION_SHEET_NAME = "regression_candidates"
N_QUARTERS = 10

EMPIRICAL_COLUMNS = [
    "model",
    "ticker",
    "model_period",
    "model_date",
    "method",
    "parameter_name",
    "parameter_value",
    "num_quarters_used",
    "last_quarter_used",
    "forecast_value",
    "actual_value",
    "forecast_max",
    "forecast_min",
    "range_width",
    "avg_penetration_pct",
    "quarterly_sales",
    "reported_sales",
    "growth_rate_pct",
    "sales_captured_in_db_pct",
    "source_file",
]

REGRESSION_COLUMNS = [
    "model",
    "ticker",
    "model_period",
    "model_date",
    "method",
    "parameter_name",
    "parameter_value",
    "num_quarters_used",
    "forecast_value",
    "actual_value",
    "forecast_max",
    "forecast_min",
    "range_width",
    "intercept",
    "slope",
    "source_file",
]


@dataclass(frozen=True)
class ModelLabel:
    model: str
    ticker: str
    model_period: str
    model_date: str


def normalize_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, str):
        text = value
    else:
        text = str(value)
    return re.sub(r"\s+", " ", text.strip().lower())


def is_blank(value: Any) -> bool:
    return value is None or (isinstance(value, str) and value.strip() == "")


def to_2d(values: Any, expected_rows: int) -> List[List[Any]]:
    if expected_rows <= 0:
        return []

    if expected_rows == 1:
        if isinstance(values, list):
            if values and isinstance(values[0], list):
                return values
            return [values]
        return [[values]]

    if not isinstance(values, list):
        return [[values]]

    if values and isinstance(values[0], list):
        return values

    return [[item] for item in values]


def to_1d(values: Any, expected_len: int) -> List[Any]:
    if expected_len <= 0:
        return []

    if not isinstance(values, list):
        return [values]

    if values and isinstance(values[0], list):
        return [row[0] if row else None for row in values]

    return values


def safe_subtract(left: Any, right: Any) -> Optional[float]:
    try:
        if is_blank(left) or is_blank(right):
            return None
        return float(left) - float(right)
    except (TypeError, ValueError):
        return None


def parse_model_label(file_name: str) -> Optional[ModelLabel]:
    stem = Path(file_name).stem
    parts = [segment.strip() for segment in stem.split(" - ") if segment.strip()]
    if len(parts) < 3:
        return None

    ticker = parts[1].upper()
    period_match = re.search(
        r"(Early|Mid|Late)(Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)(20\d{2})",
        parts[2],
        flags=re.IGNORECASE,
    )
    if not period_match:
        return None

    timing_raw, month_raw, year = period_match.groups()
    timing = timing_raw.capitalize()
    month = month_raw.capitalize()

    month_to_number = {
        "Jan": 1,
        "Feb": 2,
        "Mar": 3,
        "Apr": 4,
        "May": 5,
        "Jun": 6,
        "Jul": 7,
        "Aug": 8,
        "Sep": 9,
        "Oct": 10,
        "Nov": 11,
        "Dec": 12,
    }
    day_by_timing = {"Early": 5, "Mid": 15, "Late": 25}

    month_number = month_to_number[month]
    day = day_by_timing[timing]
    model_period = f"{timing}{month}_{year}"
    model_date = f"{year}-{month_number:02d}-{day:02d}"
    model = f"{ticker}_{model_period}"
    return ModelLabel(
        model=model,
        ticker=ticker,
        model_period=model_period,
        model_date=model_date,
    )


def resolve_output_path(source_input_dir: Path, target_output_dir: Path) -> Path:
    base_name = f"{source_input_dir.name}_PARAM"
    candidate = target_output_dir / f"{base_name}.xlsx"
    if not candidate.exists():
        return candidate

    suffix = 1
    while True:
        candidate = target_output_dir / f"{base_name}.{suffix}.xlsx"
        if not candidate.exists():
            return candidate
        suffix += 1


def close_workbook_safely(workbook: xw.Book) -> None:
    try:
        workbook.close(save=False)
        return
    except TypeError:
        pass
    except Exception:
        pass

    try:
        workbook.api.Close(SaveChanges=False)
        return
    except Exception:
        pass

    try:
        workbook.close()
    except Exception:
        print("Warning: workbook close fallback failed.")


def get_sheet_by_name(workbook: xw.Book, sheet_name: str) -> Optional[xw.Sheet]:
    for sheet in workbook.sheets:
        if sheet.name == sheet_name:
            return sheet
    return None


def find_anchor_cell(sheet: xw.Sheet, anchor_text: str = "max") -> Optional[Tuple[int, int]]:
    used = sheet.used_range
    values = used.value
    if values is None:
        return None

    matrix = to_2d(values, expected_rows=max(1, used.rows.count))
    start_row = used.row
    start_col = used.column
    normalized_anchor = normalize_text(anchor_text)

    for r_idx, row_values in enumerate(matrix):
        for c_idx, cell_value in enumerate(row_values):
            if normalize_text(cell_value) == normalized_anchor:
                return start_row + r_idx, start_col + c_idx
    return None


def map_header_offsets(
    sheet: xw.Sheet,
    anchor_row: int,
    anchor_col: int,
    col_span: int = 25,
) -> Dict[str, int]:
    left_col = max(1, anchor_col - col_span)
    right_col = anchor_col + col_span
    row_values = sheet.range((anchor_row, left_col), (anchor_row, right_col)).value
    if not isinstance(row_values, list):
        row_values = [row_values]

    mapping: Dict[str, int] = {}
    for idx, value in enumerate(row_values):
        header = normalize_text(value)
        if header and header not in mapping:
            absolute_col = left_col + idx
            mapping[header] = absolute_col - anchor_col
    return mapping


def pick_offset(
    header_offsets: Dict[str, int],
    keyword_groups: Sequence[Sequence[str]],
    fallback: int,
) -> int:
    for keyword_group in keyword_groups:
        candidates: List[Tuple[int, int]] = []
        for header, offset in header_offsets.items():
            if all(keyword in header for keyword in keyword_group):
                candidates.append((abs(offset), offset))
        if candidates:
            candidates.sort(key=lambda item: item[0])
            return candidates[0][1]
    return fallback


def read_offset_block(
    sheet: xw.Sheet,
    anchor_col: int,
    start_row: int,
    row_count: int,
    offsets: Iterable[int],
) -> Tuple[List[List[Any]], int]:
    offsets_list = sorted(set(offsets))
    min_offset = min(offsets_list)
    max_offset = max(offsets_list)

    start_col = max(1, anchor_col + min_offset)
    end_col = max(1, anchor_col + max_offset)
    range_values = sheet.range((start_row, start_col), (start_row + row_count - 1, end_col)).value
    matrix = to_2d(range_values, expected_rows=row_count)
    return matrix, start_col


def get_block_value(
    block_values: List[List[Any]],
    block_start_col: int,
    anchor_col: int,
    row_idx: int,
    col_offset: int,
) -> Any:
    if row_idx < 0 or row_idx >= len(block_values):
        return None
    row_values = block_values[row_idx]
    absolute_col = anchor_col + col_offset
    col_idx = absolute_col - block_start_col
    if col_idx < 0 or col_idx >= len(row_values):
        return None
    return row_values[col_idx]


def process_empirical_sheet(
    workbook: xw.Book,
    sheet: xw.Sheet,
    label: ModelLabel,
    source_file_name: str,
) -> List[Dict[str, Any]]:
    anchor = find_anchor_cell(sheet, anchor_text="max")
    if not anchor:
        print(f"Skipped empirical extraction for {source_file_name}: 'max' anchor not found.")
        return []

    anchor_row, anchor_col = anchor
    header_offsets = map_header_offsets(sheet, anchor_row=anchor_row, anchor_col=anchor_col)

    num_quarters_offset = pick_offset(
        header_offsets,
        keyword_groups=[("num", "quarter"), ("quarter", "used"), ("quarters", "used")],
        fallback=-10,
    )
    last_quarter_offset = pick_offset(
        header_offsets,
        keyword_groups=[("last", "quarter"), ("latest", "quarter")],
        fallback=-9,
    )
    forecast_value_offset = pick_offset(
        header_offsets,
        keyword_groups=[
            ("estimated", "total", "sold"),
            ("tot", "fcst"),
            ("forecast", "value"),
            ("forecast", "total"),
        ],
        fallback=-1,
    )
    actual_value_offset = pick_offset(
        header_offsets,
        keyword_groups=[("reported", "sales"), ("actual", "sales"), ("actual", "value")],
        fallback=-2,
    )
    forecast_min_offset = pick_offset(
        header_offsets,
        keyword_groups=[("min",)],
        fallback=1,
    )
    penetration_col_offset = pick_offset(
        header_offsets,
        keyword_groups=[("penetration",), ("avg", "penetration")],
        fallback=-4,
    )
    quarterly_sales_offset = pick_offset(
        header_offsets,
        keyword_groups=[("quarterly", "sales"), ("qtr", "sales")],
        fallback=-8,
    )
    growth_rate_offset = pick_offset(
        header_offsets,
        keyword_groups=[("growth",)],
        fallback=-6,
    )
    sales_captured_offset = pick_offset(
        header_offsets,
        keyword_groups=[("captured", "db"), ("sales", "captured"), ("captured", "database")],
        fallback=-5,
    )

    data_start_row = anchor_row + 1
    row_count = N_QUARTERS

    read_offsets = [
        num_quarters_offset,
        last_quarter_offset,
        forecast_value_offset,
        actual_value_offset,
        0,
        forecast_min_offset,
        quarterly_sales_offset,
        growth_rate_offset,
        sales_captured_offset,
        penetration_col_offset,
    ]
    block_values, block_start_col = read_offset_block(
        sheet,
        anchor_col=anchor_col,
        start_row=data_start_row,
        row_count=row_count,
        offsets=read_offsets,
    )

    scratch_start_col = max(sheet.used_range.last_cell.column + 2, anchor_col + max(read_offsets) + 2)
    scratch_range = sheet.range(
        (data_start_row, scratch_start_col),
        (data_start_row + row_count - 1, scratch_start_col),
    )

    avg_pen_formulas = []
    penetration_absolute_col = anchor_col + penetration_col_offset
    for idx in range(row_count):
        end_row = data_start_row + idx
        formula = (
            f'=IFERROR(AVERAGE(R{data_start_row}C{penetration_absolute_col}:'
            f'R{end_row}C{penetration_absolute_col}),"")'
        )
        avg_pen_formulas.append([formula])

    scratch_range.formula2 = avg_pen_formulas
    workbook.app.calculate()
    avg_penetration_values = to_1d(scratch_range.value, expected_len=row_count)
    scratch_range.clear_contents()

    rows: List[Dict[str, Any]] = []
    for idx in range(row_count):
        forecast_max = get_block_value(block_values, block_start_col, anchor_col, idx, 0)
        forecast_min = get_block_value(block_values, block_start_col, anchor_col, idx, forecast_min_offset)
        forecast_value = get_block_value(block_values, block_start_col, anchor_col, idx, forecast_value_offset)
        actual_value = get_block_value(block_values, block_start_col, anchor_col, idx, actual_value_offset)

        if is_blank(forecast_max) and is_blank(forecast_min) and is_blank(forecast_value) and is_blank(actual_value):
            continue

        num_quarters_used = get_block_value(block_values, block_start_col, anchor_col, idx, num_quarters_offset)
        if is_blank(num_quarters_used):
            num_quarters_used = idx + 1

        avg_penetration = avg_penetration_values[idx] if idx < len(avg_penetration_values) else None
        quarterly_sales = get_block_value(block_values, block_start_col, anchor_col, idx, quarterly_sales_offset)
        growth_rate = get_block_value(block_values, block_start_col, anchor_col, idx, growth_rate_offset)
        sales_captured = get_block_value(block_values, block_start_col, anchor_col, idx, sales_captured_offset)
        last_quarter_used = get_block_value(block_values, block_start_col, anchor_col, idx, last_quarter_offset)

        rows.append(
            {
                "model": label.model,
                "ticker": label.ticker,
                "model_period": label.model_period,
                "model_date": label.model_date,
                "method": "empirical",
                "parameter_name": "avg_penetration_pct",
                "parameter_value": avg_penetration,
                "num_quarters_used": num_quarters_used,
                "last_quarter_used": last_quarter_used,
                "forecast_value": forecast_value,
                "actual_value": actual_value,
                "forecast_max": forecast_max,
                "forecast_min": forecast_min,
                "range_width": safe_subtract(forecast_max, forecast_min),
                "avg_penetration_pct": avg_penetration,
                "quarterly_sales": quarterly_sales,
                "reported_sales": actual_value,
                "growth_rate_pct": growth_rate,
                "sales_captured_in_db_pct": sales_captured,
                "source_file": source_file_name,
            }
        )

    return rows


def process_regression_sheet(
    workbook: xw.Book,
    sheet: xw.Sheet,
    label: ModelLabel,
    source_file_name: str,
) -> List[Dict[str, Any]]:
    anchor = find_anchor_cell(sheet, anchor_text="max")
    if not anchor:
        print(f"Skipped regression extraction for {source_file_name}: 'max' anchor not found.")
        return []

    anchor_row, anchor_col = anchor
    header_offsets = map_header_offsets(sheet, anchor_row=anchor_row, anchor_col=anchor_col)

    num_quarters_offset = pick_offset(
        header_offsets,
        keyword_groups=[("num", "quarter"), ("quarter", "used"), ("quarters", "used")],
        fallback=-10,
    )
    forecast_total_offset = pick_offset(
        header_offsets,
        keyword_groups=[
            ("tot", "fcst", "w/o", "sa"),
            ("tot", "fcst", "wo", "sa"),
            ("forecast", "without", "sa"),
            ("tot", "fcst"),
        ],
        fallback=-1,
    )
    forecast_min_offset = pick_offset(
        header_offsets,
        keyword_groups=[("min",)],
        fallback=1,
    )
    actual_value_offset = pick_offset(
        header_offsets,
        keyword_groups=[("actual", "sales"), ("reported", "sales"), ("actual", "value")],
        fallback=-2,
    )

    y_col = anchor_col - 7
    x_col = anchor_col - 11
    data_start_row = anchor_row + 1
    row_count = N_QUARTERS

    read_offsets = [num_quarters_offset, forecast_total_offset, 0, forecast_min_offset, actual_value_offset]
    block_values, block_start_col = read_offset_block(
        sheet,
        anchor_col=anchor_col,
        start_row=data_start_row,
        row_count=row_count,
        offsets=read_offsets,
    )

    scratch_start_col = max(sheet.used_range.last_cell.column + 2, anchor_col + max(read_offsets) + 2)
    intercept_range = sheet.range(
        (data_start_row, scratch_start_col),
        (data_start_row + row_count - 1, scratch_start_col),
    )
    slope_range = sheet.range(
        (data_start_row, scratch_start_col + 1),
        (data_start_row + row_count - 1, scratch_start_col + 1),
    )

    intercept_formulas = []
    slope_formulas = []
    for idx in range(row_count):
        end_row = data_start_row + idx
        y_range = f"R{data_start_row}C{y_col}:R{end_row}C{y_col}"
        x_range = f"R{data_start_row}C{x_col}:R{end_row}C{x_col}"
        intercept_formulas.append([f'=IFERROR(INTERCEPT({y_range},{x_range}),"")'])
        slope_formulas.append([f'=IFERROR(SLOPE({y_range},{x_range}),"")'])

    intercept_range.formula2 = intercept_formulas
    slope_range.formula2 = slope_formulas
    workbook.app.calculate()

    intercept_values = to_1d(intercept_range.value, expected_len=row_count)
    slope_values = to_1d(slope_range.value, expected_len=row_count)

    intercept_range.clear_contents()
    slope_range.clear_contents()

    rows: List[Dict[str, Any]] = []
    previous_signature: Optional[Tuple[Any, ...]] = None
    for idx in range(row_count):
        forecast_max = get_block_value(block_values, block_start_col, anchor_col, idx, 0)
        forecast_min = get_block_value(block_values, block_start_col, anchor_col, idx, forecast_min_offset)
        forecast_value = get_block_value(block_values, block_start_col, anchor_col, idx, forecast_total_offset)
        actual_value = get_block_value(block_values, block_start_col, anchor_col, idx, actual_value_offset)
        num_quarters_used = get_block_value(block_values, block_start_col, anchor_col, idx, num_quarters_offset)

        if is_blank(num_quarters_used):
            num_quarters_used = idx + 1

        intercept = intercept_values[idx] if idx < len(intercept_values) else None
        slope = slope_values[idx] if idx < len(slope_values) else None

        if (
            is_blank(forecast_max)
            and is_blank(forecast_min)
            and is_blank(forecast_value)
            and is_blank(intercept)
            and is_blank(slope)
        ):
            continue

        signature = (
            forecast_value,
            forecast_max,
            forecast_min,
            intercept,
            slope,
        )
        if previous_signature is not None and signature == previous_signature:
            continue
        previous_signature = signature

        rows.append(
            {
                "model": label.model,
                "ticker": label.ticker,
                "model_period": label.model_period,
                "model_date": label.model_date,
                "method": "regression",
                "parameter_name": "num_quarters_used",
                "parameter_value": num_quarters_used,
                "num_quarters_used": num_quarters_used,
                "forecast_value": forecast_value,
                "actual_value": actual_value if not is_blank(actual_value) else None,
                "forecast_max": forecast_max,
                "forecast_min": forecast_min,
                "range_width": safe_subtract(forecast_max, forecast_min),
                "intercept": intercept,
                "slope": slope,
                "source_file": source_file_name,
            }
        )

    return rows


def write_table(
    worksheet: Worksheet,
    columns: Sequence[str],
    rows: Sequence[Dict[str, Any]],
) -> None:
    for col_idx, column_name in enumerate(columns, start=1):
        cell = worksheet.cell(row=1, column=col_idx, value=column_name)
        cell.font = Font(bold=True)

    for row_idx, row_data in enumerate(rows, start=2):
        for col_idx, column_name in enumerate(columns, start=1):
            worksheet.cell(row=row_idx, column=col_idx, value=row_data.get(column_name))

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions

    for col_idx, column_name in enumerate(columns, start=1):
        max_len = len(column_name)
        for row_data in rows:
            value = row_data.get(column_name)
            if value is None:
                continue
            text = str(value)
            if len(text) > max_len:
                max_len = len(text)
        worksheet.column_dimensions[worksheet.cell(row=1, column=col_idx).column_letter].width = min(
            50,
            max(12, max_len + 2),
        )


def write_output_workbook(
    output_path: Path,
    empirical_rows: Sequence[Dict[str, Any]],
    regression_rows: Sequence[Dict[str, Any]],
) -> None:
    output_wb = Workbook()
    empirical_ws = output_wb.active
    empirical_ws.title = OUTPUT_EMPIRICAL_SHEET_NAME
    write_table(empirical_ws, EMPIRICAL_COLUMNS, empirical_rows)

    regression_ws = output_wb.create_sheet(title=OUTPUT_REGRESSION_SHEET_NAME)
    write_table(regression_ws, REGRESSION_COLUMNS, regression_rows)

    output_wb.save(output_path)


def main() -> None:
    in_dir = Path(input_dir).expanduser().resolve()
    out_dir = Path(output_dir).expanduser().resolve()

    if not in_dir.exists() or not in_dir.is_dir():
        raise FileNotFoundError(f"Input directory not found: {in_dir}")
    out_dir.mkdir(parents=True, exist_ok=True)

    source_files: List[Tuple[Path, ModelLabel]] = []
    for file_path in sorted(in_dir.iterdir()):
        if not file_path.is_file():
            continue

        if file_path.suffix.lower() != ".xlsx":
            print(f"Skipped {file_path.name}: not an .xlsx file.")
            continue
        if file_path.name.startswith("~"):
            print(f"Skipped {file_path.name}: temporary workbook.")
            continue

        if file_path.stem.startswith(f"{in_dir.name}_PARAM"):
            print(f"Skipped {file_path.name}: appears to be a prior output workbook.")
            continue

        label = parse_model_label(file_path.name)
        if not label:
            print(f"Skipped {file_path.name}: unable to parse ticker/model period from filename.")
            continue

        source_files.append((file_path, label))

    empirical_rows: List[Dict[str, Any]] = []
    regression_rows: List[Dict[str, Any]] = []
    files_processed = 0

    app = xw.App(visible=False, add_book=False)
    app.display_alerts = False
    app.screen_updating = False
    app.calculation = "manual"

    try:
        for file_path, label in source_files:
            workbook: Optional[xw.Book] = None
            try:
                workbook = app.books.open(str(file_path), update_links=False)
                files_processed += 1
                print(f"Processed {file_path.name}")

                empirical_sheet = get_sheet_by_name(workbook, EMPIRICAL_SHEET_NAME)
                if empirical_sheet is None:
                    print(f"Skipped empirical extraction for {file_path.name}: sheet missing.")
                else:
                    empirical_rows.extend(
                        process_empirical_sheet(
                            workbook=workbook,
                            sheet=empirical_sheet,
                            label=label,
                            source_file_name=file_path.name,
                        )
                    )

                regression_sheet = get_sheet_by_name(workbook, REGRESSION_SHEET_NAME)
                if regression_sheet is None:
                    print(f"Skipped regression extraction for {file_path.name}: sheet missing.")
                else:
                    regression_rows.extend(
                        process_regression_sheet(
                            workbook=workbook,
                            sheet=regression_sheet,
                            label=label,
                            source_file_name=file_path.name,
                        )
                    )
            except Exception as exc:
                print(f"Skipped {file_path.name}: {exc}")
            finally:
                if workbook is not None:
                    close_workbook_safely(workbook)
    finally:
        app.quit()

    output_path = resolve_output_path(in_dir, out_dir)
    write_output_workbook(output_path, empirical_rows=empirical_rows, regression_rows=regression_rows)

    print(f"Output path: {output_path}")
    print(f"Number of files processed: {files_processed}")
    print(f"Number of empirical rows: {len(empirical_rows)}")
    print(f"Number of regression rows: {len(regression_rows)}")


if __name__ == "__main__":
    main()
