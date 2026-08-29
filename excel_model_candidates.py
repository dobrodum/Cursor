from __future__ import annotations

import re
from datetime import date, datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Sequence, Tuple

import xlwings as xw
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

# ----------------------------- USER CONFIG -----------------------------
input_dir = Path("input")
output_dir = Path("output")
# ----------------------------------------------------------------------

EMPIRICAL_HEADERS = [
    "model",
    "ticker",
    "model_period",
    "model_date",
    "method",
    "parameter_name",
    "parameter_value",
    "num_quarters_used",
    "last_quarter_used",
    "forecast_value",
    "actual_value",
    "forecast_max",
    "forecast_min",
    "range_width",
    "avg_penetration_pct",
    "quarterly_sales",
    "reported_sales",
    "growth_rate_pct",
    "sales_captured_in_db_pct",
    "source_file",
]

REGRESSION_HEADERS = [
    "model",
    "ticker",
    "model_period",
    "model_date",
    "method",
    "parameter_name",
    "parameter_value",
    "num_quarters_used",
    "forecast_value",
    "actual_value",
    "forecast_max",
    "forecast_min",
    "range_width",
    "intercept",
    "slope",
    "source_file",
]

DAY_BY_PERIOD = {"early": 5, "mid": 15, "late": 25}
MAX_QUARTERS = 10


def normalize_text(value: Any) -> str:
    if value is None:
        return ""
    cleaned = re.sub(r"[^a-z0-9]+", " ", str(value).strip().lower())
    return re.sub(r"\s+", " ", cleaned).strip()


def to_2d(values: Any) -> List[List[Any]]:
    if values is None:
        return []
    if not isinstance(values, list):
        return [[values]]
    if not values:
        return []
    if isinstance(values[0], list):
        return values
    return [values]


def to_float(value: Any) -> Optional[float]:
    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        cleaned = value.strip().replace(",", "").replace("%", "")
        if cleaned == "":
            return None
        try:
            return float(cleaned)
        except ValueError:
            return None
    return None


def to_int(value: Any) -> Optional[int]:
    num = to_float(value)
    if num is None:
        return None
    if abs(num - round(num)) < 1e-9:
        return int(round(num))
    return None


def get_used_values(sheet: xw.Sheet) -> Tuple[List[List[Any]], int, int]:
    used = sheet.used_range
    return to_2d(used.value), used.row, used.column


def matrix_value(
    values: List[List[Any]],
    start_row: int,
    start_col: int,
    row: int,
    col: int,
) -> Any:
    r_idx = row - start_row
    c_idx = col - start_col
    if r_idx < 0 or c_idx < 0 or r_idx >= len(values):
        return None
    row_values = values[r_idx]
    if c_idx >= len(row_values):
        return None
    return row_values[c_idx]


def find_anchor_cell(
    values: List[List[Any]],
    start_row: int,
    start_col: int,
    keyword: str = "max",
) -> Optional[Tuple[int, int]]:
    normalized_keyword = normalize_text(keyword)
    first_hit: Optional[Tuple[int, int]] = None

    for r_idx, row_values in enumerate(values):
        for c_idx, cell_value in enumerate(row_values):
            if normalize_text(cell_value) != normalized_keyword:
                continue

            abs_row = start_row + r_idx
            abs_col = start_col + c_idx
            if first_hit is None:
                first_hit = (abs_row, abs_col)

            right_text = ""
            if c_idx + 1 < len(row_values):
                right_text = normalize_text(row_values[c_idx + 1])
            if right_text == "min":
                return abs_row, abs_col

    return first_hit


def find_column_by_patterns(
    header_values: Sequence[Any],
    start_col: int,
    patterns: Sequence[Sequence[str]],
    default_col: Optional[int] = None,
) -> Optional[int]:
    for idx, raw_header in enumerate(header_values):
        header_text = normalize_text(raw_header)
        if not header_text:
            continue
        for tokens in patterns:
            if all(token in header_text for token in tokens):
                return start_col + idx
    return default_col


def build_num_quarters_row_map(
    values: List[List[Any]],
    start_row: int,
    start_col: int,
    first_data_row: int,
    num_quarters_col: Optional[int],
    max_quarters: int = MAX_QUARTERS,
) -> Dict[int, int]:
    if num_quarters_col is None:
        return {}

    row_map: Dict[int, int] = {}
    for row_offset, row_values in enumerate(values):
        abs_row = start_row + row_offset
        if abs_row < first_data_row:
            continue
        col_offset = num_quarters_col - start_col
        if col_offset < 0 or col_offset >= len(row_values):
            continue

        quarter_value = to_int(row_values[col_offset])
        if quarter_value is None:
            continue
        if 1 <= quarter_value <= max_quarters and quarter_value not in row_map:
            row_map[quarter_value] = abs_row

    return row_map


def output_path_for_run(source_dir: Path, destination_dir: Path) -> Path:
    base_name = f"{source_dir.name}_PARAM"
    candidate = destination_dir / f"{base_name}.xlsx"
    if not candidate.exists():
        return candidate

    suffix = 1
    while True:
        candidate = destination_dir / f"{base_name}.{suffix}.xlsx"
        if not candidate.exists():
            return candidate
        suffix += 1


def parse_file_metadata(file_name: str) -> Dict[str, str]:
    stem = Path(file_name).stem
    pattern = re.compile(
        r"-\s*([A-Za-z0-9]+)\s*-\s*(Early|Mid|Late)([A-Za-z]{3})(\d{4})",
        flags=re.IGNORECASE,
    )
    match = pattern.search(stem)

    if not match:
        fallback_ticker = re.sub(r"[^A-Za-z0-9]+", "", stem).upper()[:12] or "UNKNOWN"
        return {
            "model": f"{fallback_ticker}_Unknown",
            "ticker": fallback_ticker,
            "model_period": "Unknown",
            "model_date": "",
        }

    ticker = match.group(1).upper()
    period_name = match.group(2).title()
    month_abbrev = match.group(3).title()
    year = int(match.group(4))

    month_number = datetime.strptime(month_abbrev, "%b").month
    day_number = DAY_BY_PERIOD[period_name.lower()]

    model_period = f"{period_name}{month_abbrev}_{year}"
    model_date = date(year, month_number, day_number).isoformat()
    model = f"{ticker}_{model_period}"
    return {
        "model": model,
        "ticker": ticker,
        "model_period": model_period,
        "model_date": model_date,
    }


def maybe_numeric_or_text(value: Any) -> Any:
    numeric = to_float(value)
    if numeric is not None:
        return numeric
    if value is None:
        return None
    stripped = str(value).strip()
    return stripped if stripped else None


def safe_close_source_workbook(workbook: xw.Book) -> None:
    try:
        workbook.close(save=False)
        return
    except TypeError:
        pass
    except Exception:
        pass

    try:
        workbook.close(False)
        return
    except Exception:
        pass

    try:
        workbook.api.Close(SaveChanges=False)
    except Exception:
        # Final fallback: attempt a plain close. We still never call wb.save().
        workbook.close()


def extract_empirical_candidates(
    workbook: xw.Book,
    sheet: xw.Sheet,
    file_meta: Dict[str, str],
    source_file_name: str,
) -> List[Dict[str, Any]]:
    values, start_row, start_col = get_used_values(sheet)
    if not values:
        return []

    anchor = find_anchor_cell(values, start_row, start_col, keyword="max")
    if anchor is None:
        return []

    anchor_row, anchor_col = anchor
    header_row_idx = anchor_row - start_row
    header_values = values[header_row_idx] if 0 <= header_row_idx < len(values) else []

    col_num_quarters = find_column_by_patterns(
        header_values,
        start_col,
        patterns=[("num", "quarter"), ("n", "quarter"), ("quarters", "used")],
        default_col=anchor_col - 8,
    )
    col_last_quarter = find_column_by_patterns(
        header_values,
        start_col,
        patterns=[("last", "quarter"), ("quarter", "used")],
        default_col=anchor_col - 9,
    )
    col_forecast_value = find_column_by_patterns(
        header_values,
        start_col,
        patterns=[
            ("estimated", "total", "sold"),
            ("tot", "fcst"),
            ("forecast", "value"),
            ("total", "sold"),
        ],
        default_col=anchor_col - 3,
    )
    col_actual_value = find_column_by_patterns(
        header_values,
        start_col,
        patterns=[("reported", "sales"), ("actual", "sales"), ("actual",)],
        default_col=anchor_col - 2,
    )
    col_forecast_min = find_column_by_patterns(
        header_values,
        start_col,
        patterns=[("min",)],
        default_col=anchor_col + 1,
    )
    col_avg_penetration = find_column_by_patterns(
        header_values,
        start_col,
        patterns=[("avg", "penetration"), ("average", "penetration"), ("penetration", "pct")],
        default_col=anchor_col - 4,
    )
    col_quarterly_sales = find_column_by_patterns(
        header_values,
        start_col,
        patterns=[("quarterly", "sales"), ("quarter", "sales")],
        default_col=anchor_col - 11,
    )
    col_growth_rate = find_column_by_patterns(
        header_values,
        start_col,
        patterns=[("growth", "rate")],
        default_col=anchor_col - 6,
    )
    col_sales_captured = find_column_by_patterns(
        header_values,
        start_col,
        patterns=[
            ("sales", "captured", "db"),
            ("captured", "db"),
            ("captured", "in", "db"),
        ],
        default_col=col_avg_penetration,
    )

    offsets = {
        "num_quarters": col_num_quarters - anchor_col if col_num_quarters is not None else None,
        "last_quarter": col_last_quarter - anchor_col if col_last_quarter is not None else None,
        "forecast_value": col_forecast_value - anchor_col if col_forecast_value is not None else None,
        "actual_value": col_actual_value - anchor_col if col_actual_value is not None else None,
        "forecast_max": 0,
        "forecast_min": col_forecast_min - anchor_col if col_forecast_min is not None else 1,
        "avg_penetration": col_avg_penetration - anchor_col if col_avg_penetration is not None else None,
        "quarterly_sales": col_quarterly_sales - anchor_col if col_quarterly_sales is not None else None,
        "growth_rate": col_growth_rate - anchor_col if col_growth_rate is not None else None,
        "sales_captured": col_sales_captured - anchor_col if col_sales_captured is not None else None,
    }

    first_data_row = anchor_row + 1
    row_map = build_num_quarters_row_map(
        values=values,
        start_row=start_row,
        start_col=start_col,
        first_data_row=first_data_row,
        num_quarters_col=col_num_quarters,
        max_quarters=MAX_QUARTERS,
    )

    penetration_calc_col = col_sales_captured or col_avg_penetration
    scratch_avg_cells: Dict[int, Tuple[int, int]] = {}

    if penetration_calc_col is not None:
        scratch_col = anchor_col + 40
        for n_quarters in range(1, MAX_QUARTERS + 1):
            data_row = row_map.get(n_quarters, anchor_row + n_quarters)
            avg_start_row = max(first_data_row, data_row - n_quarters + 1)
            if avg_start_row > data_row:
                continue
            formula = (
                f"=AVERAGE(R{avg_start_row}C{penetration_calc_col}:"
                f"R{data_row}C{penetration_calc_col})"
            )
            scratch_row = anchor_row + n_quarters
            sheet.cells(scratch_row, scratch_col).formula2 = formula
            scratch_avg_cells[n_quarters] = (scratch_row, scratch_col)
        if scratch_avg_cells:
            workbook.app.calculate()

    rows: List[Dict[str, Any]] = []

    for n_quarters in range(1, MAX_QUARTERS + 1):
        data_row = row_map.get(n_quarters, anchor_row + n_quarters)

        def value_by_offset(offset_name: str) -> Any:
            offset = offsets[offset_name]
            if offset is None:
                return None
            return matrix_value(values, start_row, start_col, data_row, anchor_col + offset)

        num_quarters_used = to_int(value_by_offset("num_quarters")) or n_quarters
        last_quarter_used = maybe_numeric_or_text(value_by_offset("last_quarter"))
        forecast_value = to_float(value_by_offset("forecast_value"))
        actual_value = to_float(value_by_offset("actual_value"))
        forecast_max = to_float(value_by_offset("forecast_max"))
        forecast_min = to_float(value_by_offset("forecast_min"))
        avg_penetration_pct = to_float(value_by_offset("avg_penetration"))
        quarterly_sales = to_float(value_by_offset("quarterly_sales"))
        growth_rate_pct = to_float(value_by_offset("growth_rate"))
        sales_captured_pct = to_float(value_by_offset("sales_captured"))

        if avg_penetration_pct is None and n_quarters in scratch_avg_cells:
            avg_row, avg_col = scratch_avg_cells[n_quarters]
            avg_penetration_pct = to_float(sheet.cells(avg_row, avg_col).value)

        if sales_captured_pct is None:
            sales_captured_pct = avg_penetration_pct

        reported_sales = actual_value
        if forecast_value is None and reported_sales not in (None, 0) and avg_penetration_pct not in (None, 0):
            forecast_value = reported_sales / avg_penetration_pct

        if all(
            value is None
            for value in (
                forecast_value,
                actual_value,
                forecast_max,
                forecast_min,
                avg_penetration_pct,
            )
        ):
            continue

        range_width = (
            forecast_max - forecast_min
            if forecast_max is not None and forecast_min is not None
            else None
        )

        rows.append(
            {
                "model": file_meta["model"],
                "ticker": file_meta["ticker"],
                "model_period": file_meta["model_period"],
                "model_date": file_meta["model_date"],
                "method": "empirical",
                "parameter_name": "avg_penetration_pct",
                "parameter_value": avg_penetration_pct,
                "num_quarters_used": num_quarters_used,
                "last_quarter_used": last_quarter_used,
                "forecast_value": forecast_value,
                "actual_value": reported_sales,
                "forecast_max": forecast_max,
                "forecast_min": forecast_min,
                "range_width": range_width,
                "avg_penetration_pct": avg_penetration_pct,
                "quarterly_sales": quarterly_sales,
                "reported_sales": reported_sales,
                "growth_rate_pct": growth_rate_pct,
                "sales_captured_in_db_pct": sales_captured_pct,
                "source_file": source_file_name,
            }
        )

    return rows


def extract_regression_candidates(
    workbook: xw.Book,
    sheet: xw.Sheet,
    file_meta: Dict[str, str],
    source_file_name: str,
) -> List[Dict[str, Any]]:
    values, start_row, start_col = get_used_values(sheet)
    if not values:
        return []

    anchor = find_anchor_cell(values, start_row, start_col, keyword="max")
    if anchor is None:
        return []
    anchor_row, anchor_col = anchor

    header_row_idx = anchor_row - start_row
    header_values = values[header_row_idx] if 0 <= header_row_idx < len(values) else []

    # Required anchor-based source columns from the prompt.
    y_col = anchor_col - 7
    x_col = anchor_col - 11

    col_num_quarters = find_column_by_patterns(
        header_values,
        start_col,
        patterns=[("num", "quarter"), ("n", "quarter"), ("quarters", "used")],
        default_col=anchor_col - 6,
    )
    col_forecast_value = find_column_by_patterns(
        header_values,
        start_col,
        patterns=[
            ("tot", "fcst", "w", "o", "sa"),
            ("tot", "fcst"),
            ("forecast", "without", "sa"),
            ("forecast",),
        ],
        default_col=anchor_col - 1,
    )
    col_forecast_min = find_column_by_patterns(
        header_values,
        start_col,
        patterns=[("min",)],
        default_col=anchor_col + 1,
    )
    col_actual = find_column_by_patterns(
        header_values,
        start_col,
        patterns=[("actual", "sales"), ("reported", "sales"), ("actual",)],
        default_col=None,
    )

    offsets = {
        "num_quarters": col_num_quarters - anchor_col if col_num_quarters is not None else None,
        "forecast_value": col_forecast_value - anchor_col if col_forecast_value is not None else None,
        "actual": col_actual - anchor_col if col_actual is not None else None,
        "forecast_max": 0,
        "forecast_min": col_forecast_min - anchor_col if col_forecast_min is not None else 1,
    }

    first_data_row = anchor_row + 1
    row_map = build_num_quarters_row_map(
        values=values,
        start_row=start_row,
        start_col=start_col,
        first_data_row=first_data_row,
        num_quarters_col=col_num_quarters,
        max_quarters=MAX_QUARTERS,
    )

    numeric_rows: List[int] = []
    for row_offset in range(len(values)):
        abs_row = start_row + row_offset
        if abs_row < first_data_row:
            continue
        x_value = to_float(matrix_value(values, start_row, start_col, abs_row, x_col))
        y_value = to_float(matrix_value(values, start_row, start_col, abs_row, y_col))
        if x_value is not None and y_value is not None:
            numeric_rows.append(abs_row)

    if len(numeric_rows) < 2:
        return []

    n_limit = min(MAX_QUARTERS, len(numeric_rows))
    scratch_col = anchor_col + 50
    calc_cells: Dict[int, Tuple[int, int, int]] = {}

    for n_quarters in range(1, n_limit + 1):
        rows_for_n = numeric_rows[-n_quarters:]
        start_n = rows_for_n[0]
        end_n = rows_for_n[-1]

        out_row = anchor_row + n_quarters
        intercept_col = scratch_col
        slope_col = scratch_col + 1
        forecast_col = scratch_col + 2

        intercept_formula = (
            f"=INTERCEPT(R{start_n}C{y_col}:R{end_n}C{y_col},"
            f"R{start_n}C{x_col}:R{end_n}C{x_col})"
        )
        slope_formula = (
            f"=SLOPE(R{start_n}C{y_col}:R{end_n}C{y_col},"
            f"R{start_n}C{x_col}:R{end_n}C{x_col})"
        )
        forecast_formula = f"=R{out_row}C{intercept_col}+R{out_row}C{slope_col}*R{end_n}C{x_col}"

        sheet.cells(out_row, intercept_col).formula2 = intercept_formula
        sheet.cells(out_row, slope_col).formula2 = slope_formula
        sheet.cells(out_row, forecast_col).formula2 = forecast_formula

        calc_cells[n_quarters] = (out_row, intercept_col, slope_col)

    workbook.app.calculate()

    def value_by_offset(row: int, offset_name: str) -> Any:
        offset = offsets[offset_name]
        if offset is None:
            return None
        return matrix_value(values, start_row, start_col, row, anchor_col + offset)

    rows: List[Dict[str, Any]] = []
    previous_signature: Optional[Tuple[Any, ...]] = None

    for n_quarters in range(1, n_limit + 1):
        summary_row = row_map.get(n_quarters, anchor_row + n_quarters)
        out_row, intercept_col, slope_col = calc_cells[n_quarters]

        num_quarters_used = to_int(value_by_offset(summary_row, "num_quarters")) or n_quarters
        intercept = to_float(sheet.cells(out_row, intercept_col).value)
        slope = to_float(sheet.cells(out_row, slope_col).value)
        forecast_calculated = to_float(sheet.cells(out_row, slope_col + 1).value)

        forecast_value = to_float(value_by_offset(summary_row, "forecast_value"))
        if forecast_value is None:
            forecast_value = forecast_calculated

        actual_value = to_float(value_by_offset(summary_row, "actual"))
        forecast_max = to_float(value_by_offset(summary_row, "forecast_max"))
        forecast_min = to_float(value_by_offset(summary_row, "forecast_min"))
        range_width = (
            forecast_max - forecast_min
            if forecast_max is not None and forecast_min is not None
            else None
        )

        signature = (
            num_quarters_used,
            round(forecast_value, 10) if forecast_value is not None else None,
            round(forecast_max, 10) if forecast_max is not None else None,
            round(forecast_min, 10) if forecast_min is not None else None,
            round(intercept, 10) if intercept is not None else None,
            round(slope, 10) if slope is not None else None,
        )
        if previous_signature is not None and signature == previous_signature:
            continue
        previous_signature = signature

        if all(
            value is None
            for value in (forecast_value, forecast_max, forecast_min, intercept, slope)
        ):
            continue

        rows.append(
            {
                "model": file_meta["model"],
                "ticker": file_meta["ticker"],
                "model_period": file_meta["model_period"],
                "model_date": file_meta["model_date"],
                "method": "regression",
                "parameter_name": "num_quarters_used",
                "parameter_value": num_quarters_used,
                "num_quarters_used": num_quarters_used,
                "forecast_value": forecast_value,
                "actual_value": actual_value,
                "forecast_max": forecast_max,
                "forecast_min": forecast_min,
                "range_width": range_width,
                "intercept": intercept,
                "slope": slope,
                "source_file": source_file_name,
            }
        )

    return rows


def write_output_sheet(
    workbook: Workbook,
    sheet_name: str,
    headers: Sequence[str],
    rows: Sequence[Dict[str, Any]],
) -> None:
    sheet = workbook.create_sheet(title=sheet_name)

    for col_idx, header in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True)

    for row_idx, row in enumerate(rows, start=2):
        for col_idx, header in enumerate(headers, start=1):
            sheet.cell(row=row_idx, column=col_idx, value=row.get(header))

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions

    for col_idx, header in enumerate(headers, start=1):
        width = len(header) + 2
        for row_idx in range(2, sheet.max_row + 1):
            value = sheet.cell(row=row_idx, column=col_idx).value
            if value is None:
                continue
            width = max(width, len(str(value)) + 2)
        sheet.column_dimensions[get_column_letter(col_idx)].width = min(max(width, 12), 48)


def process_workbooks() -> None:
    source_dir = input_dir.resolve()
    destination_dir = output_dir.resolve()
    destination_dir.mkdir(parents=True, exist_ok=True)

    if not source_dir.exists():
        raise FileNotFoundError(f"input_dir does not exist: {source_dir}")

    empirical_rows: List[Dict[str, Any]] = []
    regression_rows: List[Dict[str, Any]] = []

    processed_files = 0
    app = xw.App(visible=False, add_book=False)
    app.display_alerts = False
    app.screen_updating = False
    app.api.EnableEvents = False

    original_calculation_mode = app.calculation
    app.calculation = "manual"

    try:
        for file_path in sorted(source_dir.iterdir()):
            if not file_path.is_file():
                continue
            if file_path.name.startswith("~"):
                print(f"Skipped file: {file_path.name} (temporary file)")
                continue
            if file_path.suffix.lower() != ".xlsx":
                print(f"Skipped file: {file_path.name} (not .xlsx)")
                continue

            print(f"Processing file: {file_path.name}")
            workbook: Optional[xw.Book] = None
            try:
                workbook = app.books.open(str(file_path), update_links=False)
                file_meta = parse_file_metadata(file_path.name)
                sheet_lookup = {sheet.name: sheet for sheet in workbook.sheets}

                empirical_sheet = sheet_lookup.get("Empirical Model")
                if empirical_sheet is None:
                    print(f"Skipped empirical extraction: {file_path.name} (missing 'Empirical Model')")
                else:
                    empirical_rows.extend(
                        extract_empirical_candidates(
                            workbook=workbook,
                            sheet=empirical_sheet,
                            file_meta=file_meta,
                            source_file_name=file_path.name,
                        )
                    )

                regression_sheet = sheet_lookup.get("Regression Model")
                if regression_sheet is None:
                    print(f"Skipped regression extraction: {file_path.name} (missing 'Regression Model')")
                else:
                    regression_rows.extend(
                        extract_regression_candidates(
                            workbook=workbook,
                            sheet=regression_sheet,
                            file_meta=file_meta,
                            source_file_name=file_path.name,
                        )
                    )

                processed_files += 1
            except Exception as exc:
                print(f"Skipped file: {file_path.name} (error: {exc})")
            finally:
                if workbook is not None:
                    safe_close_source_workbook(workbook)
    finally:
        app.calculation = original_calculation_mode
        app.quit()

    output_path = output_path_for_run(source_dir, destination_dir)
    output_wb = Workbook()
    default_sheet = output_wb.active
    output_wb.remove(default_sheet)

    write_output_sheet(
        workbook=output_wb,
        sheet_name="empirical_candidates",
        headers=EMPIRICAL_HEADERS,
        rows=empirical_rows,
    )
    write_output_sheet(
        workbook=output_wb,
        sheet_name="regression_candidates",
        headers=REGRESSION_HEADERS,
        rows=regression_rows,
    )

    output_wb.save(output_path)

    print(f"Output path: {output_path}")
    print(f"Number of files processed: {processed_files}")
    print(f"Number of empirical rows: {len(empirical_rows)}")
    print(f"Number of regression rows: {len(regression_rows)}")


if __name__ == "__main__":
    process_workbooks()
