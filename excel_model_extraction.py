#!/usr/bin/env python3
"""
Extract empirical/regression candidates from Excel model workbooks.

This script opens each source workbook once, processes both model sheets while
the workbook is open, and writes a single consolidated output workbook.
"""

from __future__ import annotations

import re
import sys
from dataclasses import dataclass
from datetime import date
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Sequence, Tuple

try:
    import xlwings as xw
except ImportError as exc:  # pragma: no cover - import guard
    raise SystemExit(
        "xlwings is required for this script. Install with: pip install xlwings openpyxl"
    ) from exc

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter


# ----------------------------
# User-configurable directories
# ----------------------------
input_dir = Path("input")
output_dir = Path("output")


EMPIRICAL_COLUMNS: List[str] = [
    "model",
    "ticker",
    "model_period",
    "model_date",
    "method",
    "parameter_name",
    "parameter_value",
    "num_quarters_used",
    "last_quarter_used",
    "forecast_value",
    "actual_value",
    "forecast_max",
    "forecast_min",
    "range_width",
    "avg_penetration_pct",
    "quarterly_sales",
    "reported_sales",
    "growth_rate_pct",
    "sales_captured_in_db_pct",
    "source_file",
]

REGRESSION_COLUMNS: List[str] = [
    "model",
    "ticker",
    "model_period",
    "model_date",
    "method",
    "parameter_name",
    "parameter_value",
    "num_quarters_used",
    "forecast_value",
    "actual_value",
    "forecast_max",
    "forecast_min",
    "range_width",
    "intercept",
    "slope",
    "source_file",
]


N_QUARTERS = 10
SHEET_EMPIRICAL = "Empirical Model"
SHEET_REGRESSION = "Regression Model"

# Early/Mid/Late model period day convention.
PERIOD_DAY = {"early": 5, "mid": 15, "late": 25}

MONTHS = {
    "jan": 1,
    "january": 1,
    "feb": 2,
    "february": 2,
    "mar": 3,
    "march": 3,
    "apr": 4,
    "april": 4,
    "may": 5,
    "jun": 6,
    "june": 6,
    "jul": 7,
    "july": 7,
    "aug": 8,
    "august": 8,
    "sep": 9,
    "sept": 9,
    "september": 9,
    "oct": 10,
    "october": 10,
    "nov": 11,
    "november": 11,
    "dec": 12,
    "december": 12,
}


@dataclass
class ModelMeta:
    model: str
    ticker: str
    model_period: str
    model_date: str


def parse_model_meta(file_name: str) -> Optional[ModelMeta]:
    """
    Parse filename metadata.

    Example:
        MedMiner_Model - AORT - MidJan2026_Send.xlsx
        -> ticker=AORT
        -> model_period=MidJan_2026
        -> model_date=2026-01-15
        -> model=AORT_MidJan_2026
    """
    stem = Path(file_name).stem
    parts = [part.strip() for part in stem.split(" - ")]

    ticker = ""
    if len(parts) >= 2:
        ticker = re.sub(r"[^A-Za-z0-9]", "", parts[1]).upper()
    if not ticker:
        ticker_match = re.search(r"-\s*([A-Za-z0-9]{1,10})\s*-", stem)
        if ticker_match:
            ticker = ticker_match.group(1).upper()
    if not ticker:
        return None

    period_match = re.search(
        r"(Early|Mid|Late)\s*([A-Za-z]{3,9})\s*([12]\d{3}|\d{2})",
        stem,
        flags=re.IGNORECASE,
    )
    if not period_match:
        return None

    period_bucket_raw, month_raw, year_raw = period_match.groups()
    period_bucket = period_bucket_raw.capitalize()

    month_key = month_raw.lower()
    if month_key not in MONTHS:
        month_key = month_key[:3]
    month_num = MONTHS.get(month_key)
    if month_num is None:
        return None
    month_abbrev = date(2000, month_num, 1).strftime("%b")

    year_num = int(year_raw)
    if year_num < 100:
        year_num += 2000

    day = PERIOD_DAY[period_bucket.lower()]
    model_dt = date(year_num, month_num, day).isoformat()
    model_period = f"{period_bucket}{month_abbrev}_{year_num}"
    model = f"{ticker}_{model_period}"
    return ModelMeta(model=model, ticker=ticker, model_period=model_period, model_date=model_dt)


def next_output_path(in_dir: Path, out_dir: Path) -> Path:
    out_dir.mkdir(parents=True, exist_ok=True)
    base = f"{in_dir.name}_PARAM"
    candidate = out_dir / f"{base}.xlsx"
    idx = 1
    while candidate.exists():
        candidate = out_dir / f"{base}.{idx}.xlsx"
        idx += 1
    return candidate


def normalize_float(value: Any) -> Optional[float]:
    if value is None or value == "":
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        raw = value.strip()
        if not raw:
            return None
        pct = raw.endswith("%")
        raw = raw.replace(",", "").replace("%", "")
        try:
            out = float(raw)
        except ValueError:
            return None
        return out / 100.0 if pct else out
    return None


def normalize_int(value: Any) -> Optional[int]:
    number = normalize_float(value)
    if number is None:
        return None
    return int(round(number))


def safe_cell_value(sheet: xw.Sheet, row: int, col: int) -> Any:
    if row < 1 or col < 1:
        return None
    try:
        return sheet.cells(row, col).value
    except Exception:
        return None


def set_formula2(cell: xw.Range, formula_r1c1: str) -> None:
    try:
        cell.formula2 = formula_r1c1
        return
    except Exception:
        pass

    try:
        cell.api.Formula2 = formula_r1c1
        return
    except Exception:
        pass

    # Last-resort fallback if Formula2 is unavailable in host Excel.
    cell.formula = formula_r1c1


def find_anchor_max(sheet: xw.Sheet) -> Optional[Tuple[int, int]]:
    """
    Find the first cell whose value is exactly "max" (case-insensitive)
    within used range.
    """
    used = sheet.used_range
    values = used.value
    if values is None:
        return None

    if not isinstance(values, list):
        values = [[values]]
    elif values and not isinstance(values[0], list):
        values = [values]

    base_row = used.row
    base_col = used.column
    for r_idx, row_values in enumerate(values):
        if not isinstance(row_values, list):
            row_values = [row_values]
        for c_idx, value in enumerate(row_values):
            if isinstance(value, str) and value.strip().lower() == "max":
                return base_row + r_idx, base_col + c_idx
    return None


def get_sheet_case_insensitive(wb: xw.Book, target_name: str) -> Optional[xw.Sheet]:
    target = target_name.strip().lower()
    for sheet in wb.sheets:
        if sheet.name.strip().lower() == target:
            return sheet
    return None


def close_workbook_safely(wb: xw.Book) -> None:
    """
    Close source workbook without saving. Uses a fallback for environments
    where close(save=False) is unsupported.
    """
    try:
        wb.close(save=False)
        return
    except TypeError:
        pass
    except Exception:
        pass

    try:
        wb.api.Close(SaveChanges=False)
        return
    except Exception:
        pass

    try:
        wb.close()
    except Exception:
        pass


def extract_empirical_rows(
    wb: xw.Book,
    sheet: xw.Sheet,
    meta: ModelMeta,
    source_file: str,
) -> List[Dict[str, Any]]:
    """
    Empirical extraction using max-anchor offsets.
    """
    anchor = find_anchor_max(sheet)
    if anchor is None:
        print(f"[SKIP] {source_file}: no 'max' anchor in '{sheet.name}'")
        return []

    anchor_row, anchor_col = anchor

    # Anchor-based offsets (kept explicit for easy maintenance).
    row_start_offset = 2
    num_quarters_col_offset = -11
    last_quarter_col_offset = -10
    quarterly_sales_col_offset = -7
    reported_sales_col_offset = -4
    estimated_total_col_offset = -5
    forecast_max_col_offset = 0
    forecast_min_col_offset = 1
    growth_rate_col_offset = -6
    captured_sales_col_offset = -3
    avg_pen_formula_col_offset = 2
    pen_history_end_col_offset = -1

    rows: List[Dict[str, Any]] = []
    formula_cells: List[xw.Range] = []
    calc_rows: List[int] = []

    for i in range(N_QUARTERS):
        n_quarters_default = i + 1
        row = anchor_row + row_start_offset + i
        num_quarters = normalize_int(
            safe_cell_value(sheet, row, anchor_col + num_quarters_col_offset)
        ) or n_quarters_default

        avg_formula_col = anchor_col + avg_pen_formula_col_offset
        pen_end_col = anchor_col + pen_history_end_col_offset
        pen_start_col = pen_end_col - (num_quarters - 1)

        if pen_start_col >= 1 and avg_formula_col >= 1:
            formula = (
                f'=IFERROR(AVERAGE(R{row}C{pen_start_col}:R{row}C{pen_end_col}),"")'
            )
            formula_cell = sheet.cells(row, avg_formula_col)
            set_formula2(formula_cell, formula)
            formula_cells.append(formula_cell)
            calc_rows.append(row)

    if formula_cells:
        wb.app.calculate()

    avg_pen_by_row: Dict[int, Optional[float]] = {}
    for row in calc_rows:
        avg_pen_val = normalize_float(
            safe_cell_value(sheet, row, anchor_col + avg_pen_formula_col_offset)
        )
        avg_pen_by_row[row] = avg_pen_val

    for i in range(N_QUARTERS):
        row = anchor_row + row_start_offset + i
        num_quarters = normalize_int(
            safe_cell_value(sheet, row, anchor_col + num_quarters_col_offset)
        ) or (i + 1)

        last_quarter_raw = safe_cell_value(sheet, row, anchor_col + last_quarter_col_offset)
        last_quarter = "" if last_quarter_raw in (None, "") else str(last_quarter_raw)

        quarterly_sales = normalize_float(
            safe_cell_value(sheet, row, anchor_col + quarterly_sales_col_offset)
        )
        reported_sales = normalize_float(
            safe_cell_value(sheet, row, anchor_col + reported_sales_col_offset)
        )
        estimated_total_sold = normalize_float(
            safe_cell_value(sheet, row, anchor_col + estimated_total_col_offset)
        )
        forecast_max = normalize_float(
            safe_cell_value(sheet, row, anchor_col + forecast_max_col_offset)
        )
        forecast_min = normalize_float(
            safe_cell_value(sheet, row, anchor_col + forecast_min_col_offset)
        )
        growth_rate_pct = normalize_float(
            safe_cell_value(sheet, row, anchor_col + growth_rate_col_offset)
        )
        sales_captured_pct = normalize_float(
            safe_cell_value(sheet, row, anchor_col + captured_sales_col_offset)
        )
        avg_penetration_pct = avg_pen_by_row.get(row)

        # If estimated total sold is blank but there is a non-zero penetration ratio,
        # derive an estimate from reported sales as a fallback.
        if (
            estimated_total_sold is None
            and reported_sales is not None
            and avg_penetration_pct not in (None, 0.0)
        ):
            estimated_total_sold = reported_sales / avg_penetration_pct

        if (
            estimated_total_sold is None
            and reported_sales is None
            and forecast_max is None
            and forecast_min is None
            and avg_penetration_pct is None
        ):
            continue

        range_width = None
        if forecast_max is not None and forecast_min is not None:
            range_width = forecast_max - forecast_min

        rows.append(
            {
                "model": meta.model,
                "ticker": meta.ticker,
                "model_period": meta.model_period,
                "model_date": meta.model_date,
                "method": "empirical",
                "parameter_name": "avg_penetration_pct",
                "parameter_value": avg_penetration_pct,
                "num_quarters_used": num_quarters,
                "last_quarter_used": last_quarter,
                "forecast_value": estimated_total_sold,
                "actual_value": reported_sales,
                "forecast_max": forecast_max,
                "forecast_min": forecast_min,
                "range_width": range_width,
                "avg_penetration_pct": avg_penetration_pct,
                "quarterly_sales": quarterly_sales,
                "reported_sales": reported_sales,
                "growth_rate_pct": growth_rate_pct,
                "sales_captured_in_db_pct": sales_captured_pct,
                "source_file": source_file,
            }
        )
    return rows


def _regression_row_signature(row: Dict[str, Any]) -> Tuple[Any, ...]:
    return (
        row.get("num_quarters_used"),
        row.get("forecast_value"),
        row.get("forecast_max"),
        row.get("forecast_min"),
        row.get("intercept"),
        row.get("slope"),
    )


def extract_regression_rows(
    wb: xw.Book,
    sheet: xw.Sheet,
    meta: ModelMeta,
    source_file: str,
) -> List[Dict[str, Any]]:
    """
    Regression extraction using max-anchor offsets and R1C1 Formula2 for
    INTERCEPT/SLOPE.
    """
    anchor = find_anchor_max(sheet)
    if anchor is None:
        print(f"[SKIP] {source_file}: no 'max' anchor in '{sheet.name}'")
        return []

    anchor_row, anchor_col = anchor
    y_col = anchor_col - 7
    x_col = anchor_col - 11

    row_start_offset = 2
    num_quarters_col_offset = -11
    forecast_total_col_offset = -5
    actual_value_col_offset = -4
    forecast_max_col_offset = 0
    forecast_min_col_offset = 1
    intercept_formula_col_offset = 2
    slope_formula_col_offset = 3

    data_end_row = anchor_row - 1
    if data_end_row < 1 or x_col < 1 or y_col < 1:
        return []

    pending: List[Tuple[int, int]] = []
    for i in range(N_QUARTERS):
        row = anchor_row + row_start_offset + i
        num_quarters = normalize_int(
            safe_cell_value(sheet, row, anchor_col + num_quarters_col_offset)
        ) or (i + 1)
        start_row = data_end_row - num_quarters + 1
        if start_row < 1:
            continue

        intercept_cell = sheet.cells(row, anchor_col + intercept_formula_col_offset)
        slope_cell = sheet.cells(row, anchor_col + slope_formula_col_offset)

        intercept_formula = (
            f'=IFERROR(INTERCEPT(R{start_row}C{y_col}:R{data_end_row}C{y_col},'
            f'R{start_row}C{x_col}:R{data_end_row}C{x_col}),"")'
        )
        slope_formula = (
            f'=IFERROR(SLOPE(R{start_row}C{y_col}:R{data_end_row}C{y_col},'
            f'R{start_row}C{x_col}:R{data_end_row}C{x_col}),"")'
        )
        set_formula2(intercept_cell, intercept_formula)
        set_formula2(slope_cell, slope_formula)
        pending.append((row, num_quarters))

    if pending:
        wb.app.calculate()

    rows: List[Dict[str, Any]] = []
    prev_sig: Optional[Tuple[Any, ...]] = None
    next_x = normalize_float(safe_cell_value(sheet, data_end_row + 1, x_col))

    for row, num_quarters in pending:
        intercept = normalize_float(
            safe_cell_value(sheet, row, anchor_col + intercept_formula_col_offset)
        )
        slope = normalize_float(safe_cell_value(sheet, row, anchor_col + slope_formula_col_offset))
        forecast_total_without_sa = normalize_float(
            safe_cell_value(sheet, row, anchor_col + forecast_total_col_offset)
        )
        if forecast_total_without_sa is None and intercept is not None and slope is not None and next_x is not None:
            forecast_total_without_sa = intercept + (slope * next_x)

        actual_value = normalize_float(
            safe_cell_value(sheet, row, anchor_col + actual_value_col_offset)
        )
        forecast_max = normalize_float(
            safe_cell_value(sheet, row, anchor_col + forecast_max_col_offset)
        )
        forecast_min = normalize_float(
            safe_cell_value(sheet, row, anchor_col + forecast_min_col_offset)
        )
        if (
            forecast_total_without_sa is None
            and forecast_max is None
            and forecast_min is None
            and intercept is None
            and slope is None
        ):
            continue

        range_width = None
        if forecast_max is not None and forecast_min is not None:
            range_width = forecast_max - forecast_min

        out = {
            "model": meta.model,
            "ticker": meta.ticker,
            "model_period": meta.model_period,
            "model_date": meta.model_date,
            "method": "regression",
            "parameter_name": "num_quarters_used",
            "parameter_value": num_quarters,
            "num_quarters_used": num_quarters,
            "forecast_value": forecast_total_without_sa,
            "actual_value": actual_value if actual_value is not None else "",
            "forecast_max": forecast_max,
            "forecast_min": forecast_min,
            "range_width": range_width,
            "intercept": intercept,
            "slope": slope,
            "source_file": source_file,
        }

        sig = _regression_row_signature(out)
        # Avoid duplicate trailing row produced by model formulas.
        if prev_sig is not None and sig == prev_sig:
            continue
        prev_sig = sig
        rows.append(out)
    return rows


def write_sheet(
    ws,
    columns: Sequence[str],
    rows: Iterable[Dict[str, Any]],
) -> None:
    ws.append(list(columns))
    for row in rows:
        ws.append([row.get(col, "") for col in columns])

    for header_cell in ws[1]:
        header_cell.font = Font(bold=True)

    ws.freeze_panes = "A2"
    max_row = max(ws.max_row, 1)
    max_col = len(columns)
    ws.auto_filter.ref = f"A1:{get_column_letter(max_col)}{max_row}"

    for col_idx, col_name in enumerate(columns, start=1):
        max_len = len(col_name)
        for row_idx in range(2, ws.max_row + 1):
            val = ws.cell(row=row_idx, column=col_idx).value
            if val is None:
                continue
            max_len = max(max_len, len(str(val)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(12, max_len + 2), 44)


def should_skip(path: Path) -> Optional[str]:
    if not path.is_file():
        return "not a file"
    if path.name.startswith("~"):
        return "temporary Excel file"
    if path.suffix.lower() != ".xlsx":
        return "not an .xlsx file"
    return None


def process_workbook(
    app: xw.App,
    file_path: Path,
    meta: ModelMeta,
) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]]]:
    wb = app.books.open(str(file_path), update_links=False)
    try:
        empirical_rows: List[Dict[str, Any]] = []
        regression_rows: List[Dict[str, Any]] = []

        empirical_sheet = get_sheet_case_insensitive(wb, SHEET_EMPIRICAL)
        regression_sheet = get_sheet_case_insensitive(wb, SHEET_REGRESSION)

        if empirical_sheet is None:
            print(f"[SKIP] {file_path.name}: missing sheet '{SHEET_EMPIRICAL}'")
        else:
            empirical_rows = extract_empirical_rows(
                wb=wb,
                sheet=empirical_sheet,
                meta=meta,
                source_file=file_path.name,
            )

        if regression_sheet is None:
            print(f"[SKIP] {file_path.name}: missing sheet '{SHEET_REGRESSION}'")
        else:
            regression_rows = extract_regression_rows(
                wb=wb,
                sheet=regression_sheet,
                meta=meta,
                source_file=file_path.name,
            )

        return empirical_rows, regression_rows
    finally:
        close_workbook_safely(wb)


def main() -> int:
    in_dir = Path(input_dir).expanduser().resolve()
    out_dir = Path(output_dir).expanduser().resolve()

    if not in_dir.exists():
        print(f"[ERROR] input directory not found: {in_dir}")
        return 1

    output_path = next_output_path(in_dir, out_dir)

    files_processed = 0
    empirical_rows_all: List[Dict[str, Any]] = []
    regression_rows_all: List[Dict[str, Any]] = []

    app: Optional[xw.App] = None
    try:
        app = xw.App(visible=False, add_book=False)
        app.display_alerts = False
        app.screen_updating = False

        for file_path in sorted(in_dir.iterdir(), key=lambda p: p.name.lower()):
            skip_reason = should_skip(file_path)
            if skip_reason:
                print(f"[SKIP] {file_path.name}: {skip_reason}")
                continue

            meta = parse_model_meta(file_path.name)
            if meta is None:
                print(f"[SKIP] {file_path.name}: could not parse ticker/model period from filename")
                continue

            try:
                empirical_rows, regression_rows = process_workbook(
                    app=app,
                    file_path=file_path,
                    meta=meta,
                )
            except Exception as exc:
                print(f"[SKIP] {file_path.name}: failed to process ({exc})")
                continue

            empirical_rows_all.extend(empirical_rows)
            regression_rows_all.extend(regression_rows)
            files_processed += 1
            print(
                f"[PROCESS] {file_path.name} | "
                f"empirical_rows={len(empirical_rows)} regression_rows={len(regression_rows)}"
            )
    finally:
        if app is not None:
            try:
                app.quit()
            except Exception:
                pass

    wb_out = Workbook()
    ws_empirical = wb_out.active
    ws_empirical.title = "empirical_candidates"
    ws_regression = wb_out.create_sheet("regression_candidates")

    write_sheet(ws_empirical, EMPIRICAL_COLUMNS, empirical_rows_all)
    write_sheet(ws_regression, REGRESSION_COLUMNS, regression_rows_all)

    wb_out.save(output_path)

    print(f"[OUTPUT] {output_path}")
    print(f"[SUMMARY] files_processed={files_processed}")
    print(f"[SUMMARY] empirical_rows={len(empirical_rows_all)}")
    print(f"[SUMMARY] regression_rows={len(regression_rows_all)}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
