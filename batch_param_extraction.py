#!/usr/bin/env python3
"""Batch extract empirical and regression model candidates from .xlsx files."""

from __future__ import annotations

import datetime as dt
import math
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Sequence

import openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

try:
    import xlwings as xw
except ImportError as exc:  # pragma: no cover
    raise SystemExit("xlwings is required. Install with: pip install xlwings") from exc


# =========================
# User-configured directories
# =========================
input_dir = r"./input"
output_dir = r"./output"


EMPIRICAL_HEADERS = [
    "model",
    "ticker",
    "model_period",
    "model_date",
    "method",
    "parameter_name",
    "parameter_value",
    "num_quarters_used",
    "last_quarter_used",
    "forecast_value",
    "actual_value",
    "forecast_max",
    "forecast_min",
    "range_width",
    "avg_penetration_pct",
    "quarterly_sales",
    "reported_sales",
    "growth_rate_pct",
    "sales_captured_in_db_pct",
    "source_file",
]

REGRESSION_HEADERS = [
    "model",
    "ticker",
    "model_period",
    "model_date",
    "method",
    "parameter_name",
    "parameter_value",
    "num_quarters_used",
    "forecast_value",
    "actual_value",
    "forecast_max",
    "forecast_min",
    "range_width",
    "intercept",
    "slope",
    "source_file",
]


EMPIRICAL_OFFSETS = {
    "num_quarters_used": -8,
    "last_quarter_used": -7,
    "avg_penetration_pct": -6,
    "quarterly_sales": -5,
    "reported_sales": -4,
    "growth_rate_pct": -3,
    "sales_captured_in_db_pct": -2,
    "forecast_value": -1,
    "forecast_max": 0,
    "forecast_min": 1,
}

REGRESSION_OFFSETS = {
    "num_quarters_used": -4,
    "actual_value": -2,
    "forecast_value": -1,  # TOT FCST w/o SA
    "forecast_max": 0,
    "forecast_min": 1,
}

N_QUARTERS = 10
REGRESSION_SCAN_LIMIT = 50

MODEL_DAY = {"early": 5, "mid": 15, "late": 25}
MONTH_NUM = {
    "jan": 1,
    "feb": 2,
    "mar": 3,
    "apr": 4,
    "may": 5,
    "jun": 6,
    "jul": 7,
    "aug": 8,
    "sep": 9,
    "oct": 10,
    "nov": 11,
    "dec": 12,
}


@dataclass
class ModelLabel:
    model: str
    ticker: str
    model_period: str
    model_date: str


def normalize_text(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip().lower()
    text = re.sub(r"[^a-z0-9]+", " ", text)
    return " ".join(text.split())


def to_float(value: Any) -> float | None:
    if value is None or value == "":
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        number = float(value)
        return number if math.isfinite(number) else None

    text = str(value).strip().replace(",", "")
    if not text:
        return None
    percent = text.endswith("%")
    if percent:
        text = text[:-1]

    try:
        number = float(text)
    except ValueError:
        return None
    if not math.isfinite(number):
        return None
    return number / 100.0 if percent else number


def to_int(value: Any) -> int | None:
    number = to_float(value)
    if number is None:
        return None
    return int(round(number))


def calc_range_width(max_value: Any, min_value: Any) -> float | None:
    max_num = to_float(max_value)
    min_num = to_float(min_value)
    if max_num is None or min_num is None:
        return None
    return max_num - min_num


def is_effectively_empty(values: Sequence[Any]) -> bool:
    for value in values:
        if value is None:
            continue
        if isinstance(value, str) and not value.strip():
            continue
        return False
    return True


def parse_model_label(file_name: str) -> ModelLabel:
    """Parse labels from names like: MedMiner_Model - AORT - MidJan2026_Send.xlsx"""
    stem = Path(file_name).stem
    parts = [part.strip() for part in stem.split(" - ")]

    ticker = "UNKNOWN"
    if len(parts) >= 2:
        ticker_candidate = re.sub(r"[^A-Za-z0-9]", "", parts[1]).upper()
        if ticker_candidate:
            ticker = ticker_candidate

    period_token = ""
    if len(parts) >= 3:
        period_token = re.split(r"[_\s-]", parts[2], maxsplit=1)[0]

    period_match = re.search(
        r"(Early|Mid|Late)([A-Za-z]{3,9})(20\d{2})",
        period_token or stem,
        flags=re.IGNORECASE,
    )

    if period_match:
        cadence = period_match.group(1).lower()
        month_token = period_match.group(2)[:3].lower()
        year = int(period_match.group(3))
        month_num = MONTH_NUM.get(month_token)
        if month_num is not None:
            cadence_title = cadence.capitalize()
            month_title = month_token.capitalize()
            model_period = f"{cadence_title}{month_title}_{year}"
            model_date = dt.date(year, month_num, MODEL_DAY[cadence]).isoformat()
        else:
            model_period = "Unknown_0000"
            model_date = ""
    else:
        model_period = "Unknown_0000"
        model_date = ""

    model = f"{ticker}_{model_period}"
    return ModelLabel(model=model, ticker=ticker, model_period=model_period, model_date=model_date)


def next_output_path(source_dir: Path, target_dir: Path) -> Path:
    base_name = f"{source_dir.name}_PARAM"
    candidate = target_dir / f"{base_name}.xlsx"
    if not candidate.exists():
        return candidate

    suffix = 1
    while True:
        candidate = target_dir / f"{base_name}.{suffix}.xlsx"
        if not candidate.exists():
            return candidate
        suffix += 1


def safe_close_workbook(wb: xw.Book) -> None:
    try:
        wb.close(save=False)
        return
    except TypeError:
        pass
    except Exception:
        pass

    try:
        wb.close(save_changes=False)
        return
    except TypeError:
        pass
    except Exception:
        pass

    try:
        wb.api.Close(SaveChanges=False)  # type: ignore[attr-defined]
        return
    except Exception:
        pass

    try:
        wb.close()
    except Exception:
        pass


def get_sheet_case_insensitive(wb: xw.Book, sheet_name: str) -> xw.Sheet | None:
    target = sheet_name.strip().lower()
    for sheet in wb.sheets:
        if sheet.name.strip().lower() == target:
            return sheet
    return None


def read_used_grid(ws: xw.Sheet) -> tuple[int, int, list[list[Any]]]:
    used = ws.used_range
    start_row = used.row
    start_col = used.column
    raw = used.value
    if raw is None:
        return start_row, start_col, []
    if isinstance(raw, list):
        if raw and isinstance(raw[0], list):
            return start_row, start_col, raw
        return start_row, start_col, [raw]
    return start_row, start_col, [[raw]]


def find_max_anchor(start_row: int, start_col: int, grid: Sequence[Sequence[Any]]) -> tuple[int, int] | None:
    for row_offset, row_values in enumerate(grid):
        for col_offset, value in enumerate(row_values):
            if normalize_text(value) == "max":
                return start_row + row_offset, start_col + col_offset
    return None


def set_formula2_r1c1(cell: xw.Range, formula_r1c1: str) -> None:
    try:
        cell.formula2 = formula_r1c1
        return
    except Exception:
        pass

    try:
        cell.api.Formula2R1C1 = formula_r1c1  # type: ignore[attr-defined]
        return
    except Exception:
        pass

    try:
        cell.api.FormulaR1C1 = formula_r1c1  # type: ignore[attr-defined]
        return
    except Exception:
        pass

    cell.formula = formula_r1c1


def build_empirical_rows(wb: xw.Book, ws: xw.Sheet, labels: ModelLabel, source_file: str) -> list[dict[str, Any]]:
    start_row, start_col, grid = read_used_grid(ws)
    anchor = find_max_anchor(start_row, start_col, grid)
    if anchor is None:
        print(f"Skipped {source_file} empirical extraction: 'max' anchor not found")
        return []

    anchor_row, anchor_col = anchor
    quarter_end_col = anchor_col - 11
    if quarter_end_col < 1:
        print(f"Skipped {source_file} empirical extraction: anchor offsets out of bounds")
        return []

    cols = {key: anchor_col + offset for key, offset in EMPIRICAL_OFFSETS.items()}
    if min(cols.values()) < 1:
        print(f"Skipped {source_file} empirical extraction: anchor offsets out of bounds")
        return []

    helper_col = max(anchor_col + 6, 1)

    formula_rows: list[int] = []
    for n_quarters in range(1, N_QUARTERS + 1):
        row = anchor_row + n_quarters
        quarter_start_col = max(1, quarter_end_col - (n_quarters - 1))
        formula = f'=IFERROR(AVERAGE(R{row}C{quarter_start_col}:R{row}C{quarter_end_col}),"")'
        set_formula2_r1c1(ws.range((row, helper_col)), formula)
        formula_rows.append(row)

    if formula_rows:
        wb.app.calculate()

    rows: list[dict[str, Any]] = []
    for n_quarters in range(1, N_QUARTERS + 1):
        row = anchor_row + n_quarters
        avg_pen_cell = ws.range((row, cols["avg_penetration_pct"])).value
        avg_pen_formula = ws.range((row, helper_col)).value
        forecast_value = ws.range((row, cols["forecast_value"])).value
        forecast_max = ws.range((row, cols["forecast_max"])).value
        forecast_min = ws.range((row, cols["forecast_min"])).value
        reported_sales = ws.range((row, cols["reported_sales"])).value

        if is_effectively_empty([avg_pen_cell, forecast_value, forecast_max, forecast_min, reported_sales]):
            continue

        avg_penetration = avg_pen_cell
        if avg_penetration in (None, ""):
            avg_penetration = avg_pen_formula

        num_quarters_used = ws.range((row, cols["num_quarters_used"])).value
        if num_quarters_used in (None, ""):
            num_quarters_used = n_quarters

        rows.append(
            {
                "model": labels.model,
                "ticker": labels.ticker,
                "model_period": labels.model_period,
                "model_date": labels.model_date,
                "method": "empirical",
                "parameter_name": "avg_penetration_pct",
                "parameter_value": avg_penetration,
                "num_quarters_used": num_quarters_used,
                "last_quarter_used": ws.range((row, cols["last_quarter_used"])).value,
                "forecast_value": forecast_value,
                "actual_value": reported_sales,
                "forecast_max": forecast_max,
                "forecast_min": forecast_min,
                "range_width": calc_range_width(forecast_max, forecast_min),
                "avg_penetration_pct": avg_penetration,
                "quarterly_sales": ws.range((row, cols["quarterly_sales"])).value,
                "reported_sales": reported_sales,
                "growth_rate_pct": ws.range((row, cols["growth_rate_pct"])).value,
                "sales_captured_in_db_pct": ws.range((row, cols["sales_captured_in_db_pct"])).value,
                "source_file": source_file,
            }
        )

    return rows


def build_regression_rows(wb: xw.Book, ws: xw.Sheet, labels: ModelLabel, source_file: str) -> list[dict[str, Any]]:
    start_row, start_col, grid = read_used_grid(ws)
    anchor = find_max_anchor(start_row, start_col, grid)
    if anchor is None:
        print(f"Skipped {source_file} regression extraction: 'max' anchor not found")
        return []

    anchor_row, anchor_col = anchor
    y_col = anchor_col - 7
    x_col = anchor_col - 11
    if min(y_col, x_col) < 1:
        print(f"Skipped {source_file} regression extraction: anchor offsets out of bounds")
        return []

    cols = {key: anchor_col + offset for key, offset in REGRESSION_OFFSETS.items()}
    if min(cols.values()) < 1:
        print(f"Skipped {source_file} regression extraction: anchor offsets out of bounds")
        return []

    intercept_col = anchor_col + 4
    slope_col = anchor_col + 5

    rows_to_calculate: list[tuple[int, int]] = []
    blank_streak = 0
    for row in range(anchor_row + 1, anchor_row + REGRESSION_SCAN_LIMIT + 1):
        num_quarters_used = ws.range((row, cols["num_quarters_used"])).value
        forecast_value = ws.range((row, cols["forecast_value"])).value
        forecast_max = ws.range((row, cols["forecast_max"])).value
        forecast_min = ws.range((row, cols["forecast_min"])).value

        if is_effectively_empty([num_quarters_used, forecast_value, forecast_max, forecast_min]):
            blank_streak += 1
            if blank_streak >= 2:
                break
            continue

        blank_streak = 0
        n_quarters = to_int(num_quarters_used) or max(2, row - anchor_row)
        n_quarters = max(2, n_quarters)
        fit_start_row = max(1, row - n_quarters + 1)

        intercept_formula = (
            f'=IFERROR(INTERCEPT(R{fit_start_row}C{y_col}:R{row}C{y_col},'
            f'R{fit_start_row}C{x_col}:R{row}C{x_col}),"")'
        )
        slope_formula = (
            f'=IFERROR(SLOPE(R{fit_start_row}C{y_col}:R{row}C{y_col},'
            f'R{fit_start_row}C{x_col}:R{row}C{x_col}),"")'
        )

        set_formula2_r1c1(ws.range((row, intercept_col)), intercept_formula)
        set_formula2_r1c1(ws.range((row, slope_col)), slope_formula)
        rows_to_calculate.append((row, n_quarters))

    if rows_to_calculate:
        wb.app.calculate()

    rows: list[dict[str, Any]] = []
    previous_signature: tuple[Any, ...] | None = None
    for row, fallback_n_quarters in rows_to_calculate:
        num_quarters_used = ws.range((row, cols["num_quarters_used"])).value
        if num_quarters_used in (None, ""):
            num_quarters_used = fallback_n_quarters

        forecast_value = ws.range((row, cols["forecast_value"])).value
        forecast_max = ws.range((row, cols["forecast_max"])).value
        forecast_min = ws.range((row, cols["forecast_min"])).value
        actual_value = ws.range((row, cols["actual_value"])).value
        intercept = ws.range((row, intercept_col)).value
        slope = ws.range((row, slope_col)).value

        signature = (
            to_int(num_quarters_used) or fallback_n_quarters,
            round(to_float(forecast_value) or 0.0, 10),
            round(to_float(forecast_max) or 0.0, 10),
            round(to_float(forecast_min) or 0.0, 10),
            round(to_float(intercept) or 0.0, 10),
            round(to_float(slope) or 0.0, 10),
        )
        if previous_signature is not None and signature == previous_signature:
            continue
        previous_signature = signature

        rows.append(
            {
                "model": labels.model,
                "ticker": labels.ticker,
                "model_period": labels.model_period,
                "model_date": labels.model_date,
                "method": "regression",
                "parameter_name": "num_quarters_used",
                "parameter_value": num_quarters_used,
                "num_quarters_used": num_quarters_used,
                "forecast_value": forecast_value,
                "actual_value": actual_value if actual_value not in ("", None) else None,
                "forecast_max": forecast_max,
                "forecast_min": forecast_min,
                "range_width": calc_range_width(forecast_max, forecast_min),
                "intercept": intercept,
                "slope": slope,
                "source_file": source_file,
            }
        )

    return rows


def write_sheet(
    wb: openpyxl.Workbook,
    sheet_name: str,
    headers: Sequence[str],
    rows: Sequence[dict[str, Any]],
) -> None:
    ws = wb.create_sheet(title=sheet_name)
    ws.append(list(headers))
    for row in rows:
        ws.append([row.get(header) for header in headers])

    for cell in ws[1]:
        cell.font = Font(bold=True)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    for col_idx, header in enumerate(headers, start=1):
        max_len = len(header)
        for row in rows:
            value = row.get(header)
            if value is None:
                continue
            max_len = max(max_len, len(str(value)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 12), 60)


def write_output_workbook(
    output_path: Path,
    empirical_rows: Sequence[dict[str, Any]],
    regression_rows: Sequence[dict[str, Any]],
) -> None:
    wb = openpyxl.Workbook()
    default_sheet = wb.active
    wb.remove(default_sheet)

    write_sheet(wb, "empirical_candidates", EMPIRICAL_HEADERS, empirical_rows)
    write_sheet(wb, "regression_candidates", REGRESSION_HEADERS, regression_rows)
    wb.save(output_path)


def list_input_files(source_dir: Path, output_name_pattern: re.Pattern[str]) -> list[Path]:
    files: list[Path] = []
    for path in sorted(source_dir.iterdir()):
        if not path.is_file():
            continue
        if path.name.startswith("~"):
            print(f"Skipped {path.name}: temporary file")
            continue
        if path.suffix.lower() != ".xlsx":
            print(f"Skipped {path.name}: not an .xlsx file")
            continue
        if output_name_pattern.match(path.name):
            print(f"Skipped {path.name}: output workbook pattern")
            continue
        files.append(path)
    return files


def main() -> None:
    source_dir = Path(input_dir).expanduser().resolve()
    target_dir = Path(output_dir).expanduser().resolve()

    if not source_dir.exists() or not source_dir.is_dir():
        raise SystemExit(f"Input directory does not exist: {source_dir}")

    target_dir.mkdir(parents=True, exist_ok=True)
    output_file = next_output_path(source_dir, target_dir)
    output_name_pattern = re.compile(
        rf"^{re.escape(source_dir.name)}_PARAM(?:\.\d+)?\.xlsx$",
        flags=re.IGNORECASE,
    )

    files = list_input_files(source_dir, output_name_pattern)
    empirical_rows: list[dict[str, Any]] = []
    regression_rows: list[dict[str, Any]] = []
    processed_files = 0

    app = xw.App(visible=False, add_book=False)
    app.display_alerts = False
    app.screen_updating = False
    try:
        app.calculation = "manual"
    except Exception:
        pass

    try:
        for file_path in files:
            print(f"Processed {file_path.name}")
            labels = parse_model_label(file_path.name)
            wb: xw.Book | None = None
            try:
                wb = app.books.open(str(file_path), update_links=False)

                empirical_sheet = get_sheet_case_insensitive(wb, "Empirical Model")
                if empirical_sheet is None:
                    print(f"Skipped {file_path.name} empirical extraction: missing 'Empirical Model' sheet")
                else:
                    empirical_rows.extend(build_empirical_rows(wb, empirical_sheet, labels, file_path.name))

                regression_sheet = get_sheet_case_insensitive(wb, "Regression Model")
                if regression_sheet is None:
                    print(f"Skipped {file_path.name} regression extraction: missing 'Regression Model' sheet")
                else:
                    regression_rows.extend(build_regression_rows(wb, regression_sheet, labels, file_path.name))

                processed_files += 1
            except Exception as exc:
                print(f"Skipped {file_path.name}: extraction failed ({exc})")
            finally:
                if wb is not None:
                    safe_close_workbook(wb)
    finally:
        app.quit()

    write_output_workbook(output_file, empirical_rows, regression_rows)
    print(f"Output path: {output_file}")
    print(f"Files processed: {processed_files}")
    print(f"Empirical rows: {len(empirical_rows)}")
    print(f"Regression rows: {len(regression_rows)}")


if __name__ == "__main__":
    main()
