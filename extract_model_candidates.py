from __future__ import annotations

import re
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import xlwings as xw
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

# Configure these before running.
input_dir = Path("/path/to/input")
output_dir = Path("/path/to/output")

N_QUARTERS = 10
MAX_EXCEL_COL = 16384

EMPIRICAL_SHEET_NAME = "Empirical Model"
REGRESSION_SHEET_NAME = "Regression Model"

# Empirical table layout is anchored on the cell whose value is "max".
# Offsets are relative to the anchor column; rows are anchor_row + n_quarters_used.
EMPIRICAL_ROW_START_OFFSET = 1
EMPIRICAL_OFFSETS = {
    "last_quarter_used": -8,
    "avg_penetration_fallback": -7,
    "quarterly_sales": -6,
    "reported_sales": -5,
    "growth_rate_pct": -4,
    "sales_captured_in_db_pct": -3,
    "actual_value_fallback": -2,
    "forecast_value": -1,
    "forecast_max": 0,
    "forecast_min": 1,
}
EMPIRICAL_PENETRATION_HISTORY_ROW_OFFSET = -2
EMPIRICAL_TEMP_FORMULA_COL_OFFSET = 8

# Regression layout is anchored on "max".
# Required by spec:
# y_col = anchor_col - 7
# x_col = anchor_col - 11
REGRESSION_ROW_START_OFFSET = 1
REGRESSION_OFFSETS = {
    "actual_value": -2,
    "forecast_total_without_sa": -1,
    "forecast_max": 0,
    "forecast_min": 1,
}
REGRESSION_TEMP_INTERCEPT_COL_OFFSET = 8
REGRESSION_TEMP_SLOPE_COL_OFFSET = 9
REGRESSION_FORECAST_X_ROW_OFFSET = 0

PHASE_DAY_MAP = {
    "early": 5,
    "mid": 15,
    "late": 25,
}

EMPIRICAL_COLUMNS = [
    "model",
    "ticker",
    "model_period",
    "model_date",
    "method",
    "parameter_name",
    "parameter_value",
    "num_quarters_used",
    "last_quarter_used",
    "forecast_value",
    "actual_value",
    "forecast_max",
    "forecast_min",
    "range_width",
    "avg_penetration_pct",
    "quarterly_sales",
    "reported_sales",
    "growth_rate_pct",
    "sales_captured_in_db_pct",
    "source_file",
]

REGRESSION_COLUMNS = [
    "model",
    "ticker",
    "model_period",
    "model_date",
    "method",
    "parameter_name",
    "parameter_value",
    "num_quarters_used",
    "forecast_value",
    "actual_value",
    "forecast_max",
    "forecast_min",
    "range_width",
    "intercept",
    "slope",
    "source_file",
]


@dataclass
class ModelMetadata:
    model: str
    ticker: str
    model_period: str
    model_date: str


def parse_model_metadata(file_name: str) -> Optional[ModelMetadata]:
    stem = Path(file_name).stem
    parts = [part.strip() for part in stem.split(" - ")]
    if len(parts) < 3:
        return None

    ticker = parts[1]
    period_segment = parts[2].split("_")[0].strip()
    match = re.search(
        r"(?i)^(Early|Mid|Late)(Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)(\d{4})$",
        period_segment,
    )
    if not match:
        return None

    phase_raw, month_raw, year_str = match.groups()
    phase = phase_raw.capitalize()
    month_abbrev = month_raw.capitalize()
    day = PHASE_DAY_MAP[phase.lower()]
    month_num = datetime.strptime(month_abbrev, "%b").month
    model_date = date(int(year_str), month_num, day).isoformat()
    model_period = f"{phase}{month_abbrev}_{year_str}"

    return ModelMetadata(
        model=f"{ticker}_{model_period}",
        ticker=ticker,
        model_period=model_period,
        model_date=model_date,
    )


def resolve_output_path(input_path: Path, output_path: Path) -> Path:
    base_name = f"{input_path.name}_PARAM"
    candidate = output_path / f"{base_name}.xlsx"
    suffix = 1
    while candidate.exists():
        candidate = output_path / f"{base_name}.{suffix}.xlsx"
        suffix += 1
    return candidate


def normalize_matrix(values: Any) -> List[List[Any]]:
    if values is None:
        return []
    if isinstance(values, list):
        if not values:
            return []
        if isinstance(values[0], list):
            return values
        return [values]
    return [[values]]


def find_max_anchor(sheet: xw.Sheet) -> Optional[Tuple[int, int]]:
    used_range = sheet.used_range
    matrix = normalize_matrix(used_range.value)
    if not matrix:
        return None

    base_row = used_range.row
    base_col = used_range.column
    for r_idx, row in enumerate(matrix):
        for c_idx, cell_value in enumerate(row):
            if isinstance(cell_value, str) and cell_value.strip().lower() == "max":
                return base_row + r_idx, base_col + c_idx
    return None


def safe_get(sheet: xw.Sheet, row: int, col: int) -> Any:
    if row < 1 or col < 1:
        return None
    return sheet.cells(row, col).value


def set_formula2_r1c1(cell: xw.Range, formula_r1c1: str) -> None:
    formula = formula_r1c1 if formula_r1c1.startswith("=") else f"={formula_r1c1}"
    try:
        cell.formula2 = formula
        return
    except Exception:
        pass

    try:
        cell.api.Formula2R1C1 = formula
        return
    except Exception:
        pass

    # Older Excel APIs can still evaluate FormulaR1C1.
    cell.api.FormulaR1C1 = formula


def as_float(value: Any) -> Optional[float]:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    try:
        return float(str(value).strip())
    except (TypeError, ValueError):
        return None


def subtract(a: Any, b: Any) -> Optional[float]:
    a_float = as_float(a)
    b_float = as_float(b)
    if a_float is None or b_float is None:
        return None
    return a_float - b_float


def close_workbook_safely(wb: xw.Book) -> None:
    try:
        wb.close(save=False)
        return
    except TypeError:
        pass
    except Exception:
        pass

    try:
        wb.api.Close(SaveChanges=False)
        return
    except Exception:
        pass

    try:
        wb.close()
    except Exception:
        pass


def extract_empirical_rows(
    wb: xw.Book,
    metadata: ModelMetadata,
    source_file: str,
) -> List[Dict[str, Any]]:
    if EMPIRICAL_SHEET_NAME not in [s.name for s in wb.sheets]:
        return []

    sheet = wb.sheets[EMPIRICAL_SHEET_NAME]
    anchor = find_max_anchor(sheet)
    if anchor is None:
        return []

    anchor_row, anchor_col = anchor
    history_row = anchor_row + EMPIRICAL_PENETRATION_HISTORY_ROW_OFFSET
    temp_col = min(anchor_col + EMPIRICAL_TEMP_FORMULA_COL_OFFSET, MAX_EXCEL_COL)
    temp_cell = sheet.cells(anchor_row, temp_col)

    rows: List[Dict[str, Any]] = []
    for n_quarters_used in range(1, N_QUARTERS + 1):
        data_row = anchor_row + EMPIRICAL_ROW_START_OFFSET + (n_quarters_used - 1)
        start_col = anchor_col - n_quarters_used
        end_col = anchor_col - 1
        if start_col < 1:
            break

        formula = f"=AVERAGE(R{history_row}C{start_col}:R{history_row}C{end_col})"
        set_formula2_r1c1(temp_cell, formula)
        wb.app.calculate()
        avg_penetration_pct = temp_cell.value

        fallback_avg_pen = safe_get(
            sheet,
            data_row,
            anchor_col + EMPIRICAL_OFFSETS["avg_penetration_fallback"],
        )
        if avg_penetration_pct is None:
            avg_penetration_pct = fallback_avg_pen

        last_quarter_used = safe_get(
            sheet,
            data_row,
            anchor_col + EMPIRICAL_OFFSETS["last_quarter_used"],
        )
        quarterly_sales = safe_get(
            sheet,
            data_row,
            anchor_col + EMPIRICAL_OFFSETS["quarterly_sales"],
        )
        reported_sales = safe_get(
            sheet,
            data_row,
            anchor_col + EMPIRICAL_OFFSETS["reported_sales"],
        )
        growth_rate_pct = safe_get(
            sheet,
            data_row,
            anchor_col + EMPIRICAL_OFFSETS["growth_rate_pct"],
        )
        sales_captured_in_db_pct = safe_get(
            sheet,
            data_row,
            anchor_col + EMPIRICAL_OFFSETS["sales_captured_in_db_pct"],
        )
        forecast_value = safe_get(
            sheet,
            data_row,
            anchor_col + EMPIRICAL_OFFSETS["forecast_value"],
        )
        forecast_max = safe_get(
            sheet,
            data_row,
            anchor_col + EMPIRICAL_OFFSETS["forecast_max"],
        )
        forecast_min = safe_get(
            sheet,
            data_row,
            anchor_col + EMPIRICAL_OFFSETS["forecast_min"],
        )
        actual_value = safe_get(
            sheet,
            data_row,
            anchor_col + EMPIRICAL_OFFSETS["actual_value_fallback"],
        )
        if actual_value is None:
            actual_value = reported_sales

        row = {
            "model": metadata.model,
            "ticker": metadata.ticker,
            "model_period": metadata.model_period,
            "model_date": metadata.model_date,
            "method": "empirical",
            "parameter_name": "avg_penetration_pct",
            "parameter_value": avg_penetration_pct,
            "num_quarters_used": n_quarters_used,
            "last_quarter_used": last_quarter_used,
            "forecast_value": forecast_value,
            "actual_value": actual_value,
            "forecast_max": forecast_max,
            "forecast_min": forecast_min,
            "range_width": subtract(forecast_max, forecast_min),
            "avg_penetration_pct": avg_penetration_pct,
            "quarterly_sales": quarterly_sales,
            "reported_sales": reported_sales,
            "growth_rate_pct": growth_rate_pct,
            "sales_captured_in_db_pct": sales_captured_in_db_pct,
            "source_file": source_file,
        }
        rows.append(row)

    return rows


def rounded_signature(values: Tuple[Any, ...]) -> Tuple[Any, ...]:
    signature: List[Any] = []
    for value in values:
        value_float = as_float(value)
        if value_float is None:
            signature.append(value)
        else:
            signature.append(round(value_float, 10))
    return tuple(signature)


def extract_regression_rows(
    wb: xw.Book,
    metadata: ModelMetadata,
    source_file: str,
) -> List[Dict[str, Any]]:
    if REGRESSION_SHEET_NAME not in [s.name for s in wb.sheets]:
        return []

    sheet = wb.sheets[REGRESSION_SHEET_NAME]
    anchor = find_max_anchor(sheet)
    if anchor is None:
        return []

    anchor_row, anchor_col = anchor
    y_col = anchor_col - 7
    x_col = anchor_col - 11
    if y_col < 1 or x_col < 1:
        return []

    intercept_col = min(anchor_col + REGRESSION_TEMP_INTERCEPT_COL_OFFSET, MAX_EXCEL_COL)
    slope_col = min(anchor_col + REGRESSION_TEMP_SLOPE_COL_OFFSET, MAX_EXCEL_COL)
    intercept_cell = sheet.cells(anchor_row, intercept_col)
    slope_cell = sheet.cells(anchor_row, slope_col)
    x_forecast_value = safe_get(sheet, anchor_row + REGRESSION_FORECAST_X_ROW_OFFSET, x_col)

    rows: List[Dict[str, Any]] = []
    previous_signature: Optional[Tuple[Any, ...]] = None
    for n_quarters_used in range(1, N_QUARTERS + 1):
        y_start_row = anchor_row - n_quarters_used
        y_end_row = anchor_row - 1
        if y_start_row < 1:
            break

        intercept_formula = (
            f"=INTERCEPT(R{y_start_row}C{y_col}:R{y_end_row}C{y_col},"
            f"R{y_start_row}C{x_col}:R{y_end_row}C{x_col})"
        )
        slope_formula = (
            f"=SLOPE(R{y_start_row}C{y_col}:R{y_end_row}C{y_col},"
            f"R{y_start_row}C{x_col}:R{y_end_row}C{x_col})"
        )
        set_formula2_r1c1(intercept_cell, intercept_formula)
        set_formula2_r1c1(slope_cell, slope_formula)
        wb.app.calculate()

        intercept = intercept_cell.value
        slope = slope_cell.value

        result_row = anchor_row + REGRESSION_ROW_START_OFFSET + (n_quarters_used - 1)
        forecast_total_without_sa = safe_get(
            sheet,
            result_row,
            anchor_col + REGRESSION_OFFSETS["forecast_total_without_sa"],
        )
        if forecast_total_without_sa is None:
            intercept_float = as_float(intercept)
            slope_float = as_float(slope)
            x_forecast_float = as_float(x_forecast_value)
            if (
                intercept_float is not None
                and slope_float is not None
                and x_forecast_float is not None
            ):
                forecast_total_without_sa = intercept_float + (slope_float * x_forecast_float)

        actual_value = safe_get(
            sheet,
            result_row,
            anchor_col + REGRESSION_OFFSETS["actual_value"],
        )
        forecast_max = safe_get(
            sheet,
            result_row,
            anchor_col + REGRESSION_OFFSETS["forecast_max"],
        )
        forecast_min = safe_get(
            sheet,
            result_row,
            anchor_col + REGRESSION_OFFSETS["forecast_min"],
        )

        signature = rounded_signature(
            (intercept, slope, forecast_total_without_sa, forecast_max, forecast_min)
        )
        if previous_signature is not None and signature == previous_signature:
            continue
        previous_signature = signature

        row = {
            "model": metadata.model,
            "ticker": metadata.ticker,
            "model_period": metadata.model_period,
            "model_date": metadata.model_date,
            "method": "regression",
            "parameter_name": "num_quarters_used",
            "parameter_value": n_quarters_used,
            "num_quarters_used": n_quarters_used,
            "forecast_value": forecast_total_without_sa,
            "actual_value": actual_value,
            "forecast_max": forecast_max,
            "forecast_min": forecast_min,
            "range_width": subtract(forecast_max, forecast_min),
            "intercept": intercept,
            "slope": slope,
            "source_file": source_file,
        }
        rows.append(row)

    return rows


def write_sheet(
    workbook: Workbook,
    sheet_name: str,
    columns: List[str],
    rows: List[Dict[str, Any]],
) -> None:
    ws = workbook.create_sheet(sheet_name)
    ws.append(columns)
    for column_idx in range(1, len(columns) + 1):
        ws.cell(row=1, column=column_idx).font = Font(bold=True)

    for row_data in rows:
        ws.append([row_data.get(column) for column in columns])

    ws.freeze_panes = "A2"
    last_col_letter = get_column_letter(len(columns))
    last_row = max(1, len(rows) + 1)
    ws.auto_filter.ref = f"A1:{last_col_letter}{last_row}"

    for column_idx, column_name in enumerate(columns, start=1):
        max_len = len(column_name)
        for row_data in rows:
            value = row_data.get(column_name)
            if value is None:
                continue
            max_len = max(max_len, len(str(value)))
        ws.column_dimensions[get_column_letter(column_idx)].width = min(max(12, max_len + 2), 42)


def write_output_workbook(
    output_path: Path,
    empirical_rows: List[Dict[str, Any]],
    regression_rows: List[Dict[str, Any]],
) -> None:
    wb = Workbook()
    default_sheet = wb.active
    wb.remove(default_sheet)

    write_sheet(wb, "empirical_candidates", EMPIRICAL_COLUMNS, empirical_rows)
    write_sheet(wb, "regression_candidates", REGRESSION_COLUMNS, regression_rows)
    wb.save(output_path)


def main() -> None:
    input_path = Path(input_dir).expanduser().resolve()
    output_path_root = Path(output_dir).expanduser().resolve()

    if not input_path.exists() or not input_path.is_dir():
        raise FileNotFoundError(f"Input folder does not exist or is not a directory: {input_path}")

    output_path_root.mkdir(parents=True, exist_ok=True)
    output_path = resolve_output_path(input_path, output_path_root)

    empirical_rows: List[Dict[str, Any]] = []
    regression_rows: List[Dict[str, Any]] = []
    files_processed = 0

    app: Optional[xw.App] = None
    prior_calculation: Any = None
    try:
        app = xw.App(visible=False, add_book=False)
        app.display_alerts = False
        app.screen_updating = False
        prior_calculation = app.calculation
        app.calculation = "manual"

        for file_path in sorted(input_path.iterdir()):
            if not file_path.is_file():
                continue
            if file_path.name.startswith("~"):
                print(f"Skipped {file_path.name}: temp file")
                continue
            if file_path.suffix.lower() != ".xlsx":
                print(f"Skipped {file_path.name}: not an .xlsx file")
                continue

            metadata = parse_model_metadata(file_path.name)
            if metadata is None:
                print(f"Skipped {file_path.name}: filename pattern not recognized")
                continue

            wb: Optional[xw.Book] = None
            try:
                wb = app.books.open(str(file_path), update_links=False)
                empirical_rows.extend(
                    extract_empirical_rows(wb=wb, metadata=metadata, source_file=file_path.name)
                )
                regression_rows.extend(
                    extract_regression_rows(wb=wb, metadata=metadata, source_file=file_path.name)
                )
                files_processed += 1
                print(f"Processed {file_path.name}")
            except Exception as exc:
                print(f"Skipped {file_path.name}: processing error ({exc})")
            finally:
                if wb is not None:
                    close_workbook_safely(wb)
    finally:
        if app is not None:
            if prior_calculation is not None:
                try:
                    app.calculation = prior_calculation
                except Exception:
                    pass
            app.quit()

    write_output_workbook(output_path, empirical_rows, regression_rows)

    print(f"Output path: {output_path}")
    print(f"Number of files processed: {files_processed}")
    print(f"Number of empirical rows: {len(empirical_rows)}")
    print(f"Number of regression rows: {len(regression_rows)}")


if __name__ == "__main__":
    main()
