from __future__ import annotations

import calendar
import re
from dataclasses import dataclass
from datetime import date
from pathlib import Path
from typing import Any, Iterable

import xlwings as xw
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter


# ---------------------------
# User-configurable paths
# ---------------------------
input_dir = Path("./input")
output_dir = Path("./output")


EMPIRICAL_COLUMNS = [
    "model",
    "ticker",
    "model_period",
    "model_date",
    "method",
    "parameter_name",
    "parameter_value",
    "num_quarters_used",
    "last_quarter_used",
    "forecast_value",
    "actual_value",
    "forecast_max",
    "forecast_min",
    "range_width",
    "avg_penetration_pct",
    "quarterly_sales",
    "reported_sales",
    "growth_rate_pct",
    "sales_captured_in_db_pct",
    "source_file",
]

REGRESSION_COLUMNS = [
    "model",
    "ticker",
    "model_period",
    "model_date",
    "method",
    "parameter_name",
    "parameter_value",
    "num_quarters_used",
    "forecast_value",
    "actual_value",
    "forecast_max",
    "forecast_min",
    "range_width",
    "intercept",
    "slope",
    "source_file",
]

N_QUARTERS = 10

EMPIRICAL_SYNONYMS = {
    "num_quarters_used": [
        "num quarters used",
        "quarters used",
        "n quarters",
        "num qtrs",
    ],
    "last_quarter_used": [
        "last quarter used",
        "last qtr used",
        "last quarter",
    ],
    "forecast_value": [
        "estimated total sold",
        "est total sold",
        "total sold estimate",
        "forecast value",
        "forecast",
    ],
    "reported_sales": [
        "reported sales",
        "actual sales",
        "actual value",
        "reported",
    ],
    "quarterly_sales": [
        "quarterly sales",
        "qtr sales",
        "quarter sales",
        "sales db",
    ],
    "growth_rate_pct": [
        "growth rate",
        "growth %",
    ],
    "sales_captured_in_db_pct": [
        "sales captured in db",
        "captured in db",
        "db capture",
    ],
    "avg_penetration_pct": [
        "avg penetration",
        "average penetration",
        "avg pen",
        "penetration %",
    ],
    "forecast_min": [
        "min",
    ],
}

REGRESSION_SYNONYMS = {
    "num_quarters_used": [
        "num quarters used",
        "quarters used",
        "n quarters",
        "num qtrs",
    ],
    "forecast_value": [
        "tot fcst w/o sa",
        "tot fcst without sa",
        "total forecast w/o sa",
        "fcst w/o sa",
        "forecast w/o sa",
    ],
    "forecast_min": [
        "min",
    ],
    "actual_value": [
        "actual",
        "actual value",
        "reported sales",
    ],
}


@dataclass
class FileMeta:
    model: str
    ticker: str
    model_period: str
    model_date: str


def normalize_2d(values: Any) -> list[list[Any]]:
    if values is None:
        return []
    if isinstance(values, list):
        if not values:
            return []
        if isinstance(values[0], list):
            return values
        return [values]
    return [[values]]


def normalize_text(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip().lower()
    text = re.sub(r"[^a-z0-9]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def as_float(value: Any) -> float | None:
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        cleaned = value.replace(",", "").replace("%", "").strip()
        if cleaned == "":
            return None
        try:
            return float(cleaned)
        except ValueError:
            return None
    return None


def safe_subtract(a: Any, b: Any) -> float | None:
    af = as_float(a)
    bf = as_float(b)
    if af is None or bf is None:
        return None
    return af - bf


def parse_filename_metadata(file_path: Path) -> FileMeta:
    stem = file_path.stem
    parts = [part.strip() for part in stem.split("-")]
    ticker = parts[1].strip().upper() if len(parts) >= 2 and parts[1].strip() else "UNKNOWN"
    period_segment = parts[2].strip() if len(parts) >= 3 else stem
    period_segment = period_segment.split("_")[0].strip()

    period_match = re.search(
        r"(?i)(early|mid|late)\s*([a-z]{3,9})\s*(\d{4})",
        period_segment,
    )

    model_period = period_segment.replace(" ", "")
    model_date = ""
    if period_match:
        timing = period_match.group(1).title()
        month_text = period_match.group(2).title()
        year = int(period_match.group(3))
        month_lookup = month_text[:3]
        month = {
            "Jan": 1,
            "Feb": 2,
            "Mar": 3,
            "Apr": 4,
            "May": 5,
            "Jun": 6,
            "Jul": 7,
            "Aug": 8,
            "Sep": 9,
            "Oct": 10,
            "Nov": 11,
            "Dec": 12,
        }.get(month_lookup)
        day = {"Early": 5, "Mid": 15, "Late": 25}[timing]

        if month is not None:
            last_day = calendar.monthrange(year, month)[1]
            safe_day = min(day, last_day)
            model_period = f"{timing}{month_lookup}_{year}"
            model_date = date(year, month, safe_day).isoformat()

    model = f"{ticker}_{model_period}"
    return FileMeta(model=model, ticker=ticker, model_period=model_period, model_date=model_date)


def find_anchor_max(
    matrix: list[list[Any]],
    base_row: int,
    base_col: int,
) -> tuple[int, int] | None:
    for r_idx, row in enumerate(matrix):
        for c_idx, value in enumerate(row):
            if normalize_text(value) == "max":
                return base_row + r_idx, base_col + c_idx
    return None


def build_header_row_map(
    matrix: list[list[Any]],
    base_row: int,
    base_col: int,
    row_number: int,
) -> dict[int, str]:
    idx = row_number - base_row
    if idx < 0 or idx >= len(matrix):
        return {}
    row_values = matrix[idx]
    header_map: dict[int, str] = {}
    for c_idx, value in enumerate(row_values):
        normalized = normalize_text(value)
        if normalized:
            header_map[base_col + c_idx] = normalized
    return header_map


def locate_column(
    header_maps: Iterable[dict[int, str]],
    synonyms: list[str],
    anchor_col: int,
) -> int | None:
    normalized_synonyms = [normalize_text(s) for s in synonyms]
    candidates: list[tuple[int, int]] = []
    for header_map in header_maps:
        for col, header in header_map.items():
            if any(s and s in header for s in normalized_synonyms):
                candidates.append((abs(col - anchor_col), col))
    if not candidates:
        return None
    candidates.sort()
    return candidates[0][1]


def close_workbook_without_saving(wb: xw.Book) -> None:
    try:
        wb.close(save=False)
        return
    except TypeError:
        pass
    except Exception:
        pass

    try:
        wb.api.Close(SaveChanges=False)
        return
    except Exception:
        pass

    try:
        wb.api.Saved = True
        wb.close()
    except Exception:
        # Last resort; ignore to keep batch run moving.
        pass


def first_available_output_path(input_folder: Path, target_dir: Path) -> Path:
    target_dir.mkdir(parents=True, exist_ok=True)
    base_name = f"{input_folder.name}_PARAM"
    candidate = target_dir / f"{base_name}.xlsx"
    if not candidate.exists():
        return candidate

    counter = 1
    while True:
        candidate = target_dir / f"{base_name}.{counter}.xlsx"
        if not candidate.exists():
            return candidate
        counter += 1


def write_formula_column(
    sheet: xw.Sheet,
    top_row: int,
    col: int,
    formulas: list[str | None],
) -> list[Any]:
    formula_values: list[list[Any]] = []
    for formula in formulas:
        formula_values.append([formula if formula else None])
    bottom_row = top_row + len(formula_values) - 1
    write_range = sheet.range((top_row, col), (bottom_row, col))
    write_range.formula2 = formula_values
    return normalize_2d(write_range.value)


def write_formula_block(
    sheet: xw.Sheet,
    top_row: int,
    left_col: int,
    formulas: list[list[str | None]],
) -> list[list[Any]]:
    bottom_row = top_row + len(formulas) - 1
    right_col = left_col + len(formulas[0]) - 1
    write_range = sheet.range((top_row, left_col), (bottom_row, right_col))
    write_range.formula2 = formulas
    return normalize_2d(write_range.value)


def read_block(
    sheet: xw.Sheet,
    top_row: int,
    bottom_row: int,
    left_col: int,
    right_col: int,
) -> list[list[Any]]:
    if left_col > right_col:
        return [[None] * 1 for _ in range(bottom_row - top_row + 1)]
    values = sheet.range((top_row, left_col), (bottom_row, right_col)).value
    return normalize_2d(values)


def block_value(
    block: list[list[Any]],
    row_number: int,
    col_number: int,
    top_row: int,
    left_col: int,
) -> Any:
    r = row_number - top_row
    c = col_number - left_col
    if r < 0 or c < 0:
        return None
    if r >= len(block):
        return None
    row = block[r]
    if c >= len(row):
        return None
    return row[c]


def build_empirical_rows(wb: xw.Book, meta: FileMeta, source_file: str) -> list[dict[str, Any]]:
    if "Empirical Model" not in [sht.name for sht in wb.sheets]:
        return []

    sheet = wb.sheets["Empirical Model"]
    used = sheet.used_range
    matrix = normalize_2d(used.value)
    if not matrix:
        return []

    base_row = used.row
    base_col = used.column
    anchor = find_anchor_max(matrix, base_row, base_col)
    if anchor is None:
        return []

    anchor_row, anchor_col = anchor
    header_maps = [
        build_header_row_map(matrix, base_row, base_col, anchor_row),
        build_header_row_map(matrix, base_row, base_col, anchor_row - 1),
    ]

    col_map: dict[str, int | None] = {
        "forecast_max": anchor_col,
    }
    for key, synonyms in EMPIRICAL_SYNONYMS.items():
        col_map[key] = locate_column(header_maps, synonyms, anchor_col)

    # Default min column if not explicitly found.
    if col_map["forecast_min"] is None:
        col_map["forecast_min"] = anchor_col + 1

    top_row = anchor_row + 1
    bottom_row = anchor_row + N_QUARTERS

    candidate_cols = [col for col in col_map.values() if isinstance(col, int)]
    if not candidate_cols:
        candidate_cols = [anchor_col]
    left_col = min(candidate_cols)
    right_col = max(candidate_cols)
    table_block = read_block(sheet, top_row, bottom_row, left_col, right_col)

    # Avg penetration via temporary R1C1 formulas (single calculate).
    last_hist_row = anchor_row - 1
    penetration_col = col_map["avg_penetration_pct"]
    scratch_col = used.last_cell.column + 2
    avg_pen_formulas: list[str | None] = []
    if isinstance(penetration_col, int):
        for n in range(1, N_QUARTERS + 1):
            start_row = last_hist_row - n + 1
            if start_row > 0:
                avg_pen_formulas.append(
                    f"=AVERAGE(R{start_row}C{penetration_col}:R{last_hist_row}C{penetration_col})"
                )
            else:
                avg_pen_formulas.append(None)
    else:
        avg_pen_formulas = [None] * N_QUARTERS

    if any(avg_pen_formulas):
        write_formula_column(sheet, top_row, scratch_col, avg_pen_formulas)
        wb.app.calculate()
        avg_pen_values = normalize_2d(sheet.range((top_row, scratch_col), (bottom_row, scratch_col)).value)
    else:
        avg_pen_values = [[None] for _ in range(N_QUARTERS)]

    rows: list[dict[str, Any]] = []
    for i in range(N_QUARTERS):
        row_num = top_row + i
        n_default = i + 1
        num_quarters_used = block_value(
            table_block, row_num, col_map["num_quarters_used"] or -1, top_row, left_col
        )
        if as_float(num_quarters_used) is None:
            num_quarters_used = n_default

        forecast_value = block_value(table_block, row_num, col_map["forecast_value"] or -1, top_row, left_col)
        forecast_max = block_value(table_block, row_num, col_map["forecast_max"] or -1, top_row, left_col)
        forecast_min = block_value(table_block, row_num, col_map["forecast_min"] or -1, top_row, left_col)
        reported_sales = block_value(table_block, row_num, col_map["reported_sales"] or -1, top_row, left_col)
        quarterly_sales = block_value(table_block, row_num, col_map["quarterly_sales"] or -1, top_row, left_col)
        growth_rate_pct = block_value(table_block, row_num, col_map["growth_rate_pct"] or -1, top_row, left_col)
        sales_captured_pct = block_value(
            table_block,
            row_num,
            col_map["sales_captured_in_db_pct"] or -1,
            top_row,
            left_col,
        )
        last_quarter_used = block_value(table_block, row_num, col_map["last_quarter_used"] or -1, top_row, left_col)

        sheet_avg_pen = block_value(table_block, row_num, col_map["avg_penetration_pct"] or -1, top_row, left_col)
        calc_avg_pen = avg_pen_values[i][0] if i < len(avg_pen_values) else None
        avg_penetration_pct = sheet_avg_pen if as_float(sheet_avg_pen) is not None else calc_avg_pen

        # Fallback forecast if sheet doesn't expose a forecast value column.
        if as_float(forecast_value) is None:
            q_sales = as_float(quarterly_sales)
            pen = as_float(avg_penetration_pct)
            if q_sales is not None and pen not in (None, 0.0):
                forecast_value = q_sales / pen

        key_presence = [
            as_float(forecast_value),
            as_float(forecast_max),
            as_float(forecast_min),
            as_float(avg_penetration_pct),
        ]
        if all(val is None for val in key_presence):
            continue

        rows.append(
            {
                "model": meta.model,
                "ticker": meta.ticker,
                "model_period": meta.model_period,
                "model_date": meta.model_date,
                "method": "empirical",
                "parameter_name": "avg_penetration_pct",
                "parameter_value": avg_penetration_pct,
                "num_quarters_used": num_quarters_used,
                "last_quarter_used": last_quarter_used,
                "forecast_value": forecast_value,
                "actual_value": reported_sales,
                "forecast_max": forecast_max,
                "forecast_min": forecast_min,
                "range_width": safe_subtract(forecast_max, forecast_min),
                "avg_penetration_pct": avg_penetration_pct,
                "quarterly_sales": quarterly_sales,
                "reported_sales": reported_sales,
                "growth_rate_pct": growth_rate_pct,
                "sales_captured_in_db_pct": sales_captured_pct,
                "source_file": source_file,
            }
        )

    return rows


def numeric_rows_for_xy(
    matrix: list[list[Any]],
    base_row: int,
    base_col: int,
    anchor_row: int,
    x_col: int,
    y_col: int,
) -> list[int]:
    rows: list[int] = []
    row_count = len(matrix)

    for abs_row in range(base_row, anchor_row):
        r_idx = abs_row - base_row
        x_idx = x_col - base_col
        y_idx = y_col - base_col
        if r_idx < 0 or r_idx >= row_count:
            continue
        row_values = matrix[r_idx]
        if not isinstance(row_values, list):
            continue
        if x_idx < 0 or y_idx < 0:
            continue
        if x_idx >= len(row_values) or y_idx >= len(row_values):
            continue
        x_val = row_values[x_idx]
        y_val = row_values[y_idx]
        if as_float(x_val) is not None and as_float(y_val) is not None:
            rows.append(abs_row)
    return rows


def dedupe_key(row: dict[str, Any]) -> tuple[Any, Any, Any, Any, Any]:
    def rounded(value: Any) -> Any:
        fv = as_float(value)
        if fv is None:
            return value
        return round(fv, 8)

    return (
        rounded(row.get("intercept")),
        rounded(row.get("slope")),
        rounded(row.get("forecast_value")),
        rounded(row.get("forecast_max")),
        rounded(row.get("forecast_min")),
    )


def build_regression_rows(wb: xw.Book, meta: FileMeta, source_file: str) -> list[dict[str, Any]]:
    if "Regression Model" not in [sht.name for sht in wb.sheets]:
        return []

    sheet = wb.sheets["Regression Model"]
    used = sheet.used_range
    matrix = normalize_2d(used.value)
    if not matrix:
        return []

    base_row = used.row
    base_col = used.column
    anchor = find_anchor_max(matrix, base_row, base_col)
    if anchor is None:
        return []

    anchor_row, anchor_col = anchor
    y_col = anchor_col - 7
    x_col = anchor_col - 11

    header_maps = [
        build_header_row_map(matrix, base_row, base_col, anchor_row),
        build_header_row_map(matrix, base_row, base_col, anchor_row - 1),
    ]

    col_map: dict[str, int | None] = {
        "forecast_max": anchor_col,
    }
    for key, synonyms in REGRESSION_SYNONYMS.items():
        col_map[key] = locate_column(header_maps, synonyms, anchor_col)
    if col_map["forecast_min"] is None:
        col_map["forecast_min"] = anchor_col + 1

    top_row = anchor_row + 1
    bottom_row = anchor_row + N_QUARTERS

    candidate_cols = [col for col in col_map.values() if isinstance(col, int)]
    if not candidate_cols:
        candidate_cols = [anchor_col]
    left_col = min(candidate_cols)
    right_col = max(candidate_cols)
    table_block = read_block(sheet, top_row, bottom_row, left_col, right_col)

    valid_rows = numeric_rows_for_xy(matrix, base_row, base_col, anchor_row, x_col, y_col)
    scratch_col = used.last_cell.column + 4
    formula_block: list[list[str | None]] = []
    for n in range(1, N_QUARTERS + 1):
        if len(valid_rows) < n or n < 2:
            formula_block.append([None, None, None])
            continue
        start_row = valid_rows[-n]
        end_row = valid_rows[-1]
        intercept_formula = (
            f"=INTERCEPT(R{start_row}C{y_col}:R{end_row}C{y_col},"
            f"R{start_row}C{x_col}:R{end_row}C{x_col})"
        )
        slope_formula = (
            f"=SLOPE(R{start_row}C{y_col}:R{end_row}C{y_col},"
            f"R{start_row}C{x_col}:R{end_row}C{x_col})"
        )
        forecast_formula = f"=RC[-2]+RC[-1]*(R{end_row}C{x_col}+1)"
        formula_block.append([intercept_formula, slope_formula, forecast_formula])

    if any(any(cell for cell in row) for row in formula_block):
        write_formula_block(sheet, top_row, scratch_col, formula_block)
        wb.app.calculate()
        calc_block = normalize_2d(
            sheet.range(
                (top_row, scratch_col),
                (bottom_row, scratch_col + 2),
            ).value
        )
    else:
        calc_block = [[None, None, None] for _ in range(N_QUARTERS)]

    rows: list[dict[str, Any]] = []
    for i in range(N_QUARTERS):
        row_num = top_row + i
        n_default = i + 1
        num_quarters_used = block_value(
            table_block,
            row_num,
            col_map["num_quarters_used"] or -1,
            top_row,
            left_col,
        )
        if as_float(num_quarters_used) is None:
            num_quarters_used = n_default

        forecast_value_sheet = block_value(
            table_block,
            row_num,
            col_map["forecast_value"] or -1,
            top_row,
            left_col,
        )
        forecast_max = block_value(table_block, row_num, col_map["forecast_max"] or -1, top_row, left_col)
        forecast_min = block_value(table_block, row_num, col_map["forecast_min"] or -1, top_row, left_col)
        actual_value = block_value(table_block, row_num, col_map["actual_value"] or -1, top_row, left_col)

        intercept = calc_block[i][0] if i < len(calc_block) and len(calc_block[i]) > 0 else None
        slope = calc_block[i][1] if i < len(calc_block) and len(calc_block[i]) > 1 else None
        forecast_value_calc = calc_block[i][2] if i < len(calc_block) and len(calc_block[i]) > 2 else None
        forecast_value = (
            forecast_value_sheet if as_float(forecast_value_sheet) is not None else forecast_value_calc
        )

        key_presence = [
            as_float(forecast_value),
            as_float(forecast_max),
            as_float(forecast_min),
            as_float(intercept),
            as_float(slope),
        ]
        if all(val is None for val in key_presence):
            continue

        row = {
            "model": meta.model,
            "ticker": meta.ticker,
            "model_period": meta.model_period,
            "model_date": meta.model_date,
            "method": "regression",
            "parameter_name": "num_quarters_used",
            "parameter_value": num_quarters_used,
            "num_quarters_used": num_quarters_used,
            "forecast_value": forecast_value,
            "actual_value": actual_value if actual_value not in ("", None) else None,
            "forecast_max": forecast_max,
            "forecast_min": forecast_min,
            "range_width": safe_subtract(forecast_max, forecast_min),
            "intercept": intercept,
            "slope": slope,
            "source_file": source_file,
        }

        if rows and dedupe_key(row) == dedupe_key(rows[-1]):
            continue
        rows.append(row)

    return rows


def autofit_columns(sheet) -> None:
    for col_idx in range(1, sheet.max_column + 1):
        max_len = 0
        for row_idx in range(1, sheet.max_row + 1):
            value = sheet.cell(row=row_idx, column=col_idx).value
            if value is None:
                continue
            value_len = len(str(value))
            if value_len > max_len:
                max_len = value_len
        width = min(max(max_len + 2, 12), 50)
        sheet.column_dimensions[get_column_letter(col_idx)].width = width


def write_sheet(
    wb: Workbook,
    name: str,
    columns: list[str],
    rows: list[dict[str, Any]],
) -> None:
    ws = wb.active if wb.active.title == "Sheet" and len(wb.sheetnames) == 1 else wb.create_sheet()
    ws.title = name

    ws.append(columns)
    for header_cell in ws[1]:
        header_cell.font = Font(bold=True)

    for row in rows:
        ws.append([row.get(col) for col in columns])

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    autofit_columns(ws)


def main() -> None:
    if not input_dir.exists() or not input_dir.is_dir():
        raise FileNotFoundError(f"input_dir does not exist or is not a directory: {input_dir}")

    files_processed = 0
    empirical_rows: list[dict[str, Any]] = []
    regression_rows: list[dict[str, Any]] = []

    app = xw.App(visible=False, add_book=False)
    app.display_alerts = False
    app.screen_updating = False
    app.calculation = "manual"

    try:
        for file_path in sorted(input_dir.iterdir()):
            if not file_path.is_file():
                print(f"Skipped {file_path.name}: not a file")
                continue
            if file_path.name.startswith("~"):
                print(f"Skipped {file_path.name}: temp file")
                continue
            if file_path.suffix.lower() != ".xlsx":
                print(f"Skipped {file_path.name}: not .xlsx")
                continue

            print(f"Processing {file_path.name}")
            wb = None
            try:
                wb = app.books.open(str(file_path), update_links=False)
                meta = parse_filename_metadata(file_path)
                empirical_rows.extend(build_empirical_rows(wb, meta, file_path.name))
                regression_rows.extend(build_regression_rows(wb, meta, file_path.name))
                files_processed += 1
            except Exception as exc:
                print(f"Skipped {file_path.name}: processing error ({exc})")
            finally:
                if wb is not None:
                    close_workbook_without_saving(wb)

    finally:
        app.quit()

    output_path = first_available_output_path(input_dir, output_dir)
    result_wb = Workbook()
    write_sheet(result_wb, "empirical_candidates", EMPIRICAL_COLUMNS, empirical_rows)
    write_sheet(result_wb, "regression_candidates", REGRESSION_COLUMNS, regression_rows)
    result_wb.save(output_path)

    print(f"Output path: {output_path}")
    print(f"Files processed: {files_processed}")
    print(f"Empirical rows: {len(empirical_rows)}")
    print(f"Regression rows: {len(regression_rows)}")


if __name__ == "__main__":
    main()
