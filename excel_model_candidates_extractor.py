from __future__ import annotations

import calendar
import re
from datetime import date
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Sequence, Tuple

import xlwings as xw
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.worksheet.worksheet import Worksheet


# Configure these folders before running.
input_dir = Path("input")
output_dir = Path("output")


EMPIRICAL_HEADERS = [
    "model",
    "ticker",
    "model_period",
    "model_date",
    "method",
    "parameter_name",
    "parameter_value",
    "num_quarters_used",
    "last_quarter_used",
    "forecast_value",
    "actual_value",
    "forecast_max",
    "forecast_min",
    "range_width",
    "avg_penetration_pct",
    "quarterly_sales",
    "reported_sales",
    "growth_rate_pct",
    "sales_captured_in_db_pct",
    "source_file",
]


REGRESSION_HEADERS = [
    "model",
    "ticker",
    "model_period",
    "model_date",
    "method",
    "parameter_name",
    "parameter_value",
    "num_quarters_used",
    "forecast_value",
    "actual_value",
    "forecast_max",
    "forecast_min",
    "range_width",
    "intercept",
    "slope",
    "source_file",
]


N_QUARTERS = 10
DAY_BY_PERIOD = {"early": 5, "mid": 15, "late": 25}
MONTH_BY_ABBR = {m.lower(): i for i, m in enumerate(calendar.month_abbr) if m}


def to_float(value: Any) -> Optional[float]:
    if value is None or value == "":
        return None
    if isinstance(value, bool):
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def to_scalar_table(value: Any) -> List[List[Any]]:
    if value is None:
        return []
    if isinstance(value, list):
        if value and isinstance(value[0], list):
            return value
        return [value]
    return [[value]]


def parse_file_labels(file_path: Path) -> Dict[str, str]:
    stem = file_path.stem
    parts = [p.strip() for p in stem.split(" - ")]
    ticker = parts[1] if len(parts) >= 2 and parts[1] else "UNKNOWN"
    period_token = ""
    if len(parts) >= 3:
        period_token = parts[2].split("_", 1)[0].strip()

    model_period = "UNKNOWN"
    model_date = ""

    match = re.match(
        r"^(Early|Mid|Late)(Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)(\d{4})$",
        period_token,
        re.IGNORECASE,
    )
    if match:
        period_word = match.group(1)
        month_word = match.group(2)
        year_text = match.group(3)
        month_num = MONTH_BY_ABBR[month_word.lower()]
        day_num = DAY_BY_PERIOD[period_word.lower()]
        parsed_date = date(int(year_text), month_num, day_num)
        model_period = f"{period_word}{month_word}_{year_text}"
        model_date = parsed_date.isoformat()

    model = f"{ticker}_{model_period}" if model_period != "UNKNOWN" else ticker
    return {
        "model": model,
        "ticker": ticker,
        "model_period": model_period,
        "model_date": model_date,
    }


def build_output_path(in_dir: Path, out_dir: Path) -> Path:
    out_dir.mkdir(parents=True, exist_ok=True)
    base_stem = f"{in_dir.name}_PARAM"
    candidate = out_dir / f"{base_stem}.xlsx"
    index = 1
    while candidate.exists():
        candidate = out_dir / f"{base_stem}.{index}.xlsx"
        index += 1
    return candidate


def close_workbook_without_save(wb: xw.Book) -> None:
    try:
        wb.close(save=False)
        return
    except TypeError:
        pass
    except Exception:
        pass

    # Fallback for environments where close(save=False) is unsupported.
    try:
        wb.api.Close(SaveChanges=False)
        return
    except Exception:
        pass

    try:
        wb.api.Close(False)
    except Exception as exc:
        print(f"Warning: unable to close workbook cleanly: {exc}")


def get_sheet(wb: xw.Book, sheet_name: str) -> Optional[xw.Sheet]:
    try:
        return wb.sheets[sheet_name]
    except Exception:
        return None


def find_anchor_cell(sheet: xw.Sheet, anchor_text: str = "max") -> Optional[Tuple[int, int, int]]:
    used = sheet.used_range
    values = to_scalar_table(used.value)
    if not values:
        return None

    start_row = used.row
    start_col = used.column
    for row_idx, row_values in enumerate(values):
        for col_idx, raw in enumerate(row_values):
            if isinstance(raw, str) and raw.strip().lower() == anchor_text:
                return (start_row + row_idx, start_col + col_idx, used.last_cell.column)
    return None


def find_min_value_near_anchor(sheet: xw.Sheet, anchor_row: int, anchor_col: int) -> Optional[float]:
    for probe in range(anchor_row, anchor_row + 6):
        label = sheet.cells(probe, anchor_col).value
        if isinstance(label, str) and label.strip().lower() == "min":
            return to_float(sheet.cells(probe, anchor_col + 1).value)
    return None


def float_equal(a: Optional[float], b: Optional[float], tolerance: float = 1e-9) -> bool:
    if a is None and b is None:
        return True
    if a is None or b is None:
        return False
    return abs(a - b) <= tolerance


def process_empirical_sheet(
    wb: xw.Book,
    file_meta: Dict[str, str],
    source_file: str,
) -> List[Dict[str, Any]]:
    sheet = get_sheet(wb, "Empirical Model")
    if sheet is None:
        print(f"  skipped empirical: sheet 'Empirical Model' not found ({source_file})")
        return []

    anchor = find_anchor_cell(sheet, "max")
    if anchor is None:
        print(f"  skipped empirical: anchor 'max' not found ({source_file})")
        return []

    anchor_row, anchor_col, used_last_col = anchor
    y_col = anchor_col - 7
    x_col = anchor_col - 11
    helper_col = max(used_last_col + 1, anchor_col + 3)
    helper_row = max(anchor_row, 1)
    helper_cell = sheet.cells(helper_row, helper_col)

    forecast_max = to_float(sheet.cells(anchor_row, anchor_col + 1).value)
    forecast_min = find_min_value_near_anchor(sheet, anchor_row, anchor_col)
    range_width = (
        (forecast_max - forecast_min)
        if forecast_max is not None and forecast_min is not None
        else None
    )

    latest_row = max(anchor_row - 1, 1)
    quarterly_sales = to_float(sheet.cells(latest_row, x_col).value)
    reported_sales = to_float(sheet.cells(latest_row, y_col).value)
    growth_rate_pct = to_float(sheet.cells(latest_row, anchor_col - 4).value)
    sales_captured_pct = to_float(sheet.cells(latest_row, anchor_col - 3).value)
    last_quarter_used = sheet.cells(latest_row, x_col - 1).value

    rows: List[Dict[str, Any]] = []
    for n_quarters in range(1, N_QUARTERS + 1):
        start_offset = -n_quarters
        y_rel = y_col - helper_col
        x_rel = x_col - helper_col
        avg_pen_formula = (
            f'=AVERAGE(IFERROR(R[{start_offset}]C[{y_rel}]:R[-1]C[{y_rel}]'
            f'/R[{start_offset}]C[{x_rel}]:R[-1]C[{x_rel}], ""))'
        )
        helper_cell.formula2 = avg_pen_formula
        wb.app.calculate()

        avg_penetration = to_float(helper_cell.value)
        if avg_penetration is None:
            continue

        forecast_value = (
            avg_penetration * quarterly_sales if quarterly_sales is not None else None
        )
        row = {
            "model": file_meta["model"],
            "ticker": file_meta["ticker"],
            "model_period": file_meta["model_period"],
            "model_date": file_meta["model_date"],
            "method": "empirical",
            "parameter_name": "avg_penetration_pct",
            "parameter_value": avg_penetration,
            "num_quarters_used": n_quarters,
            "last_quarter_used": last_quarter_used,
            "forecast_value": forecast_value,
            "actual_value": reported_sales,
            "forecast_max": forecast_max,
            "forecast_min": forecast_min,
            "range_width": range_width,
            "avg_penetration_pct": avg_penetration,
            "quarterly_sales": quarterly_sales,
            "reported_sales": reported_sales,
            "growth_rate_pct": growth_rate_pct,
            "sales_captured_in_db_pct": sales_captured_pct,
            "source_file": source_file,
        }
        rows.append(row)

    helper_cell.value = None
    return rows


def process_regression_sheet(
    wb: xw.Book,
    file_meta: Dict[str, str],
    source_file: str,
) -> List[Dict[str, Any]]:
    sheet = get_sheet(wb, "Regression Model")
    if sheet is None:
        print(f"  skipped regression: sheet 'Regression Model' not found ({source_file})")
        return []

    anchor = find_anchor_cell(sheet, "max")
    if anchor is None:
        print(f"  skipped regression: anchor 'max' not found ({source_file})")
        return []

    anchor_row, anchor_col, used_last_col = anchor
    y_col = anchor_col - 7
    x_col = anchor_col - 11
    helper_col = max(used_last_col + 1, anchor_col + 3)
    helper_row = max(anchor_row, 1)
    intercept_cell = sheet.cells(helper_row, helper_col)
    slope_cell = sheet.cells(helper_row, helper_col + 1)

    forecast_max = to_float(sheet.cells(anchor_row, anchor_col + 1).value)
    forecast_min = find_min_value_near_anchor(sheet, anchor_row, anchor_col)
    range_width = (
        (forecast_max - forecast_min)
        if forecast_max is not None and forecast_min is not None
        else None
    )
    baseline_x = to_float(sheet.cells(anchor_row, x_col).value)
    if baseline_x is None:
        baseline_x = to_float(sheet.cells(max(anchor_row - 1, 1), x_col).value)

    rows: List[Dict[str, Any]] = []
    previous_row: Optional[Dict[str, Any]] = None

    for n_quarters in range(1, N_QUARTERS + 1):
        start_offset = -n_quarters
        y_rel = y_col - helper_col
        x_rel = x_col - helper_col

        intercept_cell.formula2 = (
            f"=INTERCEPT(R[{start_offset}]C[{y_rel}]:R[-1]C[{y_rel}],"
            f"R[{start_offset}]C[{x_rel}]:R[-1]C[{x_rel}])"
        )
        slope_cell.formula2 = (
            f"=SLOPE(R[{start_offset}]C[{y_rel}]:R[-1]C[{y_rel}],"
            f"R[{start_offset}]C[{x_rel}]:R[-1]C[{x_rel}])"
        )
        wb.app.calculate()

        intercept = to_float(intercept_cell.value)
        slope = to_float(slope_cell.value)
        if intercept is None or slope is None:
            continue

        forecast_total_without_sa = None
        if baseline_x is not None:
            forecast_total_without_sa = intercept + slope * baseline_x

        current_row = {
            "model": file_meta["model"],
            "ticker": file_meta["ticker"],
            "model_period": file_meta["model_period"],
            "model_date": file_meta["model_date"],
            "method": "regression",
            "parameter_name": "num_quarters_used",
            "parameter_value": n_quarters,
            "num_quarters_used": n_quarters,
            "forecast_value": forecast_total_without_sa,
            "actual_value": None,
            "forecast_max": forecast_max,
            "forecast_min": forecast_min,
            "range_width": range_width,
            "intercept": intercept,
            "slope": slope,
            "source_file": source_file,
        }

        if previous_row is not None:
            duplicate = (
                float_equal(current_row["forecast_value"], previous_row["forecast_value"])
                and float_equal(current_row["forecast_max"], previous_row["forecast_max"])
                and float_equal(current_row["forecast_min"], previous_row["forecast_min"])
                and float_equal(current_row["intercept"], previous_row["intercept"])
                and float_equal(current_row["slope"], previous_row["slope"])
            )
            if duplicate:
                continue

        rows.append(current_row)
        previous_row = current_row

    intercept_cell.value = None
    slope_cell.value = None
    return rows


def write_sheet(ws: Worksheet, headers: Sequence[str], rows: Iterable[Dict[str, Any]]) -> None:
    ws.append(list(headers))
    for item in rows:
        ws.append([item.get(col) for col in headers])

    header_font = Font(bold=True)
    for cell in ws[1]:
        cell.font = header_font

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    for idx, header in enumerate(headers, start=1):
        max_len = len(header)
        for value in ws.iter_cols(min_col=idx, max_col=idx, min_row=2, values_only=True):
            for cell_value in value:
                if cell_value is None:
                    continue
                max_len = max(max_len, len(str(cell_value)))
        ws.column_dimensions[chr(64 + idx) if idx <= 26 else ws.cell(row=1, column=idx).column_letter].width = min(
            50, max(12, max_len + 2)
        )


def iter_candidate_files(folder: Path) -> Iterable[Path]:
    if not folder.exists():
        print(f"Input folder does not exist: {folder.resolve()}")
        return []

    for file_path in sorted(folder.iterdir()):
        if not file_path.is_file():
            continue
        if file_path.name.startswith("~"):
            print(f"skipped file: {file_path.name} (temporary file)")
            continue
        if file_path.suffix.lower() != ".xlsx":
            print(f"skipped file: {file_path.name} (not .xlsx)")
            continue
        yield file_path


def main() -> None:
    output_path = build_output_path(input_dir, output_dir)
    empirical_rows: List[Dict[str, Any]] = []
    regression_rows: List[Dict[str, Any]] = []
    processed_count = 0

    app = xw.App(visible=False, add_book=False)
    app.display_alerts = False
    app.screen_updating = False
    app.enable_events = False
    app.calculation = "manual"

    try:
        for file_path in iter_candidate_files(input_dir):
            print(f"processed file: {file_path.name}")
            file_meta = parse_file_labels(file_path)
            wb = None
            try:
                wb = app.books.open(str(file_path), update_links=False)
                empirical_rows.extend(
                    process_empirical_sheet(wb, file_meta, file_path.name)
                )
                regression_rows.extend(
                    process_regression_sheet(wb, file_meta, file_path.name)
                )
                processed_count += 1
            except Exception as exc:
                print(f"  skipped file: {file_path.name} (error: {exc})")
            finally:
                if wb is not None:
                    close_workbook_without_save(wb)
    finally:
        app.quit()

    out_wb = Workbook()
    ws_empirical = out_wb.active
    ws_empirical.title = "empirical_candidates"
    ws_regression = out_wb.create_sheet("regression_candidates")

    write_sheet(ws_empirical, EMPIRICAL_HEADERS, empirical_rows)
    write_sheet(ws_regression, REGRESSION_HEADERS, regression_rows)
    out_wb.save(output_path)

    print(f"output path: {output_path.resolve()}")
    print(f"number of files processed: {processed_count}")
    print(f"number of empirical rows: {len(empirical_rows)}")
    print(f"number of regression rows: {len(regression_rows)}")


if __name__ == "__main__":
    main()
