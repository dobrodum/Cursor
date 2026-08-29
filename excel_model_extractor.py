#!/usr/bin/env python3
"""
Extract empirical and regression model candidates from Excel model workbooks.

Runtime priorities:
- Open each source workbook once.
- Process both model sheets while it is open.
- Keep one hidden Excel app for the entire run.
- Avoid unnecessary recalculations.
"""

from __future__ import annotations

import datetime as dt
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Sequence, Tuple

from openpyxl import Workbook
from openpyxl.styles import Font

import xlwings as xw

# ---------------------------------------------------------------------------
# User-configurable paths
# ---------------------------------------------------------------------------
input_dir = Path("./input")
output_dir = Path("./output")


EMPIRICAL_COLUMNS: List[str] = [
    "model",
    "ticker",
    "model_period",
    "model_date",
    "method",
    "parameter_name",
    "parameter_value",
    "num_quarters_used",
    "last_quarter_used",
    "forecast_value",
    "actual_value",
    "forecast_max",
    "forecast_min",
    "range_width",
    "avg_penetration_pct",
    "quarterly_sales",
    "reported_sales",
    "growth_rate_pct",
    "sales_captured_in_db_pct",
    "source_file",
]

REGRESSION_COLUMNS: List[str] = [
    "model",
    "ticker",
    "model_period",
    "model_date",
    "method",
    "parameter_name",
    "parameter_value",
    "num_quarters_used",
    "forecast_value",
    "actual_value",
    "forecast_max",
    "forecast_min",
    "range_width",
    "intercept",
    "slope",
    "source_file",
]

N_QUARTERS = 10
XL_CALCULATION_MANUAL = -4135
XL_CALCULATION_AUTOMATIC = -4105
MONTH_MAP = {
    "jan": 1,
    "january": 1,
    "feb": 2,
    "february": 2,
    "mar": 3,
    "march": 3,
    "apr": 4,
    "april": 4,
    "may": 5,
    "jun": 6,
    "june": 6,
    "jul": 7,
    "july": 7,
    "aug": 8,
    "august": 8,
    "sep": 9,
    "sept": 9,
    "september": 9,
    "oct": 10,
    "october": 10,
    "nov": 11,
    "november": 11,
    "dec": 12,
    "december": 12,
}


@dataclass(frozen=True)
class FileLabel:
    model: str
    ticker: str
    model_period: str
    model_date: str


def normalize_header(text: Any) -> str:
    if text is None:
        return ""
    value = str(text).strip().lower()
    value = re.sub(r"[^a-z0-9]+", " ", value)
    return re.sub(r"\s+", " ", value).strip()


def to_float(value: Any) -> Optional[float]:
    if value is None or value == "":
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def to_int(value: Any) -> Optional[int]:
    numeric = to_float(value)
    if numeric is None:
        return None
    return int(round(numeric))


def normalize_matrix(values: Any) -> List[List[Any]]:
    if values is None:
        return []
    if not isinstance(values, list):
        return [[values]]
    if values and not isinstance(values[0], list):
        return [values]
    return values


def build_output_path(in_dir: Path, out_dir: Path) -> Path:
    out_dir.mkdir(parents=True, exist_ok=True)
    stem = f"{in_dir.name}_PARAM"
    candidate = out_dir / f"{stem}.xlsx"
    i = 1
    while candidate.exists():
        candidate = out_dir / f"{stem}.{i}.xlsx"
        i += 1
    return candidate


def parse_file_label(file_path: Path) -> Optional[FileLabel]:
    # Example: MedMiner_Model - AORT - MidJan2026_Send.xlsx
    pattern = re.compile(
        r" - (?P<ticker>[A-Za-z0-9]+) - "
        r"(?P<period>(?P<window>Early|Mid|Late)(?P<month>[A-Za-z]+)(?P<year>\d{4}))",
        re.IGNORECASE,
    )
    match = pattern.search(file_path.stem)
    if not match:
        return None

    ticker = match.group("ticker").upper()
    window_raw = match.group("window").title()
    month_raw = match.group("month").title()
    year = int(match.group("year"))

    month_key = month_raw.lower()
    month_num = MONTH_MAP.get(month_key) or MONTH_MAP.get(month_key[:3])
    if month_num is None:
        return None

    day_map = {"Early": 5, "Mid": 15, "Late": 25}
    day = day_map[window_raw]

    month_abbrev = dt.date(2000, month_num, 1).strftime("%b")
    model_period = f"{window_raw}{month_abbrev}_{year}"
    model_date = dt.date(year, month_num, day).isoformat()
    model = f"{ticker}_{model_period}"

    return FileLabel(
        model=model,
        ticker=ticker,
        model_period=model_period,
        model_date=model_date,
    )


def iter_input_workbooks(folder: Path) -> Iterable[Path]:
    if not folder.exists():
        return []
    files: List[Path] = []
    for item in sorted(folder.iterdir(), key=lambda p: p.name.lower()):
        if not item.is_file():
            continue
        if item.name.startswith("~"):
            print(f"SKIPPED {item.name}: temporary workbook")
            continue
        if item.suffix.lower() != ".xlsx":
            print(f"SKIPPED {item.name}: not an .xlsx file")
            continue
        files.append(item)
    return files


def safe_close_workbook(wb: xw.Book) -> None:
    # Never save source files.
    try:
        wb.close(save=False)
        return
    except TypeError:
        pass
    except Exception:
        pass

    try:
        wb.api.Close(False)
        return
    except Exception:
        pass

    try:
        wb.api.Close(SaveChanges=False)
    except Exception:
        pass


def get_sheet_case_insensitive(wb: xw.Book, sheet_name: str) -> Optional[xw.Sheet]:
    sheet_map = {s.name.lower(): s for s in wb.sheets}
    return sheet_map.get(sheet_name.lower())


def find_max_anchor(
    matrix: List[List[Any]], start_row: int, start_col: int
) -> Optional[Tuple[int, int]]:
    candidates: List[Tuple[int, int, int]] = []
    for r_idx, row in enumerate(matrix):
        for c_idx, value in enumerate(row):
            if normalize_header(value) != "max":
                continue
            score = 0
            right_val = row[c_idx + 1] if c_idx + 1 < len(row) else None
            if normalize_header(right_val) == "min":
                score += 10
            if r_idx <= 30:
                score += 2
            candidates.append((score, start_row + r_idx, start_col + c_idx))

    if not candidates:
        return None
    candidates.sort(key=lambda x: (-x[0], x[1], x[2]))
    _, row_num, col_num = candidates[0]
    return row_num, col_num


def headers_for_row(matrix: List[List[Any]], matrix_row_idx: int, start_col: int) -> Dict[str, int]:
    if matrix_row_idx < 0 or matrix_row_idx >= len(matrix):
        return {}
    header_map: Dict[str, int] = {}
    row = matrix[matrix_row_idx]
    for c_idx, value in enumerate(row):
        key = normalize_header(value)
        if key:
            header_map[key] = start_col + c_idx
    return header_map


def find_col_by_keywords(
    header_map: Dict[str, int], keywords: Sequence[str], default_col: int
) -> int:
    for key, col in header_map.items():
        if any(term in key for term in keywords):
            return col
    return default_col


def find_col_by_keywords_optional(
    header_map: Dict[str, int], keywords: Sequence[str]
) -> Optional[int]:
    for key, col in header_map.items():
        if any(term in key for term in keywords):
            return col
    return None


def detect_num_quarter_rows(
    matrix: List[List[Any]],
    start_row: int,
    start_col: int,
    anchor_row: int,
    num_quarters_col: Optional[int],
) -> List[Tuple[int, int]]:
    if num_quarters_col is None:
        return [(anchor_row + i, i) for i in range(1, N_QUARTERS + 1)]

    local_col = num_quarters_col - start_col
    rows: List[Tuple[int, int]] = []
    for r_idx, row_values in enumerate(matrix):
        row_num = start_row + r_idx
        if row_num <= anchor_row:
            continue
        if local_col < 0 or local_col >= len(row_values):
            continue
        value = row_values[local_col]
        numeric = to_float(value)
        if numeric is None:
            continue
        int_numeric = int(numeric)
        if abs(numeric - int_numeric) > 1e-9:
            continue
        if 1 <= int_numeric <= N_QUARTERS:
            rows.append((row_num, int_numeric))

    rows.sort(key=lambda x: x[0])
    if len(rows) >= N_QUARTERS:
        return rows[:N_QUARTERS]

    existing = {n for _, n in rows}
    fallback_rows = list(rows)
    for n in range(1, N_QUARTERS + 1):
        if n in existing:
            continue
        fallback_rows.append((anchor_row + n, n))
    fallback_rows.sort(key=lambda x: x[0])
    return fallback_rows[:N_QUARTERS]


def as_blank_if_none(value: Any) -> Any:
    return "" if value is None else value


def r1c1_ref(row_offset: int = 0, col_offset: int = 0) -> str:
    row_part = "R" if row_offset == 0 else f"R[{row_offset}]"
    col_part = "C" if col_offset == 0 else f"C[{col_offset}]"
    return row_part + col_part


def extract_empirical_rows(
    wb: xw.Book,
    sheet: xw.Sheet,
    file_label: FileLabel,
    source_file: str,
) -> List[Dict[str, Any]]:
    used = sheet.used_range
    matrix = normalize_matrix(used.value)
    if not matrix:
        return []

    start_row = used.row
    start_col = used.column

    anchor = find_max_anchor(matrix, start_row, start_col)
    if anchor is None:
        return []
    anchor_row, anchor_col = anchor

    local_anchor_idx = anchor_row - start_row
    header_map = headers_for_row(matrix, local_anchor_idx, start_col)

    num_quarters_col = find_col_by_keywords(
        header_map,
        keywords=["num quarters used", "num quarters", "quarters used", "n quarters"],
        default_col=anchor_col - 9,
    )
    quarter_rows = detect_num_quarter_rows(
        matrix=matrix,
        start_row=start_row,
        start_col=start_col,
        anchor_row=anchor_row,
        num_quarters_col=num_quarters_col,
    )

    avg_penetration_col = find_col_by_keywords(
        header_map,
        keywords=["avg penetration pct", "avg penetration", "average penetration"],
        default_col=anchor_col - 7,
    )

    forecast_value_col = find_col_by_keywords(
        header_map,
        keywords=["estimated total sold", "est total sold", "forecast", "tot fcst"],
        default_col=anchor_col - 2,
    )
    actual_value_col = find_col_by_keywords(
        header_map,
        keywords=["reported sales", "actual sales", "actual"],
        default_col=anchor_col - 1,
    )
    forecast_max_col = anchor_col
    forecast_min_col = find_col_by_keywords(
        header_map, keywords=["min"], default_col=anchor_col + 1
    )
    last_quarter_col = find_col_by_keywords(
        header_map, keywords=["last quarter used", "last quarter"], default_col=anchor_col - 8
    )
    quarterly_sales_col = find_col_by_keywords(
        header_map, keywords=["quarterly sales"], default_col=anchor_col - 6
    )
    reported_sales_col = find_col_by_keywords(
        header_map, keywords=["reported sales"], default_col=anchor_col - 5
    )
    growth_rate_col = find_col_by_keywords(
        header_map, keywords=["growth rate"], default_col=anchor_col - 4
    )
    sales_captured_col = find_col_by_keywords(
        header_map, keywords=["captured in db", "capture in db"], default_col=anchor_col - 3
    )

    avg_formula_col = avg_penetration_col
    relative_capture = sales_captured_col - avg_formula_col
    formulas_written = False
    for row_num, n_quarters in quarter_rows:
        span = max(n_quarters - 1, 0)
        start_ref = r1c1_ref(row_offset=-span, col_offset=relative_capture)
        end_ref = r1c1_ref(row_offset=0, col_offset=relative_capture)
        formula = f'=IFERROR(AVERAGE({start_ref}:{end_ref}),"")'
        sheet.cells(row_num, avg_formula_col).formula2 = formula
        formulas_written = True

    if formulas_written:
        wb.app.calculate()

    rows: List[Dict[str, Any]] = []
    for row_num, n_quarters in quarter_rows:
        avg_pen = sheet.cells(row_num, avg_formula_col).value
        forecast_value = sheet.cells(row_num, forecast_value_col).value
        actual_value = sheet.cells(row_num, actual_value_col).value
        forecast_max = sheet.cells(row_num, forecast_max_col).value
        forecast_min = sheet.cells(row_num, forecast_min_col).value
        forecast_max_num = to_float(forecast_max)
        forecast_min_num = to_float(forecast_min)
        range_width = (
            forecast_max_num - forecast_min_num
            if forecast_max_num is not None and forecast_min_num is not None
            else None
        )

        row: Dict[str, Any] = {
            "model": file_label.model,
            "ticker": file_label.ticker,
            "model_period": file_label.model_period,
            "model_date": file_label.model_date,
            "method": "empirical",
            "parameter_name": "avg_penetration_pct",
            "parameter_value": as_blank_if_none(avg_pen),
            "num_quarters_used": n_quarters,
            "last_quarter_used": as_blank_if_none(sheet.cells(row_num, last_quarter_col).value),
            "forecast_value": as_blank_if_none(forecast_value),
            "actual_value": as_blank_if_none(actual_value),
            "forecast_max": as_blank_if_none(forecast_max),
            "forecast_min": as_blank_if_none(forecast_min),
            "range_width": as_blank_if_none(range_width),
            "avg_penetration_pct": as_blank_if_none(avg_pen),
            "quarterly_sales": as_blank_if_none(sheet.cells(row_num, quarterly_sales_col).value),
            "reported_sales": as_blank_if_none(sheet.cells(row_num, reported_sales_col).value),
            "growth_rate_pct": as_blank_if_none(sheet.cells(row_num, growth_rate_col).value),
            "sales_captured_in_db_pct": as_blank_if_none(
                sheet.cells(row_num, sales_captured_col).value
            ),
            "source_file": source_file,
        }
        rows.append(row)

    return rows


def extract_regression_rows(
    wb: xw.Book,
    sheet: xw.Sheet,
    file_label: FileLabel,
    source_file: str,
) -> List[Dict[str, Any]]:
    used = sheet.used_range
    matrix = normalize_matrix(used.value)
    if not matrix:
        return []

    start_row = used.row
    start_col = used.column

    anchor = find_max_anchor(matrix, start_row, start_col)
    if anchor is None:
        return []
    anchor_row, anchor_col = anchor

    local_anchor_idx = anchor_row - start_row
    header_map = headers_for_row(matrix, local_anchor_idx, start_col)

    num_quarters_col = find_col_by_keywords(
        header_map,
        keywords=["num quarters used", "num quarters", "quarters used", "n quarters"],
        default_col=anchor_col - 4,
    )
    quarter_rows = detect_num_quarter_rows(
        matrix=matrix,
        start_row=start_row,
        start_col=start_col,
        anchor_row=anchor_row,
        num_quarters_col=num_quarters_col,
    )

    y_col = anchor_col - 7
    x_col = anchor_col - 11

    forecast_value_col = find_col_by_keywords(
        header_map,
        keywords=[
            "tot fcst w o sa",
            "tot fcst wo sa",
            "tot fcst without sa",
            "forecast total without sa",
            "forecast",
        ],
        default_col=anchor_col - 1,
    )
    forecast_max_col = anchor_col
    forecast_min_col = find_col_by_keywords(
        header_map, keywords=["min"], default_col=anchor_col + 1
    )
    actual_value_col = find_col_by_keywords_optional(header_map, keywords=["actual", "reported"])
    intercept_formula_col = find_col_by_keywords(
        header_map, keywords=["intercept"], default_col=anchor_col - 3
    )
    slope_formula_col = find_col_by_keywords(
        header_map, keywords=["slope"], default_col=anchor_col - 2
    )

    formulas_written = False
    for row_num, n_quarters in quarter_rows:
        num_quarters = to_int(sheet.cells(row_num, num_quarters_col).value) or n_quarters
        span = max(num_quarters - 1, 0)

        intercept_y_start = r1c1_ref(row_offset=-span, col_offset=(y_col - intercept_formula_col))
        intercept_y_end = r1c1_ref(row_offset=0, col_offset=(y_col - intercept_formula_col))
        intercept_x_start = r1c1_ref(row_offset=-span, col_offset=(x_col - intercept_formula_col))
        intercept_x_end = r1c1_ref(row_offset=0, col_offset=(x_col - intercept_formula_col))
        intercept_formula = (
            f'=IFERROR(INTERCEPT({intercept_y_start}:{intercept_y_end},'
            f"{intercept_x_start}:{intercept_x_end}),\"\")"
        )

        slope_y_start = r1c1_ref(row_offset=-span, col_offset=(y_col - slope_formula_col))
        slope_y_end = r1c1_ref(row_offset=0, col_offset=(y_col - slope_formula_col))
        slope_x_start = r1c1_ref(row_offset=-span, col_offset=(x_col - slope_formula_col))
        slope_x_end = r1c1_ref(row_offset=0, col_offset=(x_col - slope_formula_col))
        slope_formula = (
            f'=IFERROR(SLOPE({slope_y_start}:{slope_y_end},'
            f"{slope_x_start}:{slope_x_end}),\"\")"
        )
        sheet.cells(row_num, intercept_formula_col).formula2 = intercept_formula
        sheet.cells(row_num, slope_formula_col).formula2 = slope_formula
        formulas_written = True

    if formulas_written:
        wb.app.calculate()

    rows: List[Dict[str, Any]] = []
    prev_signature: Optional[Tuple[Any, ...]] = None
    for row_num, n_quarters in quarter_rows:
        num_quarters_used = to_int(sheet.cells(row_num, num_quarters_col).value) or n_quarters
        intercept = sheet.cells(row_num, intercept_formula_col).value
        slope = sheet.cells(row_num, slope_formula_col).value
        forecast_value = sheet.cells(row_num, forecast_value_col).value
        forecast_max = sheet.cells(row_num, forecast_max_col).value
        forecast_min = sheet.cells(row_num, forecast_min_col).value
        actual_value = (
            sheet.cells(row_num, actual_value_col).value
            if actual_value_col is not None
            else ""
        )

        max_num = to_float(forecast_max)
        min_num = to_float(forecast_min)
        range_width = max_num - min_num if max_num is not None and min_num is not None else None

        signature = (
            round(to_float(intercept) or 0.0, 10),
            round(to_float(slope) or 0.0, 10),
            round(to_float(forecast_value) or 0.0, 10),
            round(to_float(forecast_max) or 0.0, 10),
            round(to_float(forecast_min) or 0.0, 10),
        )
        if prev_signature == signature:
            continue
        prev_signature = signature

        row: Dict[str, Any] = {
            "model": file_label.model,
            "ticker": file_label.ticker,
            "model_period": file_label.model_period,
            "model_date": file_label.model_date,
            "method": "regression",
            "parameter_name": "num_quarters_used",
                "parameter_value": num_quarters_used,
                "num_quarters_used": num_quarters_used,
            "forecast_value": as_blank_if_none(forecast_value),
            "actual_value": as_blank_if_none(actual_value),
            "forecast_max": as_blank_if_none(forecast_max),
            "forecast_min": as_blank_if_none(forecast_min),
            "range_width": as_blank_if_none(range_width),
            "intercept": as_blank_if_none(intercept),
            "slope": as_blank_if_none(slope),
            "source_file": source_file,
        }
        rows.append(row)

    return rows


def write_output_workbook(
    output_file: Path, empirical_rows: List[Dict[str, Any]], regression_rows: List[Dict[str, Any]]
) -> None:
    wb = Workbook()
    wb.remove(wb.active)

    ws_emp = wb.create_sheet("empirical_candidates")
    ws_reg = wb.create_sheet("regression_candidates")

    ws_emp.append(EMPIRICAL_COLUMNS)
    for row in empirical_rows:
        ws_emp.append([row.get(col, "") for col in EMPIRICAL_COLUMNS])

    ws_reg.append(REGRESSION_COLUMNS)
    for row in regression_rows:
        ws_reg.append([row.get(col, "") for col in REGRESSION_COLUMNS])

    for ws in (ws_emp, ws_reg):
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        for cell in ws[1]:
            cell.font = Font(bold=True)
        for col_idx, col_name in enumerate(ws.iter_cols(min_row=1, max_row=ws.max_row), start=1):
            max_len = 0
            for cell in col_name:
                value = "" if cell.value is None else str(cell.value)
                if len(value) > max_len:
                    max_len = len(value)
            width = min(60, max(12, max_len + 2))
            ws.column_dimensions[ws.cell(1, col_idx).column_letter].width = width

    wb.save(output_file)


def main() -> None:
    if not input_dir.exists():
        print(f"Input directory not found: {input_dir.resolve()}")
        return

    files = list(iter_input_workbooks(input_dir))
    if not files:
        output_file = build_output_path(input_dir, output_dir)
        write_output_workbook(output_file, empirical_rows=[], regression_rows=[])
        print(f"Output written: {output_file.resolve()}")
        print("Files processed: 0")
        print("Empirical rows: 0")
        print("Regression rows: 0")
        return

    empirical_rows: List[Dict[str, Any]] = []
    regression_rows: List[Dict[str, Any]] = []
    processed_files = 0

    app = xw.App(visible=False, add_book=False)
    app.display_alerts = False
    app.screen_updating = False

    original_calc_mode = None
    try:
        try:
            original_calc_mode = app.api.Calculation
            app.api.Calculation = XL_CALCULATION_MANUAL
        except Exception:
            original_calc_mode = None

        for file_path in files:
            label = parse_file_label(file_path)
            if label is None:
                print(f"SKIPPED {file_path.name}: filename does not match expected pattern")
                continue

            wb: Optional[xw.Book] = None
            try:
                wb = app.books.open(str(file_path), update_links=False)

                empirical_sheet = get_sheet_case_insensitive(wb, "Empirical Model")
                regression_sheet = get_sheet_case_insensitive(wb, "Regression Model")

                if empirical_sheet is None and regression_sheet is None:
                    print(
                        f"SKIPPED {file_path.name}: missing both 'Empirical Model' and "
                        f"'Regression Model' sheets"
                    )
                    continue

                if empirical_sheet is not None:
                    empirical_rows.extend(
                        extract_empirical_rows(
                            wb=wb,
                            sheet=empirical_sheet,
                            file_label=label,
                            source_file=file_path.name,
                        )
                    )
                else:
                    print(f"SKIPPED empirical in {file_path.name}: sheet not found")

                if regression_sheet is not None:
                    regression_rows.extend(
                        extract_regression_rows(
                            wb=wb,
                            sheet=regression_sheet,
                            file_label=label,
                            source_file=file_path.name,
                        )
                    )
                else:
                    print(f"SKIPPED regression in {file_path.name}: sheet not found")

                processed_files += 1
                print(f"PROCESSED {file_path.name}")
            except Exception as exc:
                print(f"SKIPPED {file_path.name}: {exc}")
            finally:
                if wb is not None:
                    safe_close_workbook(wb)
    finally:
        try:
            if original_calc_mode is not None:
                app.api.Calculation = original_calc_mode
            else:
                app.api.Calculation = XL_CALCULATION_AUTOMATIC
        except Exception:
            pass
        app.quit()

    output_file = build_output_path(input_dir, output_dir)
    write_output_workbook(output_file, empirical_rows, regression_rows)

    print(f"Output written: {output_file.resolve()}")
    print(f"Files processed: {processed_files}")
    print(f"Empirical rows: {len(empirical_rows)}")
    print(f"Regression rows: {len(regression_rows)}")


if __name__ == "__main__":
    main()
