#!/usr/bin/env python3
"""Extract empirical and regression model candidates from Excel workbooks.

This script reads every non-temporary .xlsx file in input_dir, opens each source
workbook exactly once, processes both required sheets while that workbook is
open, and writes one consolidated output workbook.
"""

from __future__ import annotations

from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
import re
from typing import Any

import xlwings as xw
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter


# =========================
# User-configurable paths
# =========================
input_dir = Path("/path/to/input")
output_dir = Path("/path/to/output")


EMPIRICAL_COLUMNS = [
    "model",
    "ticker",
    "model_period",
    "model_date",
    "method",
    "parameter_name",
    "parameter_value",
    "num_quarters_used",
    "last_quarter_used",
    "forecast_value",
    "actual_value",
    "forecast_max",
    "forecast_min",
    "range_width",
    "avg_penetration_pct",
    "quarterly_sales",
    "reported_sales",
    "growth_rate_pct",
    "sales_captured_in_db_pct",
    "source_file",
]

REGRESSION_COLUMNS = [
    "model",
    "ticker",
    "model_period",
    "model_date",
    "method",
    "parameter_name",
    "parameter_value",
    "num_quarters_used",
    "forecast_value",
    "actual_value",
    "forecast_max",
    "forecast_min",
    "range_width",
    "intercept",
    "slope",
    "source_file",
]

MODEL_FILE_PATTERN = re.compile(
    r"Model\s*-\s*(?P<ticker>[A-Za-z0-9]+)\s*-\s*(?P<label>(?P<period>Early|Mid|Late)(?P<month>[A-Za-z]+)(?P<year>\d{4}))",
    flags=re.IGNORECASE,
)

PERIOD_DAY = {
    "early": 5,
    "mid": 15,
    "late": 25,
}

EMPIRICAL_N_QUARTERS = 10
REGRESSION_N_QUARTERS = 10


@dataclass(frozen=True)
class ModelLabel:
    model: str
    ticker: str
    model_period: str
    model_date: str


def log(message: str) -> None:
    print(message, flush=True)


def normalize_text(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip().lower()
    text = re.sub(r"[^a-z0-9]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def as_2d(values: Any) -> list[list[Any]]:
    if values is None:
        return []
    if not isinstance(values, list):
        if isinstance(values, tuple):
            values = list(values)
        else:
            return [[values]]
    if not values:
        return []
    first = values[0]
    if not isinstance(first, (list, tuple)):
        return [list(values)]
    rows: list[list[Any]] = []
    for row in values:
        if isinstance(row, tuple):
            rows.append(list(row))
        else:
            rows.append(row)
    return rows


def as_column(values: Any) -> list[list[Any]]:
    if values is None:
        return []
    if isinstance(values, tuple):
        values = list(values)
    if isinstance(values, list):
        if not values:
            return []
        first = values[0]
        if isinstance(first, (list, tuple)):
            rows: list[list[Any]] = []
            for row in values:
                if isinstance(row, tuple):
                    rows.append(list(row))
                else:
                    rows.append(row)
            return rows
        return [[item] for item in values]
    return [[values]]


def to_number(value: Any) -> float | None:
    if value is None:
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip()
    if not text:
        return None
    text = text.replace(",", "")
    try:
        if text.endswith("%"):
            return float(text[:-1]) / 100.0
        return float(text)
    except ValueError:
        return None


def number_or_original(value: Any) -> Any:
    parsed = to_number(value)
    if parsed is None:
        return value
    return parsed


def subtract_if_numeric(max_value: Any, min_value: Any) -> float | None:
    max_num = to_number(max_value)
    min_num = to_number(min_value)
    if max_num is None or min_num is None:
        return None
    return max_num - min_num


def parse_model_label(file_path: Path) -> ModelLabel | None:
    match = MODEL_FILE_PATTERN.search(file_path.stem)
    if not match:
        return None

    ticker = match.group("ticker").upper()
    period = match.group("period").title()
    month_text = match.group("month")
    year = int(match.group("year"))

    month_num = None
    for candidate in (month_text, month_text[:3]):
        for fmt in ("%B", "%b"):
            try:
                month_num = datetime.strptime(candidate.title(), fmt).month
                break
            except ValueError:
                continue
        if month_num is not None:
            break
    if month_num is None:
        return None

    month_abbrev = datetime(year=2000, month=month_num, day=1).strftime("%b")
    model_period = f"{period}{month_abbrev}_{year}"
    model_date = date(year, month_num, PERIOD_DAY[period.lower()]).isoformat()
    model = f"{ticker}_{model_period}"

    return ModelLabel(
        model=model,
        ticker=ticker,
        model_period=model_period,
        model_date=model_date,
    )


def output_path_for_run(input_folder: Path, out_folder: Path) -> Path:
    out_folder.mkdir(parents=True, exist_ok=True)
    base_name = f"{input_folder.name}_PARAM"
    candidate = out_folder / f"{base_name}.xlsx"
    i = 1
    while candidate.exists():
        candidate = out_folder / f"{base_name}.{i}.xlsx"
        i += 1
    return candidate


def find_anchor(sheet: xw.Sheet, anchor_text: str = "max") -> tuple[int, int, int, int, list[list[Any]]]:
    used = sheet.used_range
    start_row = used.row
    start_col = used.column
    values_2d = as_2d(used.value)

    needle = normalize_text(anchor_text)
    for r_idx, row in enumerate(values_2d):
        for c_idx, value in enumerate(row):
            if normalize_text(value) == needle:
                return (
                    start_row + r_idx,
                    start_col + c_idx,
                    start_row,
                    start_col,
                    values_2d,
                )
    raise ValueError(f"Anchor '{anchor_text}' not found in sheet '{sheet.name}'.")


def read_cell(sheet: xw.Sheet, row: int, col: int) -> Any:
    if row < 1 or col < 1:
        return None
    try:
        return sheet.cells(row, col).value
    except Exception:
        return None


def set_formula_r1c1_formula2(cell: xw.Range, formula_r1c1: str) -> None:
    try:
        cell.formula2 = formula_r1c1
    except Exception:
        cell.formula = formula_r1c1


def safe_close_workbook(wb: xw.Book) -> None:
    try:
        wb.close(save=False)
        return
    except TypeError:
        pass
    except Exception:
        pass

    try:
        wb.close(False)
        return
    except Exception:
        pass

    try:
        wb.api.Close(SaveChanges=False)
        return
    except Exception:
        pass

    try:
        wb.api.Close(False)
    except Exception:
        pass


def header_offsets_from_rows(
    values_2d: list[list[Any]],
    used_start_row: int,
    used_start_col: int,
    anchor_col: int,
    candidate_header_rows: list[int],
    aliases: dict[str, list[str]],
) -> dict[str, int]:
    offsets: dict[str, int] = {}
    for header_row in candidate_header_rows:
        local_row = header_row - used_start_row
        if local_row < 0 or local_row >= len(values_2d):
            continue
        headers = values_2d[local_row]
        for key, synonyms in aliases.items():
            if key in offsets:
                continue
            best_col = None
            best_distance = None
            for idx, cell_value in enumerate(headers):
                normalized = normalize_text(cell_value)
                if not normalized:
                    continue
                if any(s in normalized for s in synonyms):
                    col = used_start_col + idx
                    distance = abs(col - anchor_col)
                    if best_distance is None or distance < best_distance:
                        best_col = col
                        best_distance = distance
            if best_col is not None:
                offsets[key] = best_col - anchor_col
    return offsets


def first_matching_sheet(wb: xw.Book, sheet_name: str) -> xw.Sheet | None:
    wanted = normalize_text(sheet_name)
    for sheet in wb.sheets:
        if normalize_text(sheet.name) == wanted:
            return sheet
    return None


def get_candidate_rows(
    sheet: xw.Sheet,
    anchor_row: int,
    num_quarters_col: int | None,
    max_rows: int,
) -> list[tuple[int, int]]:
    if num_quarters_col is not None:
        range_end = anchor_row + 80
        values = as_column(sheet.range((anchor_row + 1, num_quarters_col), (range_end, num_quarters_col)).value)
        rows: list[tuple[int, int]] = []
        seen_n: set[int] = set()
        for idx, row_values in enumerate(values):
            row_number = anchor_row + 1 + idx
            raw = row_values[0] if row_values else None
            maybe_n = to_number(raw)
            if maybe_n is None:
                continue
            n = int(round(maybe_n))
            if n < 1 or n > max_rows or n in seen_n:
                continue
            rows.append((row_number, n))
            seen_n.add(n)
            if len(rows) >= max_rows:
                break
        if rows:
            return rows

    return [(anchor_row + i, i) for i in range(1, max_rows + 1)]


def find_penetration_series_rows(
    sheet: xw.Sheet,
    values_2d: list[list[Any]],
    used_start_row: int,
    used_start_col: int,
    anchor_row: int,
    anchor_col: int,
) -> tuple[int | None, list[int]]:
    best_col = None
    best_score = None

    for r_idx, row in enumerate(values_2d):
        abs_row = used_start_row + r_idx
        if abs_row >= anchor_row:
            break
        for c_idx, value in enumerate(row):
            normalized = normalize_text(value)
            if not normalized:
                continue
            if "penetration" in normalized and "avg" not in normalized and "average" not in normalized:
                abs_col = used_start_col + c_idx
                score = abs(abs_col - anchor_col)
                if best_score is None or score < best_score:
                    best_col = abs_col
                    best_score = score

    if best_col is None:
        return None, []

    data_start = used_start_row
    data_end = anchor_row - 1
    if data_end < data_start:
        return best_col, []

    col_values = as_column(sheet.range((data_start, best_col), (data_end, best_col)).value)
    numeric_rows: list[int] = []
    for i, row_value in enumerate(col_values):
        value = row_value[0] if row_value else None
        if to_number(value) is not None:
            numeric_rows.append(data_start + i)

    return best_col, numeric_rows


def extract_empirical_rows(
    wb: xw.Book,
    label: ModelLabel,
    source_file: str,
) -> list[dict[str, Any]]:
    sheet = first_matching_sheet(wb, "Empirical Model")
    if sheet is None:
        log(f"Skipped empirical extraction in {source_file}: sheet 'Empirical Model' missing")
        return []

    try:
        anchor_row, anchor_col, used_start_row, used_start_col, values_2d = find_anchor(sheet, "max")
    except Exception as exc:
        log(f"Skipped empirical extraction in {source_file}: {exc}")
        return []

    aliases = {
        "num_quarters_used": ["num quarters", "quarters used", "n quarters", "num qtrs", "qtrs used"],
        "last_quarter_used": ["last quarter", "last qtr", "quarter used"],
        "forecast_value": ["estimated total sold", "tot fcst", "forecast", "total sold"],
        "actual_value": ["actual sales", "actual value", "actual", "reported sales"],
        "forecast_max": ["max"],
        "forecast_min": ["min"],
        "avg_penetration_pct": ["avg penetration", "average penetration"],
        "quarterly_sales": ["quarterly sales", "qtr sales"],
        "reported_sales": ["reported sales"],
        "growth_rate_pct": ["growth rate", "growth pct", "growth %"],
        "sales_captured_in_db_pct": ["sales captured in db", "captured in db", "db pct", "db %"],
    }
    offsets = header_offsets_from_rows(
        values_2d=values_2d,
        used_start_row=used_start_row,
        used_start_col=used_start_col,
        anchor_col=anchor_col,
        candidate_header_rows=[anchor_row, anchor_row - 1, anchor_row + 1],
        aliases=aliases,
    )

    num_quarters_col = None
    if "num_quarters_used" in offsets:
        num_quarters_col = anchor_col + offsets["num_quarters_used"]
    candidate_rows = get_candidate_rows(sheet, anchor_row, num_quarters_col, EMPIRICAL_N_QUARTERS)

    penetration_col, penetration_rows = find_penetration_series_rows(
        sheet=sheet,
        values_2d=values_2d,
        used_start_row=used_start_row,
        used_start_col=used_start_col,
        anchor_row=anchor_row,
        anchor_col=anchor_col,
    )

    temp_formula_row = max(1, anchor_row - 1)
    temp_formula_col = anchor_col + 6
    temp_avg_cell = sheet.cells(temp_formula_row, temp_formula_col)

    rows: list[dict[str, Any]] = []
    for row_idx, default_n in candidate_rows:
        n_quarters_raw = read_cell(sheet, row_idx, num_quarters_col) if num_quarters_col else default_n
        n_quarters_num = to_number(n_quarters_raw)
        n_quarters_used = int(round(n_quarters_num)) if n_quarters_num is not None else default_n
        if n_quarters_used < 1:
            continue

        avg_penetration_formula_value = None
        if penetration_col is not None and penetration_rows:
            n_for_avg = min(n_quarters_used, len(penetration_rows))
            if n_for_avg > 0:
                selected_rows = penetration_rows[-n_for_avg:]
                start_row = selected_rows[0]
                end_row = selected_rows[-1]
                set_formula_r1c1_formula2(
                    temp_avg_cell,
                    f"=AVERAGE(R{start_row}C{penetration_col}:R{end_row}C{penetration_col})",
                )
                wb.app.calculate()
                avg_penetration_formula_value = temp_avg_cell.value

        def by_offset(key: str) -> Any:
            if key not in offsets:
                return None
            return read_cell(sheet, row_idx, anchor_col + offsets[key])

        forecast_max = by_offset("forecast_max")
        forecast_min = by_offset("forecast_min")
        forecast_value = by_offset("forecast_value")
        reported_sales = by_offset("reported_sales")
        actual_value = by_offset("actual_value") if by_offset("actual_value") is not None else reported_sales
        avg_penetration_raw = by_offset("avg_penetration_pct")
        avg_penetration_pct = (
            avg_penetration_formula_value if avg_penetration_formula_value is not None else avg_penetration_raw
        )

        row = {
            "model": label.model,
            "ticker": label.ticker,
            "model_period": label.model_period,
            "model_date": label.model_date,
            "method": "empirical",
            "parameter_name": "avg_penetration_pct",
            "parameter_value": number_or_original(avg_penetration_pct),
            "num_quarters_used": n_quarters_used,
            "last_quarter_used": by_offset("last_quarter_used"),
            "forecast_value": number_or_original(forecast_value),
            "actual_value": number_or_original(actual_value),
            "forecast_max": number_or_original(forecast_max),
            "forecast_min": number_or_original(forecast_min),
            "range_width": subtract_if_numeric(forecast_max, forecast_min),
            "avg_penetration_pct": number_or_original(avg_penetration_pct),
            "quarterly_sales": number_or_original(by_offset("quarterly_sales")),
            "reported_sales": number_or_original(reported_sales),
            "growth_rate_pct": number_or_original(by_offset("growth_rate_pct")),
            "sales_captured_in_db_pct": number_or_original(by_offset("sales_captured_in_db_pct")),
            "source_file": source_file,
        }
        rows.append(row)

    return rows


def extract_regression_rows(
    wb: xw.Book,
    label: ModelLabel,
    source_file: str,
) -> list[dict[str, Any]]:
    sheet = first_matching_sheet(wb, "Regression Model")
    if sheet is None:
        log(f"Skipped regression extraction in {source_file}: sheet 'Regression Model' missing")
        return []

    try:
        anchor_row, anchor_col, used_start_row, used_start_col, values_2d = find_anchor(sheet, "max")
    except Exception as exc:
        log(f"Skipped regression extraction in {source_file}: {exc}")
        return []

    y_col = anchor_col - 7
    x_col = anchor_col - 11

    aliases = {
        "num_quarters_used": ["num quarters", "quarters used", "n quarters", "num qtrs", "qtrs used"],
        "forecast_value": ["tot fcst w/o sa", "tot fcst wo sa", "forecast w/o sa", "without sa"],
        "actual_value": ["actual sales", "actual value", "actual", "reported sales"],
        "forecast_max": ["max"],
        "forecast_min": ["min"],
    }
    offsets = header_offsets_from_rows(
        values_2d=values_2d,
        used_start_row=used_start_row,
        used_start_col=used_start_col,
        anchor_col=anchor_col,
        candidate_header_rows=[anchor_row, anchor_row - 1, anchor_row + 1],
        aliases=aliases,
    )
    num_quarters_col = None
    if "num_quarters_used" in offsets:
        num_quarters_col = anchor_col + offsets["num_quarters_used"]
    candidate_rows = get_candidate_rows(sheet, anchor_row, num_quarters_col, REGRESSION_N_QUARTERS)

    history_start = used_start_row
    history_end = anchor_row - 1
    history_rows: list[int] = []
    if history_end >= history_start:
        x_vals = as_column(sheet.range((history_start, x_col), (history_end, x_col)).value)
        y_vals = as_column(sheet.range((history_start, y_col), (history_end, y_col)).value)
        row_count = min(len(x_vals), len(y_vals))
        for i in range(row_count):
            x_raw = x_vals[i][0] if x_vals[i] else None
            y_raw = y_vals[i][0] if y_vals[i] else None
            if to_number(x_raw) is not None and to_number(y_raw) is not None:
                history_rows.append(history_start + i)

    if len(history_rows) < 2:
        log(f"Skipped regression extraction in {source_file}: insufficient numeric history for INTERCEPT/SLOPE")
        return []

    temp_formula_row = max(1, anchor_row - 1)
    intercept_cell = sheet.cells(temp_formula_row, anchor_col + 6)
    slope_cell = sheet.cells(temp_formula_row, anchor_col + 7)

    def by_offset(row_idx: int, key: str) -> Any:
        if key not in offsets:
            return None
        return read_cell(sheet, row_idx, anchor_col + offsets[key])

    rows: list[dict[str, Any]] = []
    previous_signature: tuple[Any, ...] | None = None

    for row_idx, default_n in candidate_rows:
        n_quarters_raw = read_cell(sheet, row_idx, num_quarters_col) if num_quarters_col else default_n
        n_quarters_num = to_number(n_quarters_raw)
        n_quarters_used = int(round(n_quarters_num)) if n_quarters_num is not None else default_n
        if n_quarters_used < 2:
            continue

        n_for_regression = min(n_quarters_used, len(history_rows))
        if n_for_regression < 2:
            continue

        selected_rows = history_rows[-n_for_regression:]
        reg_start = selected_rows[0]
        reg_end = selected_rows[-1]

        set_formula_r1c1_formula2(
            intercept_cell,
            f"=INTERCEPT(R{reg_start}C{y_col}:R{reg_end}C{y_col},R{reg_start}C{x_col}:R{reg_end}C{x_col})",
        )
        set_formula_r1c1_formula2(
            slope_cell,
            f"=SLOPE(R{reg_start}C{y_col}:R{reg_end}C{y_col},R{reg_start}C{x_col}:R{reg_end}C{x_col})",
        )
        wb.app.calculate()

        intercept = intercept_cell.value
        slope = slope_cell.value

        forecast_total_without_sa = by_offset(row_idx, "forecast_value")
        if to_number(forecast_total_without_sa) is None:
            x_for_forecast = read_cell(sheet, reg_end + 1, x_col)
            x_for_forecast_num = to_number(x_for_forecast)
            intercept_num = to_number(intercept)
            slope_num = to_number(slope)
            if x_for_forecast_num is None:
                x_for_forecast_num = to_number(read_cell(sheet, reg_end, x_col))
            if intercept_num is not None and slope_num is not None and x_for_forecast_num is not None:
                forecast_total_without_sa = intercept_num + (slope_num * x_for_forecast_num)

        forecast_max = by_offset(row_idx, "forecast_max")
        forecast_min = by_offset(row_idx, "forecast_min")

        signature = (
            round(to_number(forecast_total_without_sa) or 0.0, 10),
            round(to_number(forecast_max) or 0.0, 10),
            round(to_number(forecast_min) or 0.0, 10),
            round(to_number(intercept) or 0.0, 10),
            round(to_number(slope) or 0.0, 10),
        )
        if previous_signature is not None and signature == previous_signature:
            continue
        previous_signature = signature

        row = {
            "model": label.model,
            "ticker": label.ticker,
            "model_period": label.model_period,
            "model_date": label.model_date,
            "method": "regression",
            "parameter_name": "num_quarters_used",
            "parameter_value": n_quarters_used,
            "num_quarters_used": n_quarters_used,
            "forecast_value": number_or_original(forecast_total_without_sa),
            "actual_value": number_or_original(by_offset(row_idx, "actual_value")),
            "forecast_max": number_or_original(forecast_max),
            "forecast_min": number_or_original(forecast_min),
            "range_width": subtract_if_numeric(forecast_max, forecast_min),
            "intercept": number_or_original(intercept),
            "slope": number_or_original(slope),
            "source_file": source_file,
        }
        rows.append(row)

    return rows


def apply_basic_sheet_formatting(ws) -> None:
    bold = Font(bold=True)
    for col in range(1, ws.max_column + 1):
        ws.cell(row=1, column=col).font = bold

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    for col in range(1, ws.max_column + 1):
        max_len = 0
        for row in range(1, ws.max_row + 1):
            value = ws.cell(row=row, column=col).value
            if value is None:
                continue
            max_len = max(max_len, len(str(value)))
        ws.column_dimensions[get_column_letter(col)].width = min(max(max_len + 2, 12), 42)


def write_output_workbook(
    output_path: Path,
    empirical_rows: list[dict[str, Any]],
    regression_rows: list[dict[str, Any]],
) -> None:
    wb = Workbook()
    default_sheet = wb.active
    wb.remove(default_sheet)

    empirical_ws = wb.create_sheet("empirical_candidates")
    empirical_ws.append(EMPIRICAL_COLUMNS)
    for row in empirical_rows:
        empirical_ws.append([row.get(col) for col in EMPIRICAL_COLUMNS])
    apply_basic_sheet_formatting(empirical_ws)

    regression_ws = wb.create_sheet("regression_candidates")
    regression_ws.append(REGRESSION_COLUMNS)
    for row in regression_rows:
        regression_ws.append([row.get(col) for col in REGRESSION_COLUMNS])
    apply_basic_sheet_formatting(regression_ws)

    wb.save(output_path)


def source_files(folder: Path) -> list[Path]:
    if not folder.exists():
        raise FileNotFoundError(f"Input directory not found: {folder}")
    if not folder.is_dir():
        raise NotADirectoryError(f"Input path is not a directory: {folder}")

    candidates = sorted(folder.iterdir(), key=lambda p: p.name.lower())
    selected: list[Path] = []
    for file_path in candidates:
        if not file_path.is_file():
            continue
        if file_path.name.startswith("~"):
            log(f"Skipped {file_path.name}: temporary file")
            continue
        if file_path.suffix.lower() != ".xlsx":
            log(f"Skipped {file_path.name}: not an .xlsx file")
            continue
        selected.append(file_path)
    return selected


def run() -> None:
    files = source_files(input_dir)
    output_path = output_path_for_run(input_dir, output_dir)

    processed_files = 0
    empirical_rows: list[dict[str, Any]] = []
    regression_rows: list[dict[str, Any]] = []

    app = xw.App(visible=False, add_book=False)
    app.display_alerts = False
    app.screen_updating = False
    try:
        app.calculation = "manual"
    except Exception:
        pass

    try:
        for file_path in files:
            label = parse_model_label(file_path)
            if label is None:
                log(f"Skipped {file_path.name}: filename pattern not recognized")
                continue

            log(f"Processing {file_path.name}")
            wb = None
            try:
                wb = app.books.open(str(file_path), update_links=False)
                empirical_rows.extend(extract_empirical_rows(wb, label, file_path.name))
                regression_rows.extend(extract_regression_rows(wb, label, file_path.name))
                processed_files += 1
            except Exception as exc:
                log(f"Skipped {file_path.name}: workbook processing error ({exc})")
            finally:
                if wb is not None:
                    safe_close_workbook(wb)
    finally:
        app.quit()

    write_output_workbook(output_path, empirical_rows, regression_rows)

    log(f"Output path: {output_path}")
    log(f"Files processed: {processed_files}")
    log(f"Empirical rows: {len(empirical_rows)}")
    log(f"Regression rows: {len(regression_rows)}")


if __name__ == "__main__":
    run()
