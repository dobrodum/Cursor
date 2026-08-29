from __future__ import annotations

import re
from datetime import date, datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import xlwings as xw
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

# ===== User-configurable paths =====
input_dir = Path("/path/to/input")
output_dir = Path("/path/to/output")


N_QUARTERS = 10
DAY_MAP = {"early": 5, "mid": 15, "late": 25}
FILE_PATTERN = re.compile(
    r"Model\s*-\s*([A-Za-z0-9]+)\s*-\s*(Early|Mid|Late)([A-Za-z]{3})(\d{4})",
    re.IGNORECASE,
)

EMPIRICAL_HEADERS = [
    "model",
    "ticker",
    "model_period",
    "model_date",
    "method",
    "parameter_name",
    "parameter_value",
    "num_quarters_used",
    "last_quarter_used",
    "forecast_value",
    "actual_value",
    "forecast_max",
    "forecast_min",
    "range_width",
    "avg_penetration_pct",
    "quarterly_sales",
    "reported_sales",
    "growth_rate_pct",
    "sales_captured_in_db_pct",
    "source_file",
]

REGRESSION_HEADERS = [
    "model",
    "ticker",
    "model_period",
    "model_date",
    "method",
    "parameter_name",
    "parameter_value",
    "num_quarters_used",
    "forecast_value",
    "actual_value",
    "forecast_max",
    "forecast_min",
    "range_width",
    "intercept",
    "slope",
    "source_file",
]

EMPIRICAL_DEFAULT_OFFSETS = {
    "num_quarters_used": -8,
    "last_quarter_used": -7,
    "avg_penetration_pct": -6,
    "quarterly_sales": -5,
    "reported_sales": -4,
    "growth_rate_pct": -3,
    "sales_captured_in_db_pct": -2,
    "forecast_value": -1,
    "forecast_max": 0,
    "forecast_min": 1,
    "actual_value": 2,
}

REGRESSION_DEFAULT_OFFSETS = {
    "num_quarters_used": -8,
    "forecast_value": -1,
    "forecast_max": 0,
    "forecast_min": 1,
    "actual_value": 2,
}

EMPIRICAL_ALIASES = {
    "num_quarters_used": ["num quarters used", "num quarters", "quarters used", "n quarters"],
    "last_quarter_used": ["last quarter used", "last quarter", "last qtr"],
    "forecast_value": ["estimated total sold", "forecast total", "tot fcst", "forecast"],
    "actual_value": ["actual value", "actual sales", "reported sales", "actual"],
    "forecast_max": ["max"],
    "forecast_min": ["min"],
    "avg_penetration_pct": ["avg penetration", "average penetration", "avg pen", "penetration"],
    "quarterly_sales": ["quarterly sales", "quarter sales", "q sales"],
    "reported_sales": ["reported sales", "reported"],
    "growth_rate_pct": ["growth rate", "growth pct", "growth"],
    "sales_captured_in_db_pct": [
        "sales captured in db",
        "captured in db",
        "db capture",
        "capture pct",
    ],
}

REGRESSION_ALIASES = {
    "num_quarters_used": ["num quarters used", "num quarters", "quarters used", "n quarters"],
    "forecast_value": [
        "tot fcst w/o sa",
        "tot fcst w o sa",
        "tot fcst without sa",
        "forecast total without sa",
    ],
    "actual_value": ["actual value", "actual sales", "actual", "reported sales"],
    "forecast_max": ["max"],
    "forecast_min": ["min"],
}


def normalize_text(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip().lower()
    text = text.replace("%", " pct ")
    text = re.sub(r"[^a-z0-9]+", " ", text)
    return " ".join(text.split())


def as_int(value: Any) -> Optional[int]:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return int(value)
    cleaned = str(value).strip().replace(",", "")
    if not cleaned:
        return None
    try:
        return int(float(cleaned))
    except ValueError:
        return None


def as_float(value: Any) -> Optional[float]:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    cleaned = str(value).strip().replace(",", "")
    if not cleaned or cleaned.startswith("#"):
        return None
    pct = cleaned.endswith("%")
    if pct:
        cleaned = cleaned[:-1]
    try:
        number = float(cleaned)
    except ValueError:
        return None
    return number / 100.0 if pct else number


def safe_subtract(left: Any, right: Any) -> Optional[float]:
    l_val = as_float(left)
    r_val = as_float(right)
    if l_val is None or r_val is None:
        return None
    return l_val - r_val


def close_workbook_no_save(wb: xw.Book) -> None:
    try:
        wb.close(save=False)
        return
    except TypeError:
        pass
    except Exception:
        pass
    try:
        wb.api.Close(SaveChanges=False)
    except Exception:
        try:
            wb.close()
        except Exception:
            pass


def parse_file_metadata(file_name: str) -> Dict[str, str]:
    stem = Path(file_name).stem
    match = FILE_PATTERN.search(stem)
    if not match:
        ticker_guess = stem.split("-")[1].strip().upper() if "-" in stem else stem.upper()
        return {
            "model": ticker_guess,
            "ticker": ticker_guess,
            "model_period": "unknown",
            "model_date": "",
        }

    ticker, period_prefix, month_text, year_text = match.groups()
    ticker = ticker.upper()
    period_prefix = period_prefix.title()
    month_text = month_text.title()
    year = int(year_text)
    month_num = datetime.strptime(month_text, "%b").month
    model_period = f"{period_prefix}{month_text}_{year_text}"
    model_day = DAY_MAP[period_prefix.lower()]
    model_date = date(year, month_num, model_day).isoformat()
    model = f"{ticker}_{model_period}"
    return {
        "model": model,
        "ticker": ticker,
        "model_period": model_period,
        "model_date": model_date,
    }


def get_output_path(in_dir: Path, out_dir: Path) -> Path:
    out_dir.mkdir(parents=True, exist_ok=True)
    root_name = f"{in_dir.name}_PARAM"
    candidate = out_dir / f"{root_name}.xlsx"
    if not candidate.exists():
        return candidate
    suffix = 1
    while True:
        candidate = out_dir / f"{root_name}.{suffix}.xlsx"
        if not candidate.exists():
            return candidate
        suffix += 1


def used_range_snapshot(sheet: xw.Sheet) -> Tuple[int, int, int, int, List[List[Any]]]:
    used = sheet.used_range
    top = used.row
    left = used.column
    values = used.options(ndim=2).value
    if values is None:
        values = [[]]
    if not values:
        values = [[]]
    row_count = len(values)
    col_count = len(values[0]) if values and values[0] is not None else 1
    bottom = top + row_count - 1
    right = left + col_count - 1
    return top, left, bottom, right, values


def snapshot_cell(
    values: List[List[Any]], top: int, left: int, row: int, col: int
) -> Any:
    row_idx = row - top
    col_idx = col - left
    if row_idx < 0 or col_idx < 0:
        return None
    if row_idx >= len(values):
        return None
    if col_idx >= len(values[row_idx]):
        return None
    return values[row_idx][col_idx]


def find_anchor(
    values: List[List[Any]], top: int, left: int, bottom: int, right: int
) -> Tuple[int, int]:
    for row in range(top, bottom + 1):
        for col in range(left, right + 1):
            value = snapshot_cell(values, top, left, row, col)
            if isinstance(value, str) and value.strip().lower() == "max":
                return row, col
    raise ValueError('Anchor cell "max" not found')


def choose_header_row(
    values: List[List[Any]],
    top: int,
    left: int,
    bottom: int,
    right: int,
    anchor_row: int,
    alias_map: Dict[str, List[str]],
) -> int:
    normalized_aliases = {
        key: [normalize_text(alias) for alias in aliases]
        for key, aliases in alias_map.items()
    }
    best_row = anchor_row
    best_score = -1
    row_start = max(top, anchor_row - 5)
    row_end = min(bottom, anchor_row + 5)
    for row in range(row_start, row_end + 1):
        score = 0
        for col in range(left, right + 1):
            header = normalize_text(snapshot_cell(values, top, left, row, col))
            if not header:
                continue
            for aliases in normalized_aliases.values():
                if any(alias and alias in header for alias in aliases):
                    score += 1
                    break
        if score > best_score:
            best_score = score
            best_row = row
    return best_row


def build_column_map(
    values: List[List[Any]],
    top: int,
    left: int,
    right: int,
    header_row: int,
    alias_map: Dict[str, List[str]],
) -> Dict[str, int]:
    normalized_aliases = {
        key: [normalize_text(alias) for alias in aliases]
        for key, aliases in alias_map.items()
    }
    col_map: Dict[str, int] = {}
    for col in range(left, right + 1):
        header = normalize_text(snapshot_cell(values, top, left, header_row, col))
        if not header:
            continue
        for key, aliases in normalized_aliases.items():
            if key in col_map:
                continue
            if any(alias and alias in header for alias in aliases):
                col_map[key] = col
    return col_map


def resolve_column(
    col_map: Dict[str, int], key: str, anchor_col: int, default_offsets: Dict[str, int]
) -> int:
    if key in col_map:
        return col_map[key]
    return anchor_col + default_offsets[key]


def collect_candidate_rows(
    values: List[List[Any]],
    top: int,
    left: int,
    bottom: int,
    header_row: int,
    num_quarters_col: int,
) -> List[int]:
    rows: List[int] = []
    blank_streak = 0
    for row in range(header_row + 1, bottom + 1):
        candidate = snapshot_cell(values, top, left, row, num_quarters_col)
        n_quarters = as_int(candidate)
        if n_quarters is None:
            blank_streak += 1
            if rows and blank_streak >= 3:
                break
            continue
        blank_streak = 0
        rows.append(row)
        if len(rows) >= N_QUARTERS:
            break

    if rows:
        return rows
    fallback_start = max(header_row + 1, top)
    fallback_end = min(fallback_start + N_QUARTERS - 1, bottom)
    return list(range(fallback_start, fallback_end + 1))


def set_formula2(cell: xw.Range, formula: str) -> None:
    try:
        cell.formula2 = formula
    except Exception:
        cell.formula = formula


def pull_formula_column(
    sheet: xw.Sheet,
    rows: List[int],
    target_col: int,
) -> Dict[int, Any]:
    if not rows:
        return {}
    first_row = min(rows)
    last_row = max(rows)
    values = sheet.range((first_row, target_col), (last_row, target_col)).options(ndim=2).value
    by_row: Dict[int, Any] = {}
    for row in rows:
        by_row[row] = values[row - first_row][0]
    return by_row


def clear_column_range(sheet: xw.Sheet, first_row: int, last_row: int, col: int) -> None:
    sheet.range((first_row, col), (last_row, col)).value = None


def process_empirical_sheet(
    wb: xw.Book, metadata: Dict[str, str], source_file: str
) -> List[Dict[str, Any]]:
    sheet = wb.sheets["Empirical Model"]
    top, left, bottom, right, values = used_range_snapshot(sheet)
    anchor_row, anchor_col = find_anchor(values, top, left, bottom, right)
    header_row = choose_header_row(
        values, top, left, bottom, right, anchor_row, EMPIRICAL_ALIASES
    )
    col_map = build_column_map(values, top, left, right, header_row, EMPIRICAL_ALIASES)

    num_col = resolve_column(
        col_map, "num_quarters_used", anchor_col, EMPIRICAL_DEFAULT_OFFSETS
    )
    rows = collect_candidate_rows(values, top, left, bottom, header_row, num_col)
    if not rows:
        return []

    avg_source_col = col_map.get("sales_captured_in_db_pct") or col_map.get("avg_penetration_pct")
    if avg_source_col is None:
        avg_source_col = resolve_column(
            col_map, "avg_penetration_pct", anchor_col, EMPIRICAL_DEFAULT_OFFSETS
        )

    scratch_col = right + 3
    first_row = min(rows)
    last_row = max(rows)
    for row in rows:
        n_quarters = as_int(snapshot_cell(values, top, left, row, num_col)) or 1
        start_row = max(header_row + 1, row - n_quarters + 1)
        formula = f"=AVERAGE(R{start_row}C{avg_source_col}:R{row}C{avg_source_col})"
        set_formula2(sheet.range((row, scratch_col)), formula)
    wb.app.calculate()
    avg_penetration_by_row = pull_formula_column(sheet, rows, scratch_col)
    clear_column_range(sheet, first_row, last_row, scratch_col)

    resolved_cols = {
        key: resolve_column(col_map, key, anchor_col, EMPIRICAL_DEFAULT_OFFSETS)
        for key in EMPIRICAL_DEFAULT_OFFSETS
    }

    output_rows: List[Dict[str, Any]] = []
    for row in rows:
        num_quarters_used = as_int(snapshot_cell(values, top, left, row, resolved_cols["num_quarters_used"]))
        last_quarter_used = snapshot_cell(values, top, left, row, resolved_cols["last_quarter_used"])
        forecast_value = snapshot_cell(values, top, left, row, resolved_cols["forecast_value"])
        forecast_max = snapshot_cell(values, top, left, row, resolved_cols["forecast_max"])
        forecast_min = snapshot_cell(values, top, left, row, resolved_cols["forecast_min"])
        reported_sales = snapshot_cell(values, top, left, row, resolved_cols["reported_sales"])
        actual_value = snapshot_cell(values, top, left, row, resolved_cols["actual_value"])
        if actual_value in (None, ""):
            actual_value = reported_sales

        avg_penetration_pct = avg_penetration_by_row.get(row)
        output_rows.append(
            {
                "model": metadata["model"],
                "ticker": metadata["ticker"],
                "model_period": metadata["model_period"],
                "model_date": metadata["model_date"],
                "method": "empirical",
                "parameter_name": "avg_penetration_pct",
                "parameter_value": avg_penetration_pct,
                "num_quarters_used": num_quarters_used,
                "last_quarter_used": last_quarter_used,
                "forecast_value": forecast_value,
                "actual_value": actual_value,
                "forecast_max": forecast_max,
                "forecast_min": forecast_min,
                "range_width": safe_subtract(forecast_max, forecast_min),
                "avg_penetration_pct": avg_penetration_pct,
                "quarterly_sales": snapshot_cell(
                    values, top, left, row, resolved_cols["quarterly_sales"]
                ),
                "reported_sales": reported_sales,
                "growth_rate_pct": snapshot_cell(
                    values, top, left, row, resolved_cols["growth_rate_pct"]
                ),
                "sales_captured_in_db_pct": snapshot_cell(
                    values, top, left, row, resolved_cols["sales_captured_in_db_pct"]
                ),
                "source_file": source_file,
            }
        )
    return output_rows


def almost_equal(left: Any, right: Any, tolerance: float = 1e-9) -> bool:
    left_float = as_float(left)
    right_float = as_float(right)
    if left_float is not None and right_float is not None:
        scale = max(1.0, abs(left_float), abs(right_float))
        return abs(left_float - right_float) <= tolerance * scale
    return left == right


def process_regression_sheet(
    wb: xw.Book, metadata: Dict[str, str], source_file: str
) -> List[Dict[str, Any]]:
    sheet = wb.sheets["Regression Model"]
    top, left, bottom, right, values = used_range_snapshot(sheet)
    anchor_row, anchor_col = find_anchor(values, top, left, bottom, right)
    header_row = choose_header_row(
        values, top, left, bottom, right, anchor_row, REGRESSION_ALIASES
    )
    col_map = build_column_map(values, top, left, right, header_row, REGRESSION_ALIASES)

    y_col = anchor_col - 7
    x_col = anchor_col - 11
    num_col = resolve_column(
        col_map, "num_quarters_used", anchor_col, REGRESSION_DEFAULT_OFFSETS
    )
    rows = collect_candidate_rows(values, top, left, bottom, header_row, num_col)
    if not rows:
        return []

    scratch_intercept_col = right + 3
    scratch_slope_col = right + 4
    first_row = min(rows)
    last_row = max(rows)
    for row in rows:
        n_quarters = as_int(snapshot_cell(values, top, left, row, num_col)) or 1
        start_row = max(header_row + 1, row - n_quarters + 1)
        if row - start_row < 1:
            intercept_formula = "=NA()"
            slope_formula = "=NA()"
        else:
            intercept_formula = (
                f"=INTERCEPT(R{start_row}C{y_col}:R{row}C{y_col},R{start_row}C{x_col}:R{row}C{x_col})"
            )
            slope_formula = (
                f"=SLOPE(R{start_row}C{y_col}:R{row}C{y_col},R{start_row}C{x_col}:R{row}C{x_col})"
            )
        set_formula2(sheet.range((row, scratch_intercept_col)), intercept_formula)
        set_formula2(sheet.range((row, scratch_slope_col)), slope_formula)
    wb.app.calculate()
    intercept_by_row = pull_formula_column(sheet, rows, scratch_intercept_col)
    slope_by_row = pull_formula_column(sheet, rows, scratch_slope_col)
    clear_column_range(sheet, first_row, last_row, scratch_intercept_col)
    clear_column_range(sheet, first_row, last_row, scratch_slope_col)

    resolved_cols = {
        key: resolve_column(col_map, key, anchor_col, REGRESSION_DEFAULT_OFFSETS)
        for key in REGRESSION_DEFAULT_OFFSETS
    }

    output_rows: List[Dict[str, Any]] = []
    for idx, row in enumerate(rows):
        num_quarters_used = as_int(snapshot_cell(values, top, left, row, resolved_cols["num_quarters_used"]))
        forecast_value = snapshot_cell(values, top, left, row, resolved_cols["forecast_value"])
        actual_value = snapshot_cell(values, top, left, row, resolved_cols["actual_value"])
        forecast_max = snapshot_cell(values, top, left, row, resolved_cols["forecast_max"])
        forecast_min = snapshot_cell(values, top, left, row, resolved_cols["forecast_min"])
        candidate = {
            "model": metadata["model"],
            "ticker": metadata["ticker"],
            "model_period": metadata["model_period"],
            "model_date": metadata["model_date"],
            "method": "regression",
            "parameter_name": "num_quarters_used",
            "parameter_value": num_quarters_used,
            "num_quarters_used": num_quarters_used,
            "forecast_value": forecast_value,
            "actual_value": actual_value,
            "forecast_max": forecast_max,
            "forecast_min": forecast_min,
            "range_width": safe_subtract(forecast_max, forecast_min),
            "intercept": intercept_by_row.get(row),
            "slope": slope_by_row.get(row),
            "source_file": source_file,
        }

        is_last = idx == len(rows) - 1
        if is_last and output_rows:
            prev = output_rows[-1]
            duplicate = (
                almost_equal(candidate["forecast_value"], prev["forecast_value"])
                and almost_equal(candidate["forecast_max"], prev["forecast_max"])
                and almost_equal(candidate["forecast_min"], prev["forecast_min"])
                and almost_equal(candidate["intercept"], prev["intercept"])
                and almost_equal(candidate["slope"], prev["slope"])
            )
            if duplicate:
                continue
        output_rows.append(candidate)

    return output_rows


def add_sheet(
    wb: Workbook, sheet_name: str, headers: List[str], rows: List[Dict[str, Any]]
) -> None:
    ws = wb.create_sheet(title=sheet_name)
    ws.append(headers)
    for row in rows:
        ws.append([row.get(header, "") for header in headers])

    for cell in ws[1]:
        cell.font = Font(bold=True)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    for col_idx, header in enumerate(headers, start=1):
        max_len = len(header)
        for row_idx in range(2, ws.max_row + 1):
            value = ws.cell(row=row_idx, column=col_idx).value
            if value is None:
                continue
            max_len = max(max_len, len(str(value)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 42)


def write_output(
    out_path: Path, empirical_rows: List[Dict[str, Any]], regression_rows: List[Dict[str, Any]]
) -> None:
    out_wb = Workbook()
    default_sheet = out_wb.active
    out_wb.remove(default_sheet)

    add_sheet(out_wb, "empirical_candidates", EMPIRICAL_HEADERS, empirical_rows)
    add_sheet(out_wb, "regression_candidates", REGRESSION_HEADERS, regression_rows)
    out_wb.save(out_path)


def should_skip_file(file_path: Path, input_folder_name: str) -> Optional[str]:
    if not file_path.is_file():
        return "not a file"
    if file_path.name.startswith("~"):
        return "temporary Excel lock file"
    if file_path.suffix.lower() != ".xlsx":
        return "not an .xlsx file"
    if file_path.name.startswith(f"{input_folder_name}_PARAM"):
        return "generated PARAM output file"
    return None


def main() -> None:
    in_dir = Path(input_dir).expanduser().resolve()
    out_dir = Path(output_dir).expanduser().resolve()
    if not in_dir.exists():
        raise FileNotFoundError(f"input_dir does not exist: {in_dir}")

    output_path = get_output_path(in_dir, out_dir)
    folder_name = in_dir.name

    empirical_rows: List[Dict[str, Any]] = []
    regression_rows: List[Dict[str, Any]] = []
    processed_files = 0

    app = xw.App(visible=False, add_book=False)
    app.display_alerts = False
    app.screen_updating = False
    app.calculation = "manual"

    try:
        for file_path in sorted(in_dir.iterdir()):
            skip_reason = should_skip_file(file_path, folder_name)
            if skip_reason:
                print(f"Skipped: {file_path.name} ({skip_reason})")
                continue

            print(f"Processed: {file_path.name}")
            metadata = parse_file_metadata(file_path.name)
            wb = None
            try:
                wb = app.books.open(str(file_path), update_links=False)

                try:
                    empirical_rows.extend(
                        process_empirical_sheet(wb, metadata, file_path.name)
                    )
                except KeyError:
                    print(f"Skipped empirical extraction in {file_path.name}: sheet not found")
                except Exception as exc:
                    print(f"Skipped empirical extraction in {file_path.name}: {exc}")

                try:
                    regression_rows.extend(
                        process_regression_sheet(wb, metadata, file_path.name)
                    )
                except KeyError:
                    print(f"Skipped regression extraction in {file_path.name}: sheet not found")
                except Exception as exc:
                    print(f"Skipped regression extraction in {file_path.name}: {exc}")

                processed_files += 1
            except Exception as exc:
                print(f"Skipped: {file_path.name} (open failed: {exc})")
            finally:
                if wb is not None:
                    close_workbook_no_save(wb)
    finally:
        app.quit()

    write_output(output_path, empirical_rows, regression_rows)

    print(f"Output: {output_path}")
    print(f"Files processed: {processed_files}")
    print(f"Empirical rows: {len(empirical_rows)}")
    print(f"Regression rows: {len(regression_rows)}")


if __name__ == "__main__":
    main()
