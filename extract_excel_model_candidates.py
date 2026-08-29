#!/usr/bin/env python3
"""Extract empirical and regression candidate parameters from Excel model files.

Design goals:
- Open each source workbook once.
- Process both model sheets while the workbook is open.
- Never save or modify source files on disk.
- Write one formatted output workbook with two sheets:
  - empirical_candidates
  - regression_candidates
"""

from __future__ import annotations

import re
from dataclasses import dataclass
from datetime import date
from pathlib import Path
from typing import Any, Iterable, Sequence

import xlwings as xw
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# User-configurable paths
# ---------------------------------------------------------------------------
input_dir = "/path/to/input"
output_dir = "/path/to/output"

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------
N_QUARTERS = 10
EMPIRICAL_SHEET_NAME = "Empirical Model"
REGRESSION_SHEET_NAME = "Regression Model"

PERIOD_DAY_MAP = {"early": 5, "mid": 15, "late": 25}
MONTH_MAP = {
    "jan": 1,
    "feb": 2,
    "mar": 3,
    "apr": 4,
    "may": 5,
    "jun": 6,
    "jul": 7,
    "aug": 8,
    "sep": 9,
    "oct": 10,
    "nov": 11,
    "dec": 12,
}

# Anchor-relative offsets from the "max" column.
EMPIRICAL_DEFAULT_OFFSETS = {
    "sales_captured_in_db_pct": -8,
    "growth_rate_pct": -7,
    "quarterly_sales": -6,
    "num_quarters_used": -5,
    "last_quarter_used": -4,
    "avg_penetration_pct": -3,
    "reported_sales": -2,
    "forecast_value": -1,
    "forecast_max": 0,
    "forecast_min": 1,
}

REGRESSION_DEFAULT_OFFSETS = {
    "num_quarters_used": -5,
    "actual_value": -2,
    "forecast_value": -1,  # TOT FCST w/o SA
    "forecast_max": 0,
    "forecast_min": 1,
}

EMPIRICAL_COLUMNS = [
    "model",
    "ticker",
    "model_period",
    "model_date",
    "method",
    "parameter_name",
    "parameter_value",
    "num_quarters_used",
    "last_quarter_used",
    "forecast_value",
    "actual_value",
    "forecast_max",
    "forecast_min",
    "range_width",
    "avg_penetration_pct",
    "quarterly_sales",
    "reported_sales",
    "growth_rate_pct",
    "sales_captured_in_db_pct",
    "source_file",
]

REGRESSION_COLUMNS = [
    "model",
    "ticker",
    "model_period",
    "model_date",
    "method",
    "parameter_name",
    "parameter_value",
    "num_quarters_used",
    "forecast_value",
    "actual_value",
    "forecast_max",
    "forecast_min",
    "range_width",
    "intercept",
    "slope",
    "source_file",
]


@dataclass(frozen=True)
class ModelMetadata:
    model: str
    ticker: str
    model_period: str
    model_date: str


@dataclass(frozen=True)
class SheetSnapshot:
    start_row: int
    start_col: int
    end_row: int
    end_col: int
    values: list[list[Any]]

    def get(self, row: int, col: int) -> Any:
        if row < self.start_row or col < self.start_col:
            return None
        r_idx = row - self.start_row
        c_idx = col - self.start_col
        if r_idx < 0 or c_idx < 0 or r_idx >= len(self.values):
            return None
        row_values = self.values[r_idx]
        if c_idx >= len(row_values):
            return None
        value = row_values[c_idx]
        if value == "":
            return None
        return value


def ensure_2d(values: Any) -> list[list[Any]]:
    if values is None:
        return []
    if isinstance(values, tuple):
        values = list(values)
    if not isinstance(values, list):
        return [[values]]
    if not values:
        return []
    if isinstance(values[0], tuple):
        return [list(row) for row in values]
    if isinstance(values[0], list):
        return values
    return [values]


def normalize_text(value: Any) -> str:
    if value is None:
        return ""
    return re.sub(r"[^a-z0-9]+", "", str(value).strip().lower())


def is_blank(value: Any) -> bool:
    return value is None or (isinstance(value, str) and value.strip() == "")


def to_float(value: Any) -> float | None:
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        text = value.strip().replace(",", "")
        if not text:
            return None
        if text.endswith("%"):
            text = text[:-1]
            try:
                return float(text) / 100.0
            except ValueError:
                return None
        try:
            return float(text)
        except ValueError:
            return None
    return None


def numeric_difference(left: Any, right: Any) -> float | None:
    left_num = to_float(left)
    right_num = to_float(right)
    if left_num is None or right_num is None:
        return None
    return left_num - right_num


def value_signature(value: Any) -> Any:
    number_value = to_float(value)
    if number_value is not None:
        return round(number_value, 10)
    if is_blank(value):
        return None
    return str(value).strip()


def set_r1c1_formula2(cell: Any, formula_r1c1: str) -> None:
    """Set R1C1 formula with Formula2 first, then safe fallbacks."""
    try:
        cell.formula2 = formula_r1c1
        return
    except Exception:
        pass
    try:
        cell.api.Formula2R1C1 = formula_r1c1
        return
    except Exception:
        pass
    try:
        cell.api.FormulaR1C1 = formula_r1c1
        return
    except Exception:
        pass
    cell.formula = formula_r1c1


def recalculate(workbook: xw.Book) -> None:
    try:
        workbook.app.calculate()
    except Exception:
        workbook.app.api.Calculate()


def close_source_workbook(workbook: xw.Book | None) -> None:
    """Close source workbook without saving, with safe fallbacks."""
    if workbook is None:
        return
    try:
        workbook.close(save=False)
        return
    except TypeError:
        pass
    except Exception:
        pass
    try:
        workbook.api.Close(SaveChanges=False)
        return
    except Exception:
        pass
    try:
        workbook.api.Close(False)
        return
    except Exception:
        pass
    try:
        workbook.close()
    except Exception:
        pass


def parse_file_metadata(file_path: Path) -> ModelMetadata:
    """Parse labels from filename.

    Example:
    MedMiner_Model - AORT - MidJan2026_Send.xlsx
    """
    match = re.search(
        r"-\s*(?P<ticker>[A-Za-z0-9]+)\s*-\s*(?P<period>Early|Mid|Late)(?P<month>[A-Za-z]+)(?P<year>\d{4})",
        file_path.stem,
        flags=re.IGNORECASE,
    )
    if not match:
        raise ValueError("filename does not match expected pattern")

    ticker = match.group("ticker").upper()
    period_token = match.group("period").lower()
    month_key = match.group("month")[:3].lower()
    year_num = int(match.group("year"))

    if month_key not in MONTH_MAP:
        raise ValueError(f"unrecognized month token: {match.group('month')}")

    day = PERIOD_DAY_MAP[period_token]
    month_num = MONTH_MAP[month_key]

    model_period = f"{period_token.title()}{month_key.title()}_{year_num}"
    model_date = date(year_num, month_num, day).isoformat()
    model = f"{ticker}_{model_period}"

    return ModelMetadata(
        model=model,
        ticker=ticker,
        model_period=model_period,
        model_date=model_date,
    )


def choose_output_path(input_path: Path, output_path: Path) -> Path:
    input_folder_name = input_path.name
    base = output_path / f"{input_folder_name}_PARAM.xlsx"
    if not base.exists():
        return base
    suffix = 1
    while True:
        candidate = output_path / f"{input_folder_name}_PARAM.{suffix}.xlsx"
        if not candidate.exists():
            return candidate
        suffix += 1


def should_skip_file(file_path: Path, input_folder_name: str) -> str | None:
    if not file_path.is_file():
        return "not a file"
    if file_path.name.startswith("~"):
        return "temporary lock file"
    if file_path.suffix.lower() != ".xlsx":
        return "not an .xlsx file"
    if re.match(
        rf"^{re.escape(input_folder_name)}_PARAM(\.\d+)?\.xlsx$",
        file_path.name,
        flags=re.IGNORECASE,
    ):
        return "looks like previously generated output"
    return None


def build_sheet_snapshot(sheet: xw.Sheet) -> SheetSnapshot:
    used = sheet.used_range
    values = ensure_2d(used.value)
    if not values:
        return SheetSnapshot(
            start_row=used.row,
            start_col=used.column,
            end_row=used.row,
            end_col=used.column,
            values=[],
        )

    max_cols = max(len(row) for row in values)
    normalized: list[list[Any]] = []
    for row in values:
        normalized.append(list(row) + [None] * (max_cols - len(row)))

    return SheetSnapshot(
        start_row=used.row,
        start_col=used.column,
        end_row=used.row + len(normalized) - 1,
        end_col=used.column + max_cols - 1,
        values=normalized,
    )


def find_anchor(snapshot: SheetSnapshot, anchor_text: str = "max") -> tuple[int, int] | None:
    target = anchor_text.strip().lower()
    for r_idx, row in enumerate(snapshot.values):
        for c_idx, value in enumerate(row):
            if isinstance(value, str) and value.strip().lower() == target:
                return snapshot.start_row + r_idx, snapshot.start_col + c_idx
    return None


def get_header_cells(
    sheet: xw.Sheet,
    header_row: int,
    anchor_col: int,
    span: int = 35,
) -> list[tuple[int, str]]:
    start_col = max(1, anchor_col - span)
    end_col = anchor_col + span
    row_values = sheet.range((header_row, start_col), (header_row, end_col)).value
    if isinstance(row_values, tuple):
        row_values = list(row_values)
    if not isinstance(row_values, list):
        row_values = [row_values]
    return [(start_col + idx, normalize_text(value)) for idx, value in enumerate(row_values)]


def pick_column_near_anchor(
    header_cells: Sequence[tuple[int, str]],
    anchor_col: int,
    keyword_patterns: Sequence[str],
) -> int | None:
    candidates: list[int] = []
    for col, normalized_header in header_cells:
        if any(pattern in normalized_header for pattern in keyword_patterns):
            candidates.append(col)
    if not candidates:
        return None
    return min(candidates, key=lambda col: abs(col - anchor_col))


def column_has_any_data(sheet: xw.Sheet, col: int, start_row: int, end_row: int) -> bool:
    if start_row > end_row:
        return False
    values = sheet.range((start_row, col), (end_row, col)).value
    if isinstance(values, tuple):
        values = list(values)
    if not isinstance(values, list):
        values = [values]
    for value in values:
        if isinstance(value, list):
            if any(not is_blank(item) for item in value):
                return True
        elif not is_blank(value):
            return True
    return False


def choose_column(
    sheet: xw.Sheet,
    header_cells: Sequence[tuple[int, str]],
    anchor_col: int,
    start_row: int,
    end_row: int,
    default_offset: int,
    keyword_patterns: Sequence[str],
) -> int:
    default_col = anchor_col + default_offset
    if column_has_any_data(sheet, default_col, start_row, end_row):
        return default_col

    header_col = pick_column_near_anchor(header_cells, anchor_col, keyword_patterns)
    if header_col is not None:
        return header_col
    return default_col


def read_block(
    sheet: xw.Sheet,
    start_row: int,
    end_row: int,
    start_col: int,
    end_col: int,
) -> list[list[Any]]:
    values = sheet.range((start_row, start_col), (end_row, end_col)).options(ndim=2).value
    return ensure_2d(values)


def block_value(block: Sequence[Sequence[Any]], row_idx: int, abs_col: int, min_col: int) -> Any:
    col_idx = abs_col - min_col
    if row_idx < 0 or row_idx >= len(block):
        return None
    row = block[row_idx]
    if col_idx < 0 or col_idx >= len(row):
        return None
    value = row[col_idx]
    if value == "":
        return None
    return value


def collect_numeric_rows_single_col(sheet: xw.Sheet, col: int, max_row: int) -> list[int]:
    if max_row < 1:
        return []
    start_row = max(1, sheet.used_range.row)
    if max_row < start_row:
        return []
    values = sheet.range((start_row, col), (max_row, col)).value
    if isinstance(values, tuple):
        values = list(values)
    if not isinstance(values, list):
        values = [values]
    rows: list[int] = []
    for idx, value in enumerate(values):
        if to_float(value) is not None:
            rows.append(start_row + idx)
    return rows


def collect_numeric_rows_pair(sheet: xw.Sheet, x_col: int, y_col: int, max_row: int) -> list[int]:
    if max_row < 1:
        return []
    start_row = max(1, sheet.used_range.row)
    if max_row < start_row:
        return []

    x_values = sheet.range((start_row, x_col), (max_row, x_col)).value
    y_values = sheet.range((start_row, y_col), (max_row, y_col)).value
    if isinstance(x_values, tuple):
        x_values = list(x_values)
    if isinstance(y_values, tuple):
        y_values = list(y_values)
    if not isinstance(x_values, list):
        x_values = [x_values]
    if not isinstance(y_values, list):
        y_values = [y_values]

    rows: list[int] = []
    for idx, (x_value, y_value) in enumerate(zip(x_values, y_values)):
        if to_float(x_value) is not None and to_float(y_value) is not None:
            rows.append(start_row + idx)
    return rows


def extract_empirical_candidates(
    workbook: xw.Book,
    metadata: ModelMetadata,
    source_file: str,
) -> list[dict[str, Any]]:
    sheet_names = [sheet.name for sheet in workbook.sheets]
    if EMPIRICAL_SHEET_NAME not in sheet_names:
        print(f"skipped empirical: {source_file} (sheet '{EMPIRICAL_SHEET_NAME}' not found)")
        return []

    sheet = workbook.sheets[EMPIRICAL_SHEET_NAME]
    snapshot = build_sheet_snapshot(sheet)
    anchor = find_anchor(snapshot, "max")
    if anchor is None:
        print(f"skipped empirical: {source_file} ('max' anchor not found)")
        return []
    anchor_row, anchor_col = anchor

    first_data_row = anchor_row + 1
    last_data_row = first_data_row + N_QUARTERS - 1
    header_cells = get_header_cells(sheet, anchor_row, anchor_col)

    col_map = {
        "num_quarters_used": choose_column(
            sheet,
            header_cells,
            anchor_col,
            first_data_row,
            last_data_row,
            EMPIRICAL_DEFAULT_OFFSETS["num_quarters_used"],
            ["numquartersused", "quartersused", "nquarters", "numqtrs", "numquarters"],
        ),
        "last_quarter_used": choose_column(
            sheet,
            header_cells,
            anchor_col,
            first_data_row,
            last_data_row,
            EMPIRICAL_DEFAULT_OFFSETS["last_quarter_used"],
            ["lastquarterused", "lastquarter", "latestquarter"],
        ),
        "avg_penetration_pct": choose_column(
            sheet,
            header_cells,
            anchor_col,
            first_data_row,
            last_data_row,
            EMPIRICAL_DEFAULT_OFFSETS["avg_penetration_pct"],
            ["avgpenetrationpct", "avgpenetration", "averagepenetration", "penetrationpct"],
        ),
        "forecast_value": choose_column(
            sheet,
            header_cells,
            anchor_col,
            first_data_row,
            last_data_row,
            EMPIRICAL_DEFAULT_OFFSETS["forecast_value"],
            ["estimatedtotalsold", "esttotalsold", "forecastvalue", "forecasttotalsold"],
        ),
        "reported_sales": choose_column(
            sheet,
            header_cells,
            anchor_col,
            first_data_row,
            last_data_row,
            EMPIRICAL_DEFAULT_OFFSETS["reported_sales"],
            ["reportedsales", "actualsales", "salesreported"],
        ),
        "quarterly_sales": choose_column(
            sheet,
            header_cells,
            anchor_col,
            first_data_row,
            last_data_row,
            EMPIRICAL_DEFAULT_OFFSETS["quarterly_sales"],
            ["quarterlysales", "salesquarter", "quartersales"],
        ),
        "growth_rate_pct": choose_column(
            sheet,
            header_cells,
            anchor_col,
            first_data_row,
            last_data_row,
            EMPIRICAL_DEFAULT_OFFSETS["growth_rate_pct"],
            ["growthratepct", "growthrate", "growthpct"],
        ),
        "sales_captured_in_db_pct": choose_column(
            sheet,
            header_cells,
            anchor_col,
            first_data_row,
            last_data_row,
            EMPIRICAL_DEFAULT_OFFSETS["sales_captured_in_db_pct"],
            ["salescapturedindbpct", "capturedindb", "dbcapturepct", "salescapturedpct"],
        ),
        "forecast_max": choose_column(
            sheet,
            header_cells,
            anchor_col,
            first_data_row,
            last_data_row,
            EMPIRICAL_DEFAULT_OFFSETS["forecast_max"],
            ["max", "forecastmax"],
        ),
        "forecast_min": choose_column(
            sheet,
            header_cells,
            anchor_col,
            first_data_row,
            last_data_row,
            EMPIRICAL_DEFAULT_OFFSETS["forecast_min"],
            ["min", "forecastmin"],
        ),
    }

    penetration_source_col = choose_column(
        sheet,
        header_cells,
        anchor_col,
        max(1, snapshot.start_row),
        anchor_row - 1,
        EMPIRICAL_DEFAULT_OFFSETS["sales_captured_in_db_pct"],
        ["salescapturedindbpct", "penetrationpct", "penetration"],
    )

    formulas_written = False
    avg_pen_col = col_map["avg_penetration_pct"]
    hist_rows = collect_numeric_rows_single_col(sheet, penetration_source_col, anchor_row - 1)
    if hist_rows:
        for idx in range(N_QUARTERS):
            n_use = min(idx + 1, len(hist_rows))
            first_hist_row = hist_rows[-n_use]
            last_hist_row = hist_rows[-1]
            formula = (
                f'=IFERROR(AVERAGE(R{first_hist_row}C{penetration_source_col}:'
                f'R{last_hist_row}C{penetration_source_col}),"")'
            )
            target_cell = sheet.range((first_data_row + idx, avg_pen_col))
            set_r1c1_formula2(target_cell, formula)
            formulas_written = True

    if formulas_written:
        recalculate(workbook)

    tracked_cols = set(col_map.values())
    min_col = min(tracked_cols)
    max_col = max(tracked_cols)
    block = read_block(sheet, first_data_row, last_data_row, min_col, max_col)

    rows: list[dict[str, Any]] = []
    for idx in range(N_QUARTERS):
        forecast_max = block_value(block, idx, col_map["forecast_max"], min_col)
        forecast_min = block_value(block, idx, col_map["forecast_min"], min_col)
        forecast_value = block_value(block, idx, col_map["forecast_value"], min_col)
        avg_penetration = block_value(block, idx, col_map["avg_penetration_pct"], min_col)

        if (
            is_blank(forecast_max)
            and is_blank(forecast_min)
            and is_blank(forecast_value)
            and is_blank(avg_penetration)
        ):
            continue

        raw_num_quarters = block_value(block, idx, col_map["num_quarters_used"], min_col)
        num_quarters = int(round(to_float(raw_num_quarters))) if to_float(raw_num_quarters) is not None else idx + 1

        reported_sales = block_value(block, idx, col_map["reported_sales"], min_col)

        rows.append(
            {
                "model": metadata.model,
                "ticker": metadata.ticker,
                "model_period": metadata.model_period,
                "model_date": metadata.model_date,
                "method": "empirical",
                "parameter_name": "avg_penetration_pct",
                "parameter_value": avg_penetration,
                "num_quarters_used": num_quarters,
                "last_quarter_used": block_value(block, idx, col_map["last_quarter_used"], min_col),
                "forecast_value": forecast_value,
                "actual_value": reported_sales,
                "forecast_max": forecast_max,
                "forecast_min": forecast_min,
                "range_width": numeric_difference(forecast_max, forecast_min),
                "avg_penetration_pct": avg_penetration,
                "quarterly_sales": block_value(block, idx, col_map["quarterly_sales"], min_col),
                "reported_sales": reported_sales,
                "growth_rate_pct": block_value(block, idx, col_map["growth_rate_pct"], min_col),
                "sales_captured_in_db_pct": block_value(
                    block,
                    idx,
                    col_map["sales_captured_in_db_pct"],
                    min_col,
                ),
                "source_file": source_file,
            }
        )

    return rows


def extract_regression_candidates(
    workbook: xw.Book,
    metadata: ModelMetadata,
    source_file: str,
) -> list[dict[str, Any]]:
    sheet_names = [sheet.name for sheet in workbook.sheets]
    if REGRESSION_SHEET_NAME not in sheet_names:
        print(f"skipped regression: {source_file} (sheet '{REGRESSION_SHEET_NAME}' not found)")
        return []

    sheet = workbook.sheets[REGRESSION_SHEET_NAME]
    snapshot = build_sheet_snapshot(sheet)
    anchor = find_anchor(snapshot, "max")
    if anchor is None:
        print(f"skipped regression: {source_file} ('max' anchor not found)")
        return []
    anchor_row, anchor_col = anchor

    first_data_row = anchor_row + 1
    last_data_row = first_data_row + N_QUARTERS - 1
    y_col = anchor_col - 7
    x_col = anchor_col - 11

    header_cells = get_header_cells(sheet, anchor_row, anchor_col)
    col_map = {
        "num_quarters_used": choose_column(
            sheet,
            header_cells,
            anchor_col,
            first_data_row,
            last_data_row,
            REGRESSION_DEFAULT_OFFSETS["num_quarters_used"],
            ["numquartersused", "quartersused", "nquarters", "numqtrs", "numquarters"],
        ),
        "forecast_value": choose_column(
            sheet,
            header_cells,
            anchor_col,
            first_data_row,
            last_data_row,
            REGRESSION_DEFAULT_OFFSETS["forecast_value"],
            [
                "totfcstwosa",
                "totfcstwithoutsa",
                "forecasttotalwithoutsa",
                "forecastvalue",
                "totforecastwosa",
            ],
        ),
        "actual_value": choose_column(
            sheet,
            header_cells,
            anchor_col,
            first_data_row,
            last_data_row,
            REGRESSION_DEFAULT_OFFSETS["actual_value"],
            ["actualvalue", "actualsales", "reportedsales"],
        ),
        "forecast_max": choose_column(
            sheet,
            header_cells,
            anchor_col,
            first_data_row,
            last_data_row,
            REGRESSION_DEFAULT_OFFSETS["forecast_max"],
            ["max", "forecastmax"],
        ),
        "forecast_min": choose_column(
            sheet,
            header_cells,
            anchor_col,
            first_data_row,
            last_data_row,
            REGRESSION_DEFAULT_OFFSETS["forecast_min"],
            ["min", "forecastmin"],
        ),
    }

    paired_rows = collect_numeric_rows_pair(sheet, x_col, y_col, anchor_row - 1)

    used_last_col = sheet.used_range.last_cell.column
    scratch_intercept_col = max(used_last_col + 2, anchor_col + 2)
    scratch_slope_col = scratch_intercept_col + 1

    formulas_written = False
    for idx in range(N_QUARTERS):
        n_use = min(idx + 1, len(paired_rows))
        target_row = first_data_row + idx

        if n_use >= 2:
            first_hist_row = paired_rows[-n_use]
            last_hist_row = paired_rows[-1]
            intercept_formula = (
                f'=IFERROR(INTERCEPT(R{first_hist_row}C{y_col}:R{last_hist_row}C{y_col},'
                f'R{first_hist_row}C{x_col}:R{last_hist_row}C{x_col}),"")'
            )
            slope_formula = (
                f'=IFERROR(SLOPE(R{first_hist_row}C{y_col}:R{last_hist_row}C{y_col},'
                f'R{first_hist_row}C{x_col}:R{last_hist_row}C{x_col}),"")'
            )
        else:
            intercept_formula = '=""'
            slope_formula = '=""'

        set_r1c1_formula2(sheet.range((target_row, scratch_intercept_col)), intercept_formula)
        set_r1c1_formula2(sheet.range((target_row, scratch_slope_col)), slope_formula)
        formulas_written = True

    if formulas_written:
        recalculate(workbook)

    tracked_cols = [
        col_map["num_quarters_used"],
        col_map["forecast_value"],
        col_map["actual_value"],
        col_map["forecast_max"],
        col_map["forecast_min"],
        scratch_intercept_col,
        scratch_slope_col,
    ]
    min_col = min(tracked_cols)
    max_col = max(tracked_cols)
    block = read_block(sheet, first_data_row, last_data_row, min_col, max_col)

    rows: list[dict[str, Any]] = []
    previous_signature: tuple[Any, ...] | None = None
    for idx in range(N_QUARTERS):
        forecast_max = block_value(block, idx, col_map["forecast_max"], min_col)
        forecast_min = block_value(block, idx, col_map["forecast_min"], min_col)
        forecast_value = block_value(block, idx, col_map["forecast_value"], min_col)
        intercept = block_value(block, idx, scratch_intercept_col, min_col)
        slope = block_value(block, idx, scratch_slope_col, min_col)

        if (
            is_blank(forecast_max)
            and is_blank(forecast_min)
            and is_blank(forecast_value)
            and is_blank(intercept)
            and is_blank(slope)
        ):
            continue

        signature = (
            value_signature(intercept),
            value_signature(slope),
            value_signature(forecast_value),
            value_signature(forecast_max),
            value_signature(forecast_min),
        )
        if previous_signature is not None and signature == previous_signature:
            continue
        previous_signature = signature

        raw_num_quarters = block_value(block, idx, col_map["num_quarters_used"], min_col)
        num_quarters = int(round(to_float(raw_num_quarters))) if to_float(raw_num_quarters) is not None else idx + 1

        rows.append(
            {
                "model": metadata.model,
                "ticker": metadata.ticker,
                "model_period": metadata.model_period,
                "model_date": metadata.model_date,
                "method": "regression",
                "parameter_name": "num_quarters_used",
                "parameter_value": num_quarters,
                "num_quarters_used": num_quarters,
                "forecast_value": forecast_value,  # TOT FCST w/o SA
                "actual_value": block_value(block, idx, col_map["actual_value"], min_col),
                "forecast_max": forecast_max,
                "forecast_min": forecast_min,
                "range_width": numeric_difference(forecast_max, forecast_min),
                "intercept": intercept,
                "slope": slope,
                "source_file": source_file,
            }
        )

    if formulas_written:
        sheet.range((first_data_row, scratch_intercept_col), (last_data_row, scratch_slope_col)).clear_contents()

    return rows


def style_output_sheet(ws: Any, headers: Sequence[str]) -> None:
    for cell in ws[1]:
        cell.font = Font(bold=True)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    for col_idx, header in enumerate(headers, start=1):
        max_len = len(str(header))
        for row_idx in range(2, ws.max_row + 1):
            value = ws.cell(row=row_idx, column=col_idx).value
            if value is None:
                continue
            max_len = max(max_len, len(str(value)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(12, max_len + 2), 64)


def write_sheet(ws: Any, headers: Sequence[str], rows: Iterable[dict[str, Any]]) -> None:
    ws.append(list(headers))
    for row in rows:
        ws.append([row.get(header, None) for header in headers])
    style_output_sheet(ws, headers)


def write_output_workbook(
    output_file: Path,
    empirical_rows: Sequence[dict[str, Any]],
    regression_rows: Sequence[dict[str, Any]],
) -> None:
    wb = Workbook()
    default_sheet = wb.active
    wb.remove(default_sheet)

    ws_empirical = wb.create_sheet("empirical_candidates")
    write_sheet(ws_empirical, EMPIRICAL_COLUMNS, empirical_rows)

    ws_regression = wb.create_sheet("regression_candidates")
    write_sheet(ws_regression, REGRESSION_COLUMNS, regression_rows)

    wb.save(output_file)


def main() -> None:
    input_path = Path(input_dir).expanduser().resolve()
    output_path = Path(output_dir).expanduser().resolve()

    if not input_path.exists():
        raise FileNotFoundError(f"Input directory does not exist: {input_path}")
    if not input_path.is_dir():
        raise NotADirectoryError(f"Input path is not a directory: {input_path}")

    output_path.mkdir(parents=True, exist_ok=True)
    output_file = choose_output_path(input_path, output_path)

    empirical_rows: list[dict[str, Any]] = []
    regression_rows: list[dict[str, Any]] = []
    processed_files = 0

    app = xw.App(visible=False, add_book=False)
    try:
        for attr, value in (
            ("display_alerts", False),
            ("screen_updating", False),
            ("enable_events", False),
        ):
            try:
                setattr(app, attr, value)
            except Exception:
                pass
        try:
            app.calculation = "manual"
        except Exception:
            pass

        for file_path in sorted(input_path.iterdir(), key=lambda p: p.name.lower()):
            skip_reason = should_skip_file(file_path, input_path.name)
            if skip_reason:
                print(f"skipped: {file_path.name} ({skip_reason})")
                continue

            try:
                metadata = parse_file_metadata(file_path)
            except Exception as exc:
                print(f"skipped: {file_path.name} (filename parse error: {exc})")
                continue

            print(f"processing: {file_path.name}")
            workbook: xw.Book | None = None
            try:
                workbook = app.books.open(str(file_path), update_links=False)
                empirical_rows.extend(extract_empirical_candidates(workbook, metadata, file_path.name))
                regression_rows.extend(extract_regression_candidates(workbook, metadata, file_path.name))
                processed_files += 1
                print(f"processed: {file_path.name}")
            except Exception as exc:
                print(f"skipped: {file_path.name} (processing error: {exc})")
            finally:
                close_source_workbook(workbook)
    finally:
        app.quit()

    write_output_workbook(output_file, empirical_rows, regression_rows)

    print(f"output_path: {output_file}")
    print(f"number_of_files_processed: {processed_files}")
    print(f"number_of_empirical_rows: {len(empirical_rows)}")
    print(f"number_of_regression_rows: {len(regression_rows)}")


if __name__ == "__main__":
    main()
