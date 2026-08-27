from __future__ import annotations

import math
import re
from dataclasses import dataclass
from datetime import date
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Sequence, Tuple

import xlwings as xw
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

# =========================
# User-configurable inputs
# =========================
input_dir = Path(r"./input")
output_dir = Path(r"./output")


EMPIRICAL_COLUMNS = [
    "model",
    "ticker",
    "model_period",
    "model_date",
    "method",
    "parameter_name",
    "parameter_value",
    "num_quarters_used",
    "last_quarter_used",
    "forecast_value",
    "actual_value",
    "forecast_max",
    "forecast_min",
    "range_width",
    "avg_penetration_pct",
    "quarterly_sales",
    "reported_sales",
    "growth_rate_pct",
    "sales_captured_in_db_pct",
    "source_file",
]

REGRESSION_COLUMNS = [
    "model",
    "ticker",
    "model_period",
    "model_date",
    "method",
    "parameter_name",
    "parameter_value",
    "num_quarters_used",
    "forecast_value",
    "actual_value",
    "forecast_max",
    "forecast_min",
    "range_width",
    "intercept",
    "slope",
    "source_file",
]

DAY_BY_WINDOW = {"early": 5, "mid": 15, "late": 25}
MONTH_BY_ABBR = {
    "jan": 1,
    "feb": 2,
    "mar": 3,
    "apr": 4,
    "may": 5,
    "jun": 6,
    "jul": 7,
    "aug": 8,
    "sep": 9,
    "oct": 10,
    "nov": 11,
    "dec": 12,
}

EMPIRICAL_ALIASES = {
    "num_quarters_used": ["num quarters used", "number quarters used", "quarters used", "num qtrs", "n qtrs"],
    "last_quarter_used": ["last quarter used", "last qtr used", "last quarter", "last qtr"],
    "forecast_value": ["estimated total sold", "est total sold", "total sold estimate", "forecast value", "tot fcst"],
    "actual_value": ["reported sales", "actual sales", "reported value"],
    "forecast_min": ["min", "minimum"],
    "avg_penetration_pct": ["avg penetration", "average penetration", "penetration pct", "penetration %"],
    "quarterly_sales": ["quarterly sales", "qtrly sales", "quarter sales"],
    "reported_sales": ["reported sales", "reported total sales"],
    "growth_rate_pct": ["growth rate", "growth pct", "growth %"],
    "sales_captured_in_db_pct": ["sales captured in db", "captured in db", "captured %", "captured pct"],
}

REGRESSION_ALIASES = {
    "num_quarters_used": ["num quarters used", "number quarters used", "quarters used", "num qtrs", "n qtrs"],
    "forecast_value": ["tot fcst w/o sa", "total forecast without sa", "forecast value", "forecast total"],
    "actual_value": ["reported sales", "actual sales"],
    "forecast_min": ["min", "minimum"],
    "intercept": ["intercept"],
    "slope": ["slope"],
}


@dataclass(frozen=True)
class FileMeta:
    model: str
    ticker: str
    model_period: str
    model_date: str


@dataclass
class SheetSnapshot:
    values: List[List[Any]]
    top_row: int
    left_col: int
    max_row: int
    max_col: int

    def value_at(self, row: int, col: int) -> Any:
        r_idx = row - self.top_row
        c_idx = col - self.left_col
        if r_idx < 0 or c_idx < 0:
            return None
        if r_idx >= len(self.values):
            return None
        row_data = self.values[r_idx]
        if c_idx >= len(row_data):
            return None
        return row_data[c_idx]


def normalize_text(value: Any) -> str:
    if value is None:
        return ""
    return re.sub(r"[^a-z0-9]+", " ", str(value).strip().lower()).strip()


def as_number(value: Any) -> Optional[float]:
    if value is None:
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        if isinstance(value, float) and math.isnan(value):
            return None
        return float(value)
    if isinstance(value, str):
        stripped = value.strip()
        if stripped == "":
            return None
        stripped = stripped.replace(",", "")
        pct = stripped.endswith("%")
        if pct:
            stripped = stripped[:-1].strip()
        try:
            number = float(stripped)
            return number / 100.0 if pct else number
        except ValueError:
            return None
    return None


def safe_value(value: Any) -> Any:
    if isinstance(value, float) and math.isnan(value):
        return None
    return value


def to_matrix(values: Any) -> List[List[Any]]:
    if values is None:
        return []
    if not isinstance(values, list):
        return [[values]]
    if len(values) == 0:
        return []
    if isinstance(values[0], list):
        return values
    return [values]


def build_snapshot(sheet: xw.Sheet) -> SheetSnapshot:
    used = sheet.used_range
    matrix = to_matrix(used.value)
    row_count = len(matrix)
    col_count = max((len(row) for row in matrix), default=0)
    return SheetSnapshot(
        values=matrix,
        top_row=used.row,
        left_col=used.column,
        max_row=used.row + max(row_count - 1, 0),
        max_col=used.column + max(col_count - 1, 0),
    )


def find_max_anchor(snapshot: SheetSnapshot) -> Optional[Tuple[int, int]]:
    candidates: List[Tuple[int, int, int]] = []
    for r_idx, row in enumerate(snapshot.values):
        for c_idx, cell_value in enumerate(row):
            if normalize_text(cell_value) != "max":
                continue
            row_abs = snapshot.top_row + r_idx
            col_abs = snapshot.left_col + c_idx
            score = 0
            right = normalize_text(snapshot.value_at(row_abs, col_abs + 1))
            if right == "min":
                score += 10
            left = normalize_text(snapshot.value_at(row_abs, col_abs - 1))
            if "forecast" in left:
                score += 2
            candidates.append((score, row_abs, col_abs))
    if not candidates:
        return None
    candidates.sort(reverse=True)
    _, row_abs, col_abs = candidates[0]
    return row_abs, col_abs


def gather_text_cells(snapshot: SheetSnapshot) -> List[Tuple[int, int, str]]:
    text_cells: List[Tuple[int, int, str]] = []
    for r_idx, row in enumerate(snapshot.values):
        row_abs = snapshot.top_row + r_idx
        for c_idx, value in enumerate(row):
            normalized = normalize_text(value)
            if normalized:
                text_cells.append((row_abs, snapshot.left_col + c_idx, normalized))
    return text_cells


def find_label_col_offsets(
    snapshot: SheetSnapshot,
    anchor_row: int,
    anchor_col: int,
    aliases: Dict[str, Sequence[str]],
) -> Dict[str, int]:
    text_cells = gather_text_cells(snapshot)
    offsets: Dict[str, int] = {}
    for key, options in aliases.items():
        best: Optional[Tuple[int, int]] = None
        for row, col, norm in text_cells:
            if not any(option in norm for option in options):
                continue
            score = abs(row - anchor_row) * 100 + abs(col - anchor_col)
            if best is None or score < best[0]:
                best = (score, col)
        if best is not None:
            offsets[key] = best[1] - anchor_col
    return offsets


def numeric_rows_in_col(snapshot: SheetSnapshot, col: int, row_start: int, row_end: int) -> List[int]:
    rows: List[int] = []
    for row in range(row_start, row_end + 1):
        if as_number(snapshot.value_at(row, col)) is not None:
            rows.append(row)
    return rows


def read_cell_by_offset(sheet: xw.Sheet, row: int, anchor_col: int, offset: Optional[int]) -> Any:
    if offset is None:
        return None
    col = anchor_col + offset
    if col < 1 or row < 1:
        return None
    return safe_value(sheet.range((row, col)).value)


def apply_formula2_r1c1(cell: xw.Range, formula_r1c1: str) -> None:
    try:
        cell.formula2 = formula_r1c1
    except Exception:
        # Some Excel versions expose formula2 with partial support.
        cell.formula = formula_r1c1


def calc_if_needed(wb: xw.Book, formula_updated: bool) -> None:
    if formula_updated:
        wb.app.calculate()


def parse_file_meta(file_name: str) -> Optional[FileMeta]:
    stem = Path(file_name).stem
    parts = [part.strip() for part in stem.split(" - ")]
    if len(parts) < 3:
        return None
    ticker = parts[1]
    period_token = parts[2].split("_")[0]

    match = re.fullmatch(r"(Early|Mid|Late)([A-Za-z]{3})(\d{4})", period_token, flags=re.IGNORECASE)
    if not match:
        return None

    window_raw, month_raw, year_raw = match.groups()
    window_key = window_raw.lower()
    month_key = month_raw.lower()
    if window_key not in DAY_BY_WINDOW or month_key not in MONTH_BY_ABBR:
        return None

    year_num = int(year_raw)
    month_num = MONTH_BY_ABBR[month_key]
    day_num = DAY_BY_WINDOW[window_key]
    month_abbr = month_raw[:1].upper() + month_raw[1:].lower()
    window_name = window_raw[:1].upper() + window_raw[1:].lower()
    model_period = f"{window_name}{month_abbr}_{year_num}"
    model_date = date(year_num, month_num, day_num).isoformat()
    model = f"{ticker}_{model_period}"
    return FileMeta(model=model, ticker=ticker, model_period=model_period, model_date=model_date)


def close_workbook_no_save(wb: xw.Book) -> None:
    try:
        wb.close(save=False)
        return
    except TypeError:
        pass
    except Exception:
        pass

    try:
        wb.close(False)
        return
    except Exception:
        pass

    try:
        wb.api.Close(SaveChanges=False)
        return
    except Exception as exc:
        raise RuntimeError("Unable to close source workbook without saving.") from exc


def find_sheet_or_none(wb: xw.Book, sheet_name: str) -> Optional[xw.Sheet]:
    for sheet in wb.sheets:
        if sheet.name.strip().lower() == sheet_name.strip().lower():
            return sheet
    return None


def choose_empirical_penetration_col(
    snapshot: SheetSnapshot,
    anchor_row: int,
    anchor_col: int,
    offsets: Dict[str, int],
) -> Optional[int]:
    preferred_offset = offsets.get("avg_penetration_pct")
    if preferred_offset is not None:
        col = anchor_col + preferred_offset
        rows = numeric_rows_in_col(snapshot, col, snapshot.top_row, max(anchor_row - 1, snapshot.top_row))
        if rows:
            return col

    text_cells = gather_text_cells(snapshot)
    candidates: List[Tuple[int, int]] = []
    for row, col, norm in text_cells:
        if "penetration" in norm:
            numeric_rows = numeric_rows_in_col(snapshot, col, row + 1, max(anchor_row - 1, row + 1))
            if len(numeric_rows) >= 2:
                score = abs(row - anchor_row) * 100 + abs(col - anchor_col)
                candidates.append((score, col))
    if not candidates:
        return None
    candidates.sort()
    return candidates[0][1]


def extract_empirical_candidates(
    wb: xw.Book,
    sheet: xw.Sheet,
    file_meta: FileMeta,
    source_file: str,
) -> List[Dict[str, Any]]:
    snapshot = build_snapshot(sheet)
    anchor = find_max_anchor(snapshot)
    if anchor is None:
        return []
    anchor_row, anchor_col = anchor

    offsets = find_label_col_offsets(snapshot, anchor_row, anchor_col, EMPIRICAL_ALIASES)
    offsets.setdefault("forecast_max", 0)
    offsets.setdefault("forecast_min", 1)

    penetration_col = choose_empirical_penetration_col(snapshot, anchor_row, anchor_col, offsets)
    helper_col = anchor_col + 30
    helper_row = anchor_row - 1 if anchor_row > 1 else anchor_row + 12
    helper_cell = sheet.range((helper_row, helper_col))

    penetration_rows: List[int] = []
    if penetration_col is not None:
        penetration_rows = numeric_rows_in_col(snapshot, penetration_col, snapshot.top_row, max(anchor_row - 1, snapshot.top_row))

    rows_out: List[Dict[str, Any]] = []
    for n_quarters in range(1, 11):
        formula_updated = False
        avg_penetration_value = None
        if penetration_col is not None and len(penetration_rows) >= n_quarters:
            end_row = penetration_rows[-1]
            start_row = penetration_rows[-n_quarters]
            formula = (
                f'=IFERROR(AVERAGE(R{start_row}C{penetration_col}:R{end_row}C{penetration_col}),"")'
            )
            apply_formula2_r1c1(helper_cell, formula)
            formula_updated = True
            calc_if_needed(wb, formula_updated)
            avg_penetration_value = safe_value(helper_cell.value)

        row_idx = anchor_row + n_quarters
        num_quarters_used = read_cell_by_offset(sheet, row_idx, anchor_col, offsets.get("num_quarters_used"))
        if num_quarters_used is None:
            num_quarters_used = n_quarters

        last_quarter_used = read_cell_by_offset(sheet, row_idx, anchor_col, offsets.get("last_quarter_used"))
        forecast_value = read_cell_by_offset(sheet, row_idx, anchor_col, offsets.get("forecast_value"))
        actual_value = read_cell_by_offset(sheet, row_idx, anchor_col, offsets.get("actual_value"))
        forecast_max = read_cell_by_offset(sheet, row_idx, anchor_col, offsets.get("forecast_max"))
        forecast_min = read_cell_by_offset(sheet, row_idx, anchor_col, offsets.get("forecast_min"))
        quarterly_sales = read_cell_by_offset(sheet, row_idx, anchor_col, offsets.get("quarterly_sales"))
        reported_sales = read_cell_by_offset(sheet, row_idx, anchor_col, offsets.get("reported_sales"))
        growth_rate_pct = read_cell_by_offset(sheet, row_idx, anchor_col, offsets.get("growth_rate_pct"))
        sales_captured_in_db_pct = read_cell_by_offset(sheet, row_idx, anchor_col, offsets.get("sales_captured_in_db_pct"))

        avg_penetration_cell = read_cell_by_offset(sheet, row_idx, anchor_col, offsets.get("avg_penetration_pct"))
        if avg_penetration_cell is not None:
            avg_penetration_value = avg_penetration_cell

        if reported_sales is None:
            reported_sales = actual_value

        range_width = None
        max_num = as_number(forecast_max)
        min_num = as_number(forecast_min)
        if max_num is not None and min_num is not None:
            range_width = max_num - min_num

        # Skip fully empty candidate rows.
        data_check = [
            forecast_value,
            actual_value,
            forecast_max,
            forecast_min,
            avg_penetration_value,
            quarterly_sales,
            reported_sales,
        ]
        if all(value in (None, "") for value in data_check):
            continue

        rows_out.append(
            {
                "model": file_meta.model,
                "ticker": file_meta.ticker,
                "model_period": file_meta.model_period,
                "model_date": file_meta.model_date,
                "method": "empirical",
                "parameter_name": "avg_penetration_pct",
                "parameter_value": avg_penetration_value,
                "num_quarters_used": num_quarters_used,
                "last_quarter_used": last_quarter_used,
                "forecast_value": forecast_value,
                "actual_value": actual_value,
                "forecast_max": forecast_max,
                "forecast_min": forecast_min,
                "range_width": range_width,
                "avg_penetration_pct": avg_penetration_value,
                "quarterly_sales": quarterly_sales,
                "reported_sales": reported_sales,
                "growth_rate_pct": growth_rate_pct,
                "sales_captured_in_db_pct": sales_captured_in_db_pct,
                "source_file": source_file,
            }
        )
    return rows_out


def collect_numeric_xy_rows(
    snapshot: SheetSnapshot,
    x_col: int,
    y_col: int,
    row_start: int,
    row_end: int,
) -> List[Tuple[int, float, float]]:
    rows: List[Tuple[int, float, float]] = []
    for row in range(row_start, row_end + 1):
        x_value = as_number(snapshot.value_at(row, x_col))
        y_value = as_number(snapshot.value_at(row, y_col))
        if x_value is None or y_value is None:
            continue
        rows.append((row, x_value, y_value))
    return rows


def row_signature(values: Iterable[Any]) -> Tuple[Any, ...]:
    signature: List[Any] = []
    for value in values:
        num = as_number(value)
        if num is not None:
            signature.append(round(num, 10))
        else:
            signature.append(value)
    return tuple(signature)


def extract_regression_candidates(
    wb: xw.Book,
    sheet: xw.Sheet,
    file_meta: FileMeta,
    source_file: str,
) -> List[Dict[str, Any]]:
    snapshot = build_snapshot(sheet)
    anchor = find_max_anchor(snapshot)
    if anchor is None:
        return []
    anchor_row, anchor_col = anchor

    offsets = find_label_col_offsets(snapshot, anchor_row, anchor_col, REGRESSION_ALIASES)
    offsets.setdefault("forecast_max", 0)
    offsets.setdefault("forecast_min", 1)

    x_col = anchor_col - 11
    y_col = anchor_col - 7
    numeric_xy = collect_numeric_xy_rows(snapshot, x_col, y_col, snapshot.top_row, max(anchor_row - 1, snapshot.top_row))
    if len(numeric_xy) < 2:
        return []

    helper_col = anchor_col + 30
    helper_row = anchor_row - 1 if anchor_row > 1 else anchor_row + 12
    intercept_cell = sheet.range((helper_row, helper_col))
    slope_cell = sheet.range((helper_row + 1, helper_col))

    rows_out: List[Dict[str, Any]] = []
    previous_signature: Optional[Tuple[Any, ...]] = None

    for n_quarters in range(2, 11):
        if len(numeric_xy) < n_quarters:
            break
        subset = numeric_xy[-n_quarters:]
        start_row = subset[0][0]
        end_row = subset[-1][0]

        intercept_formula = (
            f'=IFERROR(INTERCEPT(R{start_row}C{y_col}:R{end_row}C{y_col},'
            f'R{start_row}C{x_col}:R{end_row}C{x_col}),"")'
        )
        slope_formula = (
            f'=IFERROR(SLOPE(R{start_row}C{y_col}:R{end_row}C{y_col},'
            f'R{start_row}C{x_col}:R{end_row}C{x_col}),"")'
        )
        apply_formula2_r1c1(intercept_cell, intercept_formula)
        apply_formula2_r1c1(slope_cell, slope_formula)
        calc_if_needed(wb, formula_updated=True)

        intercept_value = safe_value(intercept_cell.value)
        slope_value = safe_value(slope_cell.value)

        last_x = subset[-1][1]
        intercept_num = as_number(intercept_value)
        slope_num = as_number(slope_value)
        forecast_default = None
        if intercept_num is not None and slope_num is not None:
            forecast_default = intercept_num + slope_num * (last_x + 1)

        row_idx = anchor_row + n_quarters
        num_quarters_used = read_cell_by_offset(sheet, row_idx, anchor_col, offsets.get("num_quarters_used"))
        if num_quarters_used is None:
            num_quarters_used = n_quarters

        forecast_value = read_cell_by_offset(sheet, row_idx, anchor_col, offsets.get("forecast_value"))
        if forecast_value is None:
            forecast_value = forecast_default

        actual_value = read_cell_by_offset(sheet, row_idx, anchor_col, offsets.get("actual_value"))
        forecast_max = read_cell_by_offset(sheet, row_idx, anchor_col, offsets.get("forecast_max"))
        forecast_min = read_cell_by_offset(sheet, row_idx, anchor_col, offsets.get("forecast_min"))

        intercept_cell_value = read_cell_by_offset(sheet, row_idx, anchor_col, offsets.get("intercept"))
        if intercept_cell_value is not None:
            intercept_value = intercept_cell_value
        slope_cell_value = read_cell_by_offset(sheet, row_idx, anchor_col, offsets.get("slope"))
        if slope_cell_value is not None:
            slope_value = slope_cell_value

        range_width = None
        max_num = as_number(forecast_max)
        min_num = as_number(forecast_min)
        if max_num is not None and min_num is not None:
            range_width = max_num - min_num

        signature = row_signature(
            [num_quarters_used, forecast_value, actual_value, forecast_max, forecast_min, intercept_value, slope_value]
        )
        if previous_signature is not None and signature == previous_signature:
            continue
        previous_signature = signature

        data_check = [forecast_value, forecast_max, forecast_min, intercept_value, slope_value]
        if all(value in (None, "") for value in data_check):
            continue

        rows_out.append(
            {
                "model": file_meta.model,
                "ticker": file_meta.ticker,
                "model_period": file_meta.model_period,
                "model_date": file_meta.model_date,
                "method": "regression",
                "parameter_name": "num_quarters_used",
                "parameter_value": num_quarters_used,
                "num_quarters_used": num_quarters_used,
                "forecast_value": forecast_value,
                "actual_value": actual_value,
                "forecast_max": forecast_max,
                "forecast_min": forecast_min,
                "range_width": range_width,
                "intercept": intercept_value,
                "slope": slope_value,
                "source_file": source_file,
            }
        )
    return rows_out


def next_output_path(input_folder: Path, out_folder: Path) -> Path:
    base = f"{input_folder.name}_PARAM"
    candidate = out_folder / f"{base}.xlsx"
    suffix = 1
    while candidate.exists():
        candidate = out_folder / f"{base}.{suffix}.xlsx"
        suffix += 1
    return candidate


def write_rows(sheet, columns: Sequence[str], rows: Sequence[Dict[str, Any]]) -> None:
    sheet.append(list(columns))
    for row in rows:
        sheet.append([row.get(col) for col in columns])

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for cell in sheet[1]:
        cell.font = Font(bold=True)

    for idx, col_name in enumerate(columns, start=1):
        max_len = len(col_name)
        for row_idx in range(2, sheet.max_row + 1):
            value = sheet.cell(row=row_idx, column=idx).value
            display = "" if value is None else str(value)
            if len(display) > max_len:
                max_len = len(display)
        sheet.column_dimensions[get_column_letter(idx)].width = min(max(max_len + 2, 10), 55)


def save_output_workbook(
    output_path: Path,
    empirical_rows: Sequence[Dict[str, Any]],
    regression_rows: Sequence[Dict[str, Any]],
) -> None:
    wb_out = Workbook()
    ws_empirical = wb_out.active
    ws_empirical.title = "empirical_candidates"
    write_rows(ws_empirical, EMPIRICAL_COLUMNS, empirical_rows)

    ws_regression = wb_out.create_sheet("regression_candidates")
    write_rows(ws_regression, REGRESSION_COLUMNS, regression_rows)
    wb_out.save(output_path)


def configure_excel_app(app: xw.App) -> None:
    app.display_alerts = False
    app.screen_updating = False
    try:
        app.calculation = "manual"
    except Exception:
        # Keep default if manual mode is unavailable.
        pass


def iter_candidate_files(folder: Path) -> Iterable[Path]:
    for path in sorted(folder.iterdir(), key=lambda p: p.name.lower()):
        if not path.is_file():
            continue
        yield path


def main() -> None:
    if not input_dir.exists():
        print(f"Skipped input folder: {input_dir} (folder not found)")
        return

    output_dir.mkdir(parents=True, exist_ok=True)
    output_path = next_output_path(input_dir, output_dir)

    empirical_rows: List[Dict[str, Any]] = []
    regression_rows: List[Dict[str, Any]] = []
    files_processed = 0

    app = xw.App(visible=False, add_book=False)
    configure_excel_app(app)

    try:
        for file_path in iter_candidate_files(input_dir):
            file_name = file_path.name
            if file_name.startswith("~"):
                print(f"Skipped file: {file_name} (temporary file)")
                continue
            if file_path.suffix.lower() != ".xlsx":
                print(f"Skipped file: {file_name} (not an .xlsx file)")
                continue

            meta = parse_file_meta(file_name)
            if meta is None:
                print(f"Skipped file: {file_name} (filename did not match expected model pattern)")
                continue

            wb_in: Optional[xw.Book] = None
            try:
                print(f"Processed file: {file_name}")
                wb_in = app.books.open(str(file_path), update_links=False)
                empirical_sheet = find_sheet_or_none(wb_in, "Empirical Model")
                regression_sheet = find_sheet_or_none(wb_in, "Regression Model")

                if empirical_sheet is None:
                    print(f"Skipped empirical sheet in {file_name}: sheet not found")
                else:
                    empirical_rows.extend(
                        extract_empirical_candidates(
                            wb=wb_in,
                            sheet=empirical_sheet,
                            file_meta=meta,
                            source_file=file_name,
                        )
                    )

                if regression_sheet is None:
                    print(f"Skipped regression sheet in {file_name}: sheet not found")
                else:
                    regression_rows.extend(
                        extract_regression_candidates(
                            wb=wb_in,
                            sheet=regression_sheet,
                            file_meta=meta,
                            source_file=file_name,
                        )
                    )

                files_processed += 1
            except Exception as exc:
                print(f"Skipped file: {file_name} (error: {exc})")
            finally:
                if wb_in is not None:
                    close_workbook_no_save(wb_in)
    finally:
        app.quit()

    save_output_workbook(output_path, empirical_rows, regression_rows)

    print(f"Output path: {output_path}")
    print(f"Number of files processed: {files_processed}")
    print(f"Number of empirical rows: {len(empirical_rows)}")
    print(f"Number of regression rows: {len(regression_rows)}")


if __name__ == "__main__":
    main()
